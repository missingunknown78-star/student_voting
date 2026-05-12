from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, session, jsonify, Response
from extensions import db, bcrypt
from flask_login import login_user, logout_user, current_user, login_required
from datetime import datetime
from functools import wraps
from admin.models import Admin, Candidate, Position, Election, Announcement, Department, Course, CtuStudent, TallyVote, ElectionPosition
from student.models import Student, Vote
import mysql.connector
from settings import DATABASE_URL, MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DB, MYSQL_PORT
import pytz
from werkzeug.utils import secure_filename
import os
import logging
import time
import pyotp
from datetime import timedelta
from sqlalchemy import or_
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash
import json
from phe import paillier
import pickle
from .utils import check_if_tallied, get_tally_timestamp
from flask import request, render_template, send_file
from sqlalchemy import or_, desc
from datetime import datetime
import io
import csv
from admin.models import AuditLog, Setting
from admin.utils import log_audit
import pandas as pd
from admin.models import AdminTrustedDevice, TwoFADisableToken
from flask import make_response
from datetime import datetime
import pytz
from admin.models import YearLevel, DeletionRequestAudit
from student.models import PendingCandidate
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import traceback
from datetime import datetime, timedelta
import secrets
import datetime
from flask import request, jsonify, render_template_string
from flask_mail import Mail, Message
from extensions import mail

try:
    from admin.models import TallyVote
    TALLY_VOTE_AVAILABLE = True
except ImportError:
    TALLY_VOTE_AVAILABLE = False
    print("Note: TallyVote model not found. Official tally features disabled.")



# ---------------------- Blueprint ---------------------- #
admin_bp = Blueprint('admin', __name__, template_folder='templates', static_folder='static')


# ---------------------- PHE Helper Functions ---------------------- #
def get_encryption_keys():
    """Get Paillier encryption keys"""
    key_dir = "keys"
    os.makedirs(key_dir, exist_ok=True)
    
    public_key_path = os.path.join(key_dir, "public_key.pkl")
    private_key_path = os.path.join(key_dir, "private_key.pkl")
    
    if os.path.exists(public_key_path) and os.path.exists(private_key_path):
        with open(public_key_path, 'rb') as f:
            public_key = pickle.load(f)
        with open(private_key_path, 'rb') as f:
            private_key = pickle.load(f)
    else:
        public_key, private_key = paillier.generate_paillier_keypair()
        with open(public_key_path, 'wb') as f:
            pickle.dump(public_key, f)
        with open(private_key_path, 'wb') as f:
            pickle.dump(private_key, f)
    
    return public_key, private_key

# Initialize keys
public_key, private_key = get_encryption_keys()

def deserialize_encrypted_vote(encrypted_data):
    """Deserialize stored encrypted vote"""
    if not encrypted_data:
        return []
    
    data = json.loads(encrypted_data)
    return [
        paillier.EncryptedNumber(
            public_key,
            int(item["ciphertext"]),
            int(item["exponent"])
        )
        for item in data
    ]

def count_votes_for_candidate(candidate_id, election_id):
    """Count votes for a specific candidate using encrypted votes (OFFICIAL TALLY)"""
    import json
    
    votes = Vote.query.filter_by(election_id=election_id).all()
    
    if not votes:
        return 0
    
    # Get current candidates sorted by ID (for reference only)
    all_candidates = Candidate.query.filter_by(election_id=election_id)\
                                    .order_by(Candidate.id).all()
    current_sorted_ids = [c.id for c in all_candidates]
    
    total_votes = 0
    
    for vote in votes:
        if not vote.finder_hash:
            continue
            
        try:
            finder_data = json.loads(vote.finder_hash)
            
            # Get the candidate order that was used when this vote was created
            stored_order = finder_data.get('candidate_order')
            
            if not stored_order:
                # For old votes without candidate_order, use current order
                # (This will be wrong for old votes but we fixed them)
                stored_order = current_sorted_ids
            
            # Find which index this candidate is in the stored order
            try:
                candidate_index = stored_order.index(candidate_id)
            except ValueError:
                # Candidate not in this vote's candidate list (shouldn't happen)
                continue
            
            # Decrypt the vote at that index
            enc_votes = deserialize_encrypted_vote(vote.encrypted_vote)
            
            if candidate_index < len(enc_votes):
                decrypted = private_key.decrypt(enc_votes[candidate_index])
                if decrypted > 0:  # This candidate was voted for
                    total_votes += 1
                    
        except Exception as e:
            print(f"Error processing vote {vote.id}: {e}")
            continue
    
    return total_votes

def get_all_voters_for_election(election_id):
    """Get all unique voters for an election"""
    votes = Vote.query.filter_by(election_id=election_id).all()
    return list(set(vote.student_id for vote in votes))

# ---------------------- Secure Admin Required Decorator ---------------------- #
logging.basicConfig(filename='admin_access.log', level=logging.WARNING,
                    format='%(asctime)s - %(message)s')

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        ip = request.remote_addr
        attempts = session.get(f'admin_attempts_{ip}', 0)
        cooldown = session.get(f'admin_cooldown_{ip}', 0)

        if time.time() < cooldown:
            flash(f"Too many failed attempts. Try again in {int(cooldown - time.time())} seconds.", "admin-warning")
            return redirect(url_for('admin.login'))

        if not current_user.is_authenticated or getattr(current_user, 'role', None) != 'Admin':
            logging.warning(f"Unauthorized admin access attempt from IP: {ip}")

            attempts += 1
            session[f'admin_attempts_{ip}'] = attempts

            if attempts >= MAX_ATTEMPTS:
                session[f'admin_cooldown_{ip}'] = time.time() + COOLDOWN_TIME
                session[f'admin_attempts_{ip}'] = 0
                flash("Too many failed attempts. Admin login temporarily locked.", "admin-warning")
            else:
                flash("Please log in as admin to access this page.", "admin-warning")

            return redirect(url_for('admin.login'))

        session[f'admin_attempts_{ip}'] = 0
        session[f'admin_cooldown_{ip}'] = 0

        return f(*args, **kwargs)
    return decorated_function

 
def count_unique_voters(election_id):
    """Count the number of unique students who voted in an election"""
    return db.session.query(Vote.student_id).filter_by(
        election_id=election_id
    ).distinct().count()


def get_all_years():
    """Get all unique years from elections"""
    try:
        elections = Election.query.order_by(Election.start_date.asc()).all()
        years = set()
        
        for election in elections:
            if election.start_date:
                years.add(election.start_date.year)
        
        return sorted(list(years), reverse=True)
    except:
        return []

def get_position_color(position_name, index=None):
    """
    Generate consistent colors for positions based on name hash.
    This ensures the same position always gets the same color.
    """
    import hashlib
    hash_obj = hashlib.md5(position_name.encode())
    hash_hex = hash_obj.hexdigest()
    # Use first 6 characters for a consistent color
    return f'#{hash_hex[:6]}'


def calculate_chart_step_size(max_value):
    """Calculate appropriate step size for chart y-axis"""
    if max_value <= 10:
        return 2
    elif max_value <= 20:
        return 5
    elif max_value <= 50:
        return 10
    elif max_value <= 100:
        return 20
    elif max_value <= 200:
        return 50
    elif max_value <= 500:
        return 100
    elif max_value <= 1000:
        return 200
    elif max_value <= 5000:
        return 500
    else:
        return 1000


def get_candidate_color(candidate_name, position_name, index, total):
    """
    Generate automatic colors for candidates based on index.
    Uses HSL for visually distinct colors without hardcoding.
    """
    # Use prime number for good distribution
    hue = (index * 31) % 360
    return f'hsl({hue}, 70%, 55%)'



from datetime import datetime, timedelta
import pytz

# Define Philippine timezone once
PHILIPPINE_TZ = pytz.timezone('Asia/Manila')

def get_philippine_time():
    """Get current time in Philippine Timezone (UTC+8)"""
    return datetime.now(PHILIPPINE_TZ)

def get_philippine_time_naive():
    """Get current time in Philippine Timezone as naive datetime (for database)"""
    return datetime.now(PHILIPPINE_TZ).replace(tzinfo=None)

# ------------------- Configuration ------------------- #
MAX_ATTEMPTS = 3          # max allowed failed username/password attempts
COOLDOWN_TIME = 300       # cooldown in seconds (5 minutes)
MAX_2FA_ATTEMPTS = 5      # max allowed failed 2FA attempts
TWO_FA_COOLDOWN = 300     # cooldown for 2FA in seconds

from admin.models import AccessCode
from datetime import datetime
import time


ACCESS_ATTEMPTS_LIMIT = 5
ACCESS_COOLDOWN = 300  # 5 minutes in seconds

from admin.models import AccessCode
from datetime import datetime
import time
from flask import abort

ACCESS_ATTEMPTS_LIMIT = 5
ACCESS_COOLDOWN = 300  # 5 minutes in seconds


# ------------------- Dynamic Admin Access Code Entry ------------------- #
@admin_bp.route('/<path:secret_path>', methods=['GET', 'POST'])
def dynamic_access(secret_path):
    """Dynamic URL access code entry - path comes from database"""
    
    if not AccessCode.verify_path(secret_path):

        abort(404)
    
    ip = request.remote_addr
    
    if session.get('access_granted'):
        return redirect(url_for('admin.login'))
    
    attempts_key = f'access_attempts_{ip}'
    cooldown_key = f'access_cooldown_{ip}'
    
    attempts = session.get(attempts_key, 0)
    cooldown = session.get(cooldown_key, 0)
    
    if time.time() < cooldown:
        remaining = int(cooldown - time.time())
        error = f'Too many failed attempts. Try again in {remaining} seconds.'
        return render_template('access_code.html', error=error)
    
    if request.method == 'POST':
        entered_code = request.form.get('access_code', '').strip()
        
        if AccessCode.verify_code(entered_code):

            session[attempts_key] = 0
            session[cooldown_key] = 0
            session['access_granted'] = True
            session['access_time'] = datetime.now().isoformat()

            log_audit(
                action='ACCESS_CODE_SUCCESS',
                description=f"Admin access code entered successfully from IP: {ip}"
            )
            
            return redirect(url_for('admin.login'))
        else:

            attempts += 1
            session[attempts_key] = attempts
            
            log_audit(
                action='ACCESS_CODE_FAILED',
                description=f"Failed access code attempt from IP: {ip} - Attempt {attempts}"
            )
            
            if attempts >= ACCESS_ATTEMPTS_LIMIT:
                session[cooldown_key] = time.time() + ACCESS_COOLDOWN
                session[attempts_key] = 0
                error = 'Too many failed attempts. Access temporarily locked.'
                
                log_audit(
                    action='ACCESS_LOCKED',
                    description=f"Access locked for IP: {ip} due to {ACCESS_ATTEMPTS_LIMIT} failed attempts"
                )
            else:
                error = f'Invalid access code. Attempt {attempts} of {ACCESS_ATTEMPTS_LIMIT}.'
            
            return render_template('access_code.html', error=error)
    
    return render_template('access_code.html')

# ------------------- AJAX Verify Route ------------------- #
@admin_bp.route('/verify-access-code', methods=['POST'])
def verify_access_code_ajax():
    """AJAX endpoint for auto-verifying access code"""
    
    ip = request.remote_addr
    
    attempts_key = f'access_attempts_{ip}'
    cooldown_key = f'access_cooldown_{ip}'
    
    attempts = session.get(attempts_key, 0)
    cooldown = session.get(cooldown_key, 0)
    
    if time.time() < cooldown:
        remaining = int(cooldown - time.time())
        return jsonify({
            'success': False,
            'message': f'Too many attempts. Try again in {remaining} seconds.'
        })
    
    entered_code = request.form.get('access_code', '').strip()
    
    if AccessCode.verify_code(entered_code):
        # Success
        session[attempts_key] = 0
        session[cooldown_key] = 0
        session['access_granted'] = True
        session['access_time'] = datetime.now().isoformat()
        
        log_audit(
            action='ACCESS_CODE_SUCCESS',
            description=f"Admin access code entered successfully from IP: {ip}"
        )
        
        return jsonify({
            'success': True,
            'redirect': url_for('admin.login')
        })
    else:
        # Failed attempt
        attempts += 1
        session[attempts_key] = attempts
        
        log_audit(
            action='ACCESS_CODE_FAILED',
            description=f"Failed access code attempt from IP: {ip} - Attempt {attempts}"
        )
        
        if attempts >= ACCESS_ATTEMPTS_LIMIT:
            session[cooldown_key] = time.time() + ACCESS_COOLDOWN
            session[attempts_key] = 0
            return jsonify({
                'success': False,
                'message': 'Too many failed attempts. Access temporarily locked.'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Invalid code. {ACCESS_ATTEMPTS_LIMIT - attempts} attempts remaining.'
            })

# ------------------- Modified Admin Login Route ------------------- #
@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Second layer: Actual login page - only accessible after access code"""
    
    if not session.get('access_granted'):

        secret_path = AccessCode.get_secret_path()

        return redirect(url_for('admin.dynamic_access', secret_path=secret_path))
    
    ip = request.remote_addr
    attempts = session.get(f'login_attempts_{ip}', 0)
    cooldown = session.get(f'login_cooldown_{ip}', 0)


    if current_user.is_authenticated and getattr(current_user, 'role', None) == 'Admin':
        return redirect(url_for('admin.dashboard'))


    if time.time() < cooldown:
        remaining = int(cooldown - time.time())
        error = f'Too many failed attempts. Try again in {remaining} seconds.'
        return render_template('admin_login.html', error=error)

    error = None

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        admin = Admin.query.filter(
            Admin.username == username
        ).first()
        
        if admin and admin.username == username:
            if bcrypt.check_password_hash(admin.password, password) and admin.role == 'Admin':

                session[f'login_attempts_{ip}'] = 0
                session[f'login_cooldown_{ip}'] = 0
                session.permanent = True

                # ---------- AUDIT LOG: Successful login attempt ----------
                log_audit(
                    action='LOGIN_SUCCESS',
                    description=f"Admin user '{username}' logged in successfully from IP: {ip}"
                )

                # Redirect to 2FA setup if no totp_secret
                if not getattr(admin, 'totp_secret', None):
                    session['pre_2fa_admin_id'] = admin.id
                    log_audit(
                        action='2FA_REQUIRED',
                        description=f"Admin user '{username}' redirected to 2FA setup - TOTP secret not configured"
                    )
                    return redirect(url_for('admin.setup_2fa'))
                else:
                    session['pre_2fa_admin_id'] = admin.id
                    log_audit(
                        action='2FA_VERIFICATION',
                        description=f"Admin user '{username}' redirected to 2FA verification"
                    )
                    return redirect(url_for('admin.verify_2fa'))
            else:
                admin = None
        else:
            admin = None


        if not admin:
            attempts += 1
            session[f'login_attempts_{ip}'] = attempts

            username_exists = Admin.query.filter(
                func.lower(Admin.username) == func.lower(username)
            ).first()
            
            if username_exists and username_exists.username != username:
                error_msg = f"Invalid username or password."
                log_msg = f"Case mismatch: Attempted '{username}', actual username is '{username_exists.username}'"
            else:
                error_msg = f'Invalid username or password. Attempt {attempts} of {MAX_ATTEMPTS}.'
                log_msg = f"Username '{username}' not found"

            log_audit(
                action='LOGIN_FAILED',
                description=f"Failed login attempt from IP: {ip} | {log_msg}"
            )

            if attempts >= MAX_ATTEMPTS:
                session[f'login_cooldown_{ip}'] = time.time() + COOLDOWN_TIME
                session[f'login_attempts_{ip}'] = 0
                error = 'Invalid credentials. Admin login temporarily locked.'
                
                log_audit(
                    action='ACCOUNT_LOCKED',
                    description=f"Admin login temporarily locked for IP: {ip} due to {MAX_ATTEMPTS} failed attempts"
                )
            else:
                error = error_msg

    return render_template('admin_login.html', error=error)


# ------------------- FORGOT PASSWORD FROM LOGIN PAGE (NO LOGIN REQUIRED) ------------------- #
@admin_bp.route('/login-forgot-password', methods=['POST'])
def login_forgot_password():
    """
    Send password reset email - Accessible from login page (no authentication required)
    Uses reset_token column in admins table (SAME as student pattern that works!)
    """
    try:
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Request must be JSON'
            }), 400
        
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        username = data.get('username', '').strip()
        
        if not email or not username:
            return jsonify({
                'success': False,
                'message': 'Email and username are required'
            }), 400
        
        # Find admin by email AND username
        admin = Admin.query.filter(
            func.lower(Admin.email) == email,
            func.lower(Admin.username) == func.lower(username)
        ).first()
        
        # Always return success even if admin not found (security - prevents email enumeration)
        if not admin:
            return jsonify({
                'success': True,
                'message': 'If an account matches your email and username, a reset link will be sent.'
            }), 200
        
        # Check if admin is active
        if admin.status != 'Active':
            return jsonify({
                'success': True,
                'message': 'If an account matches your email and username, a reset link will be sent.'
            }), 200
        
        # Generate reset token (using secrets, same as student)
        token = secrets.token_urlsafe(32)
        
        # Store token in the reset_token column (NO expiry time, just like student!)
        admin.reset_token = token
        db.session.commit()
        
        # Generate reset URL
        reset_url = url_for('admin.reset_password_page', token=token, _external=True)
        
        # Render email template
        email_html = render_template('admin_password_reset_email.html',
                                   admin_name=f"{admin.first_name} {admin.last_name}",
                                   admin_username=admin.username,
                                   reset_url=reset_url)
        
        # Send email
        from flask_mail import Message
        from extensions import mail
        
        msg = Message(
            'Reset Your Admin Password - CTU COMELEC',
            recipients=[admin.email]
        )
        msg.html = email_html
        
        mail.send(msg)
        
        # Audit log
        log_audit(
            action='PASSWORD_RESET_REQUESTED_FROM_LOGIN',
            description=f"Password reset requested for admin '{admin.username}' from login page IP: {request.remote_addr}"
        )
        
        return jsonify({
            'success': True,
            'message': 'Password reset instructions sent to your email!'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error sending password reset from login: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500


from flask import get_flashed_messages
@admin_bp.route('/2fa/verify', methods=['GET', 'POST'])
def verify_2fa():

    get_flashed_messages()

    # Check if user is in pre-2fa session
    if 'pre_2fa_admin_id' not in session:
        flash('Please login first.', 'warning')
        return redirect(url_for('admin.login'))

    admin_id = session['pre_2fa_admin_id']
    admin = Admin.query.get(admin_id)
    if not admin:
        session.pop('pre_2fa_admin_id', None)
        return redirect(url_for('admin.login'))

    ip = request.remote_addr
    attempts = session.get(f'2fa_attempts_{ip}', 0)
    cooldown = session.get(f'2fa_cooldown_{ip}', 0)

    # Check cooldown
    if time.time() < cooldown:
        remaining = int(cooldown - time.time())
        error = f"Too many 2FA attempts. Try again in {remaining} seconds."
        return render_template('admin_2fa_verify.html', error=error)

    error = None
    
    if request.method == 'POST':
        # Check if user has TOTP secret
        if not admin.totp_secret:
            error = "2FA not properly set up. Please setup 2FA first."
            return render_template('admin_2fa_verify.html', error=error)

        totp = pyotp.TOTP(admin.totp_secret)
        code = request.form.get('code')
        
        if totp.verify(code):
            # Success: log in the user
            login_user(admin)
            session.permanent = True
            print(f"✅ User {admin.username} logged in after 2FA")

            # ===== DEVICE FINGERPRINT AND TRUST CHECK =====
            device_info = AdminTrustedDevice.get_device_info(request)
            
            # 🔥 FIXED: EXCLUDE IP address from fingerprint (IP changes!)
            # OLD (WRONG): fingerprint_data = f"{admin.id}{device_info['ip_address']}{device_info['user_agent']}{device_info['browser']}{device_info['os']}"
            # NEW (CORRECT) - NO IP ADDRESS:
            fingerprint_data = f"{admin.id}{device_info['user_agent']}{device_info['browser']}{device_info['os']}"
            device_fingerprint = hashlib.sha256(fingerprint_data.encode()).hexdigest()

            print(f"🔑 Generated fingerprint: {device_fingerprint[:20]}...")
            print(f"📊 Data used: admin_id={admin.id}, browser={device_info['browser']}, os={device_info['os']}")
            print(f"⚠️ IP address NOT used in fingerprint (it changes)")
            
            # Store in session
            session['admin_device_fingerprint'] = device_fingerprint
            
            # 🔐 CHECK IF THIS IS A NEW DEVICE
            trusted_device = AdminTrustedDevice.query.filter_by(
                admin_id=admin.id,
                device_fingerprint=device_fingerprint,
                trusted=True
            ).first()
            
            print(f"🔍 Trusted device found: {trusted_device is not None}")
            
            # Check if ANY trusted device exists for this admin
            any_trusted_device = AdminTrustedDevice.query.filter_by(
                admin_id=admin.id,
                trusted=True
            ).first()
            
            print(f"🔍 Any trusted device exists: {any_trusted_device is not None}")
            
            if not trusted_device:
                print("📱 This is a NEW device")
                # This is a new/untrusted device
                if any_trusted_device:
                    print("📧 Existing trusted devices found → Sending verification email")
                    # There are existing trusted devices → need verification
                    
                    # Save device as untrusted
                    new_device = AdminTrustedDevice(
                        admin_id=admin.id,
                        device_name=f"{device_info['browser']} on {device_info['os']}",
                        ip_address=device_info['ip_address'],
                        user_agent=device_info['user_agent'],
                        browser=device_info['browser'],
                        os=device_info['os'],
                        device_type=device_info['device_type'],
                        trusted=False
                    )
                    # Set fingerprint using the fixed method (NO IP)
                    new_device.device_fingerprint = device_fingerprint
                    db.session.add(new_device)
                    db.session.commit()
                    
                    # Send verification email
                    from admin.utils import send_admin_new_device_email
                    send_admin_new_device_email(admin, new_device)
                    
                    # Store in session for verification page
                    session['pending_admin_login'] = admin.id
                    session['pending_device_fp'] = device_fingerprint
                    
                    # Logout temporarily (they'll log in after verification)
                    from flask_login import logout_user
                    logout_user()
                    
                    # Cleanup 2FA session
                    session.pop('pre_2fa_admin_id', None)
                    session.pop(f'2fa_attempts_{ip}', None)
                    session.pop(f'2fa_cooldown_{ip}', None)
                    
                    print("↩️ Redirecting to verification page")
                    flash("New device detected. Please check your email to verify this device.", "warning")
                    return redirect(url_for('admin.verify_device_page'))
                else:
                    print("🎉 FIRST device ever - auto trusting it")
                    # FIRST device ever - auto trust it
                    new_device = AdminTrustedDevice(
                        admin_id=admin.id,
                        device_name=f"{device_info['browser']} on {device_info['os']}",
                        ip_address=device_info['ip_address'],
                        user_agent=device_info['user_agent'],
                        browser=device_info['browser'],
                        os=device_info['os'],
                        device_type=device_info['device_type'],
                        trusted=True,
                        expires_at=datetime.utcnow() + timedelta(days=30)
                    )
                    # Set fingerprint using the fixed method (NO IP)
                    new_device.device_fingerprint = device_fingerprint
                    db.session.add(new_device)
                    db.session.commit()
                    
                    # Cleanup session
                    session.pop('pre_2fa_admin_id', None)
                    session.pop(f'2fa_attempts_{ip}', None)
                    session.pop(f'2fa_cooldown_{ip}', None)
                    
                    print("🏠 Redirecting to dashboard (first device)")
                    flash("2FA verified successfully! This device has been added to trusted devices.", "success")
                    return redirect(url_for('admin.dashboard'))
            else:
                print("✅ Device ALREADY TRUSTED - direct to dashboard")
                # Device already trusted - just update last used
                trusted_device.update_last_used()
                # Update device info to keep it current (IP can change, that's fine)
                trusted_device.ip_address = device_info['ip_address']
                trusted_device.user_agent = device_info['user_agent']
                trusted_device.browser = device_info['browser']
                trusted_device.os = device_info['os']
                trusted_device.device_type = device_info['device_type']
                trusted_device.device_name = f"{device_info['browser']} on {device_info['os']}"
                db.session.commit()
                
                # Cleanup session
                session.pop('pre_2fa_admin_id', None)
                session.pop(f'2fa_attempts_{ip}', None)
                session.pop(f'2fa_cooldown_{ip}', None)
                
                print("🏠 Redirecting to dashboard (trusted device)")
                return redirect(url_for('admin.dashboard'))
        else:
            # Increment failed attempts
            attempts += 1
            session[f'2fa_attempts_{ip}'] = attempts

            if attempts >= MAX_2FA_ATTEMPTS:
                session[f'2fa_cooldown_{ip}'] = time.time() + TWO_FA_COOLDOWN
                session[f'2fa_attempts_{ip}'] = 0
                error = "Too many invalid codes. 2FA temporarily locked."
            else:
                error = f"Invalid code. Attempt {attempts} of {MAX_2FA_ATTEMPTS}."

    return render_template('admin_2fa_verify.html', error=error)

# ------------------- 2FA Secret Generation ------------------- #
def generate_2fa_secret(admin):
    if not admin.totp_secret:  # Simpler condition
        admin.totp_secret = pyotp.random_base32()
        db.session.commit()
        
        log_audit(
            action='2FA_SECRET_GENERATED',
            description=f"2FA secret generated for admin user '{admin.username}'"
        )
        
    return admin.totp_secret



# ------------------- Admin 2FA Setup ------------------- #
@admin_bp.route('/2fa/setup', methods=['GET', 'POST'])
def setup_2fa():
    # Check if user is in pre-2fa session
    if 'pre_2fa_admin_id' not in session:
        # REMOVED the flash message to avoid notification
        # flash('Please login first.', 'warning')  # <-- COMMENT THIS OUT
        return redirect(url_for('admin.login'))
    
    admin_id = session['pre_2fa_admin_id']
    admin = Admin.query.get(admin_id)
    
    if not admin:
        session.pop('pre_2fa_admin_id', None)
        return redirect(url_for('admin.login'))
    
    # Generate secret if not exists
    if not admin.totp_secret:
        admin.totp_secret = pyotp.random_base32()
        db.session.commit()
    
    secret = admin.totp_secret
    totp = pyotp.TOTP(secret)

    # Only GET request - no POST processing since you removed the form
    totp_uri = totp.provisioning_uri(
        name=admin.email,
        issuer_name="CTU-COMELEC Admin"
    )
    return render_template('admin_2fa_setup.html', totp_uri=totp_uri, secret=secret)


# ==================== 2FA DISABLE WITH EMAIL CONFIRMATION ====================

@admin_bp.route('/2fa/disable', methods=['POST'])
@login_required
def request_disable_2fa():
    """Step 1: Request to disable 2FA - sends confirmation email"""
    try:
        admin = current_user
        
        # Check if 2FA is actually enabled
        if not admin.totp_secret:
            return jsonify({'success': False, 'message': '2FA is not enabled'}), 400
        
        # Delete any existing unused tokens for this admin
        TwoFADisableToken.query.filter_by(
            admin_id=admin.id, 
            used=False
        ).delete()
        db.session.commit()
        
        # Create new token
        token = TwoFADisableToken(admin_id=admin.id)
        db.session.add(token)
        db.session.commit()
        
        # Send email
        from admin.utils import send_2fa_disable_confirmation
        email_sent = send_2fa_disable_confirmation(admin, token)
        
        if email_sent:
            # Log this request
            log_audit(
                action='2FA_DISABLE_REQUESTED',
                description=f"Admin user '{admin.username}' requested to disable 2FA from IP: {request.remote_addr}"
            )
            
            return jsonify({
                'success': True, 
                'message': 'Confirmation email sent. Please check your email to confirm disabling 2FA.'
            })
        else:
            return jsonify({
                'success': False, 
                'message': 'Failed to send confirmation email. Please try again.'
            }), 500
            
    except Exception as e:
        db.session.rollback()
        print(f"Error in request_disable_2fa: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/2fa/disable/confirm/<token>')
def confirm_disable_2fa(token):
    """Step 2: User clicks YES in email - actually disable 2FA"""
    try:
        # Find the token
        disable_token = TwoFADisableToken.query.filter_by(token=token, used=False).first()
        
        # Variables for template
        success = False
        message = ""
        title = ""
        
        if not disable_token:
            title = "Invalid Link"
            message = "This confirmation link is invalid or has already been used. Please request a new one."
        else:
            # Check if token expired
            from datetime import datetime
            if datetime.utcnow() > disable_token.expires_at:
                title = "Link Expired"
                message = "This confirmation link has expired. Please request a new one from your settings."
            else:
                # Get the admin
                admin = Admin.query.get(disable_token.admin_id)
                
                if not admin:
                    title = "Invalid Link"
                    message = "Admin account not found."
                else:
                    # Disable 2FA
                    admin.totp_secret = None
                    disable_token.used = True
                    
                    db.session.commit()
                    
                    # Only log audit if user is authenticated (has id)
                    if current_user.is_authenticated and hasattr(current_user, 'id'):
                        log_audit(
                            action='2FA_DISABLED',
                            description=f"Admin user '{admin.username}' disabled two-factor authentication via email confirmation from IP: {request.remote_addr}"
                        )
                    
                    title = "2FA Disabled Successfully"
                    message = "Two-Factor Authentication has been disabled on your account."
                    success = True
        
        # Return simple HTML with only close window button
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    max-width: 500px;
                }}
                .success {{ color: #10b981; }}
                .error {{ color: #ef4444; }}
                .icon {{ font-size: 48px; margin-bottom: 20px; }}
                h1 {{ color: #333; margin-bottom: 10px; }}
                p {{ color: #666; margin-bottom: 30px; line-height: 1.5; }}
                button {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #667eea;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    cursor: pointer;
                    font-weight: 600;
                    transition: all 0.3s ease;
                }}
                button:hover {{
                    background: #764ba2;
                    transform: translateY(-2px);
                    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                }}
                button:active {{
                    transform: translateY(0);
                }}
            </style>
            <script>
                function closeWindow() {{
                    // Check if window was opened by script
                    if (window.opener) {{
                        // Close this window/tab
                        window.close();
                    }} else {{
                        // If no opener, show message and try to close anyway
                        alert('You can now close this window/tab.');
                        window.close();
                    }}
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="icon {'success' if success else 'error'}">
                    {'✓' if success else '✗'}
                </div>
                <h1>{title}</h1>
                <p>{message}</p>
                <button onclick="closeWindow()">
                    <i style="margin-right: 8px;">✕</i> Close Window
                </button>
            </div>
        </body>
        </html>
        """
        
    except Exception as e:
        print(f"Error in confirm_disable_2fa: {str(e)}")
        db.session.rollback()
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    max-width: 500px;
                }}
                .error {{ color: #ef4444; font-size: 48px; margin-bottom: 20px; }}
                h1 {{ color: #333; margin-bottom: 10px; }}
                p {{ color: #666; margin-bottom: 30px; line-height: 1.5; }}
                button {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #667eea;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    cursor: pointer;
                    font-weight: 600;
                    transition: all 0.3s ease;
                }}
                button:hover {{
                    background: #764ba2;
                    transform: translateY(-2px);
                    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                }}
                button:active {{
                    transform: translateY(0);
                }}
            </style>
            <script>
                function closeWindow() {{
                    if (window.opener) {{
                        window.close();
                    }} else {{
                        alert('An error occurred. You can now close this window/tab.');
                        window.close();
                    }}
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="error">✗</div>
                <h1>Error</h1>
                <p>An error occurred while disabling 2FA. Please try again or contact support.</p>
                <button onclick="closeWindow()">
                    <i style="margin-right: 8px;">✕</i> Close Window
                </button>
            </div>
        </body>
        </html>
        """


@admin_bp.route('/2fa/disable/cancel/<token>')
def cancel_disable_2fa(token):
    """Step 3: User clicks NO in email - cancel the request"""
    try:
        # Find the token
        disable_token = TwoFADisableToken.query.filter_by(token=token, used=False).first()
        
        success = False
        title = ""
        message = ""
        
        if disable_token:
            # Mark as used so it can't be used later
            disable_token.used = True
            
            # Only log audit if user is authenticated
            if current_user.is_authenticated and hasattr(current_user, 'id'):
                log_audit(
                    action='2FA_DISABLE_CANCELLED',
                    description=f"Admin user '{current_user.username}' cancelled 2FA disable request from IP: {request.remote_addr}"
                )
            
            db.session.commit()
            success = True
            title = "Request Cancelled"
            message = "Your 2FA disable request has been cancelled. Your account remains protected with Two-Factor Authentication."
        else:
            title = "Invalid Link"
            message = "This cancellation link is invalid or has already been used."
        
        # Return simple HTML with only close window button
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    max-width: 500px;
                }}
                .success {{ color: #10b981; }}
                .error {{ color: #ef4444; }}
                .icon {{ font-size: 48px; margin-bottom: 20px; }}
                h1 {{ color: #333; margin-bottom: 10px; }}
                p {{ color: #666; margin-bottom: 30px; line-height: 1.5; }}
                button {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #667eea;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    cursor: pointer;
                    font-weight: 600;
                    transition: all 0.3s ease;
                }}
                button:hover {{
                    background: #764ba2;
                    transform: translateY(-2px);
                    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                }}
                button:active {{
                    transform: translateY(0);
                }}
            </style>
            <script>
                function closeWindow() {{
                    // Check if window was opened by script
                    if (window.opener) {{
                        // Close this window/tab
                        window.close();
                    }} else {{
                        // If no opener, show message and try to close anyway
                        alert('You can now close this window/tab.');
                        window.close();
                    }}
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="icon {'success' if success else 'error'}">
                    {'✓' if success else '✗'}
                </div>
                <h1>{title}</h1>
                <p>{message}</p>
                <button onclick="closeWindow()">
                    <i style="margin-right: 8px;">✕</i> Close Window
                </button>
            </div>
        </body>
        </html>
        """
        
    except Exception as e:
        print(f"Error in cancel_disable_2fa: {str(e)}")
        db.session.rollback()
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    max-width: 500px;
                }}
                .error {{ color: #ef4444; font-size: 48px; margin-bottom: 20px; }}
                h1 {{ color: #333; margin-bottom: 10px; }}
                p {{ color: #666; margin-bottom: 30px; line-height: 1.5; }}
                button {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #667eea;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    cursor: pointer;
                    font-weight: 600;
                    transition: all 0.3s ease;
                }}
                button:hover {{
                    background: #764ba2;
                    transform: translateY(-2px);
                    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                }}
                button:active {{
                    transform: translateY(0);
                }}
            </style>
            <script>
                function closeWindow() {{
                    if (window.opener) {{
                        window.close();
                    }} else {{
                        alert('An error occurred. You can now close this window/tab.');
                        window.close();
                    }}
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="error">✗</div>
                <h1>Error</h1>
                <p>An error occurred while cancelling the request. Please try again or contact support.</p>
                <button onclick="closeWindow()">
                    <i style="margin-right: 8px;">✕</i> Close Window
                </button>
            </div>
        </body>
        </html>
        """




@admin_bp.route('/2fa/check-status-changed')
@login_required
def check_2fa_status_changed():
    """Check if 2FA status changed from email confirmation"""
    changed = session.pop('2fa_status_changed', False)
    new_status = session.pop('2fa_new_status', None)
    
    print(f"2FA status check: changed={changed}, new_status={new_status}")  # Debug log
    
    return jsonify({
        'changed': changed,
        'new_status': new_status
    })

        
@admin_bp.route('/verify-device')
def verify_device_page():
    """Page shown when new device needs verification"""
    admin_id = session.get('pending_admin_login')
    fingerprint = session.get('pending_device_fp')
    
    if not admin_id or not fingerprint:
        flash("No pending device verification.", "danger")
        return redirect(url_for('admin.login'))
    
    admin = Admin.query.get(admin_id)
    if not admin:
        session.pop('pending_admin_login', None)
        session.pop('pending_device_fp', None)
        return redirect(url_for('admin.login'))
    
    device = AdminTrustedDevice.query.filter_by(
        admin_id=admin.id,
        device_fingerprint=fingerprint,
        trusted=False
    ).first()
    
    if not device:
        # Maybe already verified? Check session
        return render_template('admin_verify_device.html', admin=admin, message="Device already verified?", redirect_after=3)
    
    # 🚫 REMOVED: VIEW audit log for device verification page
    
    return render_template('admin_verify_device.html', admin=admin, device=device)


@admin_bp.route('/verify-device/resend', methods=['POST'])
def resend_admin_verification():
    """Resend verification email"""
    admin_id = session.get('pending_admin_login')
    fingerprint = session.get('pending_device_fp')
    
    if not admin_id or not fingerprint:
        return jsonify({"success": False, "message": "No pending verification"}), 400
    
    admin = Admin.query.get(admin_id)
    device = AdminTrustedDevice.query.filter_by(
        admin_id=admin_id,
        device_fingerprint=fingerprint,
        trusted=False
    ).first()
    
    if not device:
        return jsonify({"success": False, "message": "Device not found"}), 404
    
    from admin.utils import send_admin_new_device_email
    send_admin_new_device_email(admin, device)
    
    # 🚫 REMOVED: AUDIT log for resending verification (not a data modification)
    # Email resend is not changing database state
    
    return jsonify({"success": True, "message": "Verification email resent"})


@admin_bp.route('/verify-device/confirm/<token>')
def confirm_admin_device(token):
    """Confirm new device via email link (Yes, it's me)"""
    device = AdminTrustedDevice.query.filter_by(verification_token=token).first()
    
    if not device:
        return render_template('admin_device_result.html', 
                             success=False, 
                             message='Invalid or expired verification link.')
    
    # Check if token expired (15 minutes)
    if device.verification_sent_at and datetime.utcnow() > device.verification_sent_at + timedelta(minutes=15):
        db.session.delete(device)
        db.session.commit()
        return render_template('admin_device_result.html',
                             success=False,
                             message='Verification link has expired. Please try logging in again.')
    
    # Mark device as trusted
    device.trusted = True
    device.verification_token = None
    device.expires_at = datetime.utcnow() + timedelta(days=30)
    device.last_used = datetime.utcnow()
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (DEVICE VERIFIED - important security action)
    admin = device.admin
    username = getattr(admin, 'username', 'Unknown')
    ip = request.remote_addr
    log_audit(
        action='DEVICE_VERIFIED',
        description=f"Admin user '{username}' verified new device: {device.device_name} from IP: {ip}"
    )
    
    # If this is the device that was pending, we can auto-login
    # Check if this matches pending session
    pending_fp = session.get('pending_device_fp')
    if pending_fp and pending_fp == device.device_fingerprint:
        admin = device.admin
        login_user(admin)
        session.permanent = True
        session.pop('pending_admin_login', None)
        session.pop('pending_device_fp', None)
        
        return render_template('admin_device_result.html',
                             success=True,
                             message='Device verified successfully! You are now being logged in...',
                             auto_redirect=True,
                             redirect_url=url_for('admin.dashboard'))
    
    return render_template('admin_device_result.html',
                         success=True,
                         message='Device verified successfully! You can now close this window and return to login.')


@admin_bp.route('/verify-device/reject/<token>')
def reject_admin_device(token):
    """Reject new device via email link (No, it's not me)"""
    device = AdminTrustedDevice.query.filter_by(verification_token=token).first()
    
    if device:
        admin_username = getattr(device.admin, 'username', 'Unknown')
        
        db.session.delete(device)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (DEVICE REJECTED - important security action)
        log_audit(
            action='DEVICE_REJECTED',
            description=f"Admin user '{admin_username}' rejected new device via email link from IP: {request.remote_addr}"
        )
    
    return render_template('admin_device_result.html',
                         success=False,
                         message='Device rejected. No action needed.')


@admin_bp.route('/verify-device/status')
def verify_device_status():
    """Check if device has been verified (polling endpoint)"""
    admin_id = session.get('pending_admin_login')
    fingerprint = session.get('pending_device_fp')
    
    if not admin_id or not fingerprint:
        return jsonify({"status": "no_session"})
    
    # Check if device is now trusted
    device = AdminTrustedDevice.query.filter_by(
        admin_id=admin_id,
        device_fingerprint=fingerprint,
        trusted=True
    ).first()
    
    if device:
        # 🚀 AUTO-LOGIN THE ADMIN HERE
        admin = Admin.query.get(admin_id)
        if admin:
            login_user(admin)
            session.permanent = True
            
            # Update last used
            device.update_last_used()
            db.session.commit()
            
            # ✅ KEEP THIS AUDIT LOG (DEVICE VERIFIED VIA POLLING)
            log_audit(
                action='DEVICE_AUTO_VERIFIED',
                description=f"Admin user '{admin.username}' auto-verified device via polling from IP: {request.remote_addr}"
            )
            
            # Clean up session
            session.pop('pending_admin_login', None)
            session.pop('pending_device_fp', None)
            
            return jsonify({"status": "verified"})
    
    # Check if device was rejected/deleted
    device_any = AdminTrustedDevice.query.filter_by(
        admin_id=admin_id,
        device_fingerprint=fingerprint
    ).first()
    
    if device_any is None:
        # Clean up session on rejection
        session.pop('pending_admin_login', None)
        session.pop('pending_device_fp', None)
        return jsonify({"status": "rejected"})
    
    return jsonify({"status": "pending"})


@admin_bp.route('/2fa/setup-data', methods=['GET'])
def get_2fa_setup_data():
    """AJAX endpoint to get 2FA setup data for inline setup"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    admin = current_user
    
    # Generate secret if not exists
    if not admin.totp_secret:
        admin.totp_secret = pyotp.random_base32()
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (2FA SECRET GENERATED - important security action)
        log_audit(
            action='2FA_SECRET_GENERATED',
            description=f"Admin user '{admin.username}' generated 2FA secret from IP: {request.remote_addr}"
        )
    
    secret = admin.totp_secret
    totp = pyotp.TOTP(secret)
    
    totp_uri = totp.provisioning_uri(
        name=admin.email,
        issuer_name="CTU-COMELEC Admin"
    )
    
    # 🚫 REMOVED: VIEW audit log for 2FA setup data
    
    return jsonify({
        'success': True,
        'totp_uri': totp_uri,
        'secret': secret,
        'email': admin.email
    })

# ---------------------- DASHBOARD ---------------------- #
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    # Get year filter from URL parameters or session
    year = request.args.get('year')
    
    # If no year parameter, try to get from session
    if not year:
        year = session.get('admin_current_year')
    
    # Get all available years from elections first
    all_elections = Election.query.order_by(Election.start_date.asc()).all()
    available_years = set()
    for election in all_elections:
        if election.start_date:
            available_years.add(election.start_date.year)
    
    # Sort years in descending order (newest first)
    available_years = sorted(list(available_years), reverse=True)
    
    # If no year is selected, default to the latest election's year
    if not year and available_years:
        # Get the most recent election
        latest_election = Election.query.order_by(Election.start_date.desc()).first()
        if latest_election and latest_election.start_date:
            year = latest_election.start_date.year
            if year in available_years:
                session['admin_current_year'] = year
    
    # Parse year to date range (January 1 to December 31 of that year)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    # Save to session
    if year:
        session['admin_current_year'] = year
    
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)

    # ================================
    # Build base queries with year filter
    # ================================
    election_query = Election.query
    vote_query = Vote.query
    
    if start_date and end_date:
        # For elections, filter by start date in range
        election_query = election_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
        
        # For votes, filter by cast_timestamp in range
        vote_query = vote_query.filter(
            Vote.cast_timestamp >= start_date,
            Vote.cast_timestamp <= end_date
        )
    
    # ================================
    # KPI counts with year filter
    # ================================
    total_students = Student.query.count()  # Total students always same (no filter)
    total_elections = election_query.count()
    total_votes = vote_query.count()
    
    # ===== Get pending applications count filtered by year =====
    from student.models import PendingCandidate
    
    # Filter pending applications by year (based on election dates)
    pending_applications_query = PendingCandidate.query.filter_by(status='pending')
    
    if start_date and end_date:
        # Join with Election to filter by year
        pending_applications_query = pending_applications_query.join(
            Election, PendingCandidate.election_id == Election.id
        ).filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    pending_applications = pending_applications_query.count()
    
    # ===== Get pending deletion requests count filtered by year =====
    # Filter deletion requests by request date (when the request was made)
    pending_deletion_query = DeletionRequest.query.filter_by(status='pending')
    
    if start_date and end_date:
        pending_deletion_query = pending_deletion_query.filter(
            DeletionRequest.request_date >= start_date,
            DeletionRequest.request_date <= end_date
        )
    
    pending_deletion_requests = pending_deletion_query.count()
    
    # ===== Get ongoing elections count (based on current date, filtered by year) =====
    ongoing_elections_query = Election.query.filter(
        Election.start_date <= now,
        Election.end_date >= now
    )
    
    if start_date and end_date:
        ongoing_elections_query = ongoing_elections_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    ongoing_elections = ongoing_elections_query.count()

    # ================================
    # Recent elections table with year filter (with pagination)
    # ================================
    elections = election_query.order_by(Election.start_date.desc()).all()

    # Localize timezone if naive
    for election in elections:
        if election.start_date.tzinfo is None:
            election.start_date = tz.localize(election.start_date)
        if election.end_date.tzinfo is None:
            election.end_date = tz.localize(election.end_date)

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    total_elections_count = len(elections)
    total_pages = (total_elections_count + per_page - 1) // per_page if total_elections_count > 0 else 1
    
    if page < 1:
        page = 1
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, total_elections_count)
    paginated_elections = elections[start_idx:end_idx]
    
    start_index = start_idx + 1 if elections else 0
    end_index = end_idx
    
    # Generate page range
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        if page <= 3:
            page_range = [1, 2, 3, 4, '...', total_pages - 1, total_pages]
        elif page >= total_pages - 2:
            page_range = [1, 2, '...', total_pages - 3, total_pages - 2, total_pages - 1, total_pages]
        else:
            page_range = [1, '...', page - 1, page, page + 1, '...', total_pages]

    return render_template(
        'admin_dashboard.html',
        admin=current_user,
        total_students=total_students,
        total_elections=total_elections,
        ongoing_elections=ongoing_elections,
        total_votes=total_votes,
        pending_applications=pending_applications,
        pending_deletion_requests=pending_deletion_requests,
        paginated_elections=paginated_elections,
        total_elections_count=total_elections_count,
        total_pages=total_pages,
        current_page=page,
        start_index=start_index,
        end_index=end_index,
        page_range=page_range,
        available_years=available_years,  
        current_year=year,  
        now=now
    )



@admin_bp.route('/settings')
@admin_required
def settings():
    """
    System Settings page
    """
    try:
        tz = pytz.timezone('Asia/Manila')
        now = datetime.now(tz)
        
        # Get audit logs for the logs section
        page = request.args.get("page", 1, type=int)
        search = request.args.get("search", "")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        
        query = AuditLog.query
        
        # Search filter
        if search:
            query = query.filter(
                or_(
                    AuditLog.action.ilike(f"%{search}%"),
                    AuditLog.role.ilike(f"%{search}%"),
                    AuditLog.description.ilike(f"%{search}%")
                )
            )
        
        # Date filters
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(AuditLog.timestamp >= start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
                end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
                query = query.filter(AuditLog.timestamp <= end_date_obj)
            except ValueError:
                pass
        
        # Order by latest first
        query = query.order_by(desc(AuditLog.timestamp))
        
        # Pagination
        logs = query.paginate(page=page, per_page=20)
        
        # Get trusted devices for current admin
        trusted_devices = AdminTrustedDevice.query.filter_by(
            admin_id=current_user.id
        ).order_by(AdminTrustedDevice.last_used.desc()).all()
        
        # Mark current device
        current_fingerprint = session.get('admin_device_fingerprint')
        for device in trusted_devices:
            device.is_current = (device.device_fingerprint == current_fingerprint)
        
        # Get current access code - NEW
        access_code = AccessCode.get_active_code()
        
        # Get last backup info (you can implement this later)
        last_backup = None
        
        return render_template(
            'admin_settings.html',
            admin=current_user,
            now=now,
            logs=logs,
            search=search,
            start_date=start_date,
            end_date=end_date,
            pytz=pytz,
            last_backup=last_backup,
            trusted_devices=trusted_devices,
            access_code=access_code  # ADD THIS
        )
        
    except Exception as e:
        flash(f'Error loading settings: {str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/settings/save', methods=['POST'])
@admin_required
def save_settings():
    """
    Save system settings
    """
    try:
        data = request.get_json()
        section = data.get('section')
        settings = data.get('settings', {})
        
        # ✅ KEEP THIS AUDIT LOG (SETTINGS UPDATE - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        # Create a summary of what was changed
        changed_fields = ', '.join(settings.keys()) if settings else 'No fields'
        
        log_audit(
            action='SETTINGS_UPDATE',
            description=f"Admin user '{username}' updated {section} settings from IP: {ip} | Fields changed: {changed_fields}"
        )
        
        return jsonify({
            'success': True,
            'message': 'Settings saved successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500





@admin_bp.route('/settings/access-code/update', methods=['POST'])
@admin_required
def update_access_code():
    """
    Update the admin access code
    """
    try:
        data = request.get_json()
        new_code = data.get('access_code')
        expiration = data.get('expiration')
        
        if not new_code or len(new_code.strip()) < 4:
            return jsonify({
                'success': False,
                'message': 'Access code must be at least 4 characters long'
            }), 400
        
        # FIX: Use Philippine time
        from datetime import datetime, timedelta
        import pytz
        ph_tz = pytz.timezone('Asia/Manila')
        now_ph = datetime.now(ph_tz).replace(tzinfo=None)  # Naive for DB
        
        # Get current active access code
        current_code = AccessCode.get_active_code()
        
        if current_code:
            # Deactivate current code
            current_code.is_active = False
            current_code.updated_at = now_ph  # FIXED: Use Philippine time
            current_code.updated_by = current_user.id
        
        # Create new access code
        new_access_code = AccessCode(
            code=new_code.strip(),
            description=f"Updated by {current_user.username}",
            is_active=True,
            created_by=current_user.id,
            updated_by=current_user.id,
            created_at=now_ph,  # Add this if your model has it
            updated_at=now_ph   # Add this if your model has it
        )
        
        # Set expiration if provided
        if expiration:
            try:
                days = int(expiration)
                if days > 0:
                    # FIXED: Use Philippine time for expiration
                    new_access_code.expires_at = now_ph + timedelta(days=days)
            except (ValueError, TypeError):
                pass
        
        db.session.add(new_access_code)
        db.session.commit()
        
        # Audit log
        log_audit(
            action='ACCESS_CODE_UPDATED',
            description=f"Admin user '{current_user.username}' updated access code from IP: {request.remote_addr}"
        )
        
        return jsonify({
            'success': True,
            'message': 'Access code updated successfully!'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating access code: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error updating access code: {str(e)}'
        }), 500




# Dictionary to store email change requests
email_change_requests = {}

# ==================== ADMIN EMAIL CHANGE (DATABASE-BACKED) ====================
@admin_bp.route('/send-email-change-verification', methods=['POST'])
@admin_required
def send_email_change_verification():
    """Send verification email to old email address - DATABASE VERSION"""
    try:
        data = request.get_json()
        old_email = data.get('old_email')
        new_email = data.get('new_email')
        
        if not old_email or not new_email:
            return jsonify({'error': 'Emails are required'}), 400
        
        # Check if new email already exists
        existing_admin = Admin.query.filter_by(email=new_email).first()
        if existing_admin and existing_admin.id != current_user.id:
            return jsonify({'error': 'New email already in use'}), 400
        
        # Generate unique token
        import secrets
        token = secrets.token_urlsafe(32)
        
        # Get Philippine time (naive for DB)
        now_ph = get_philippine_time_naive()
        expiry_time = now_ph + timedelta(minutes=15)
        
        # ========== FIX #1: RESET confirmed flag for new request ==========
        current_user.email_change_confirmed = False  # <-- ADD THIS LINE
        
        # Store in database
        current_user.email_change_token = token
        current_user.new_email_pending = new_email
        current_user.email_change_requested_at = now_ph
        current_user.email_change_expires_at = expiry_time
        db.session.commit()
        
        # Mask email for display
        def mask_email(email):
            if not email or '@' not in email:
                return email
            name, domain = email.split('@')
            if len(name) > 2:
                return name[0] + '*' * (len(name) - 2) + name[-1] + '@' + domain
            elif len(name) == 2:
                return name[0] + '*@' + domain
            return email
        
        masked_new_email = mask_email(new_email)
        
        # Generate confirm and reject URLs
        confirm_url = url_for('admin.confirm_email_change', 
                            token=token, 
                            action='confirm', 
                            _external=True)
        reject_url = url_for('admin.confirm_email_change', 
                           token=token, 
                           action='reject', 
                           _external=True)
        
        # Render the email template
        email_html = render_template('verify_admin_email_change.html',
                                   masked_email=masked_new_email,
                                   confirm_url=confirm_url,
                                   reject_url=reject_url)
        
        # Send verification email to OLD email
        msg = Message(
            'Confirm Email Change Request - Admin Account',
            recipients=[old_email]
        )
        msg.html = email_html
        
        mail.send(msg)
        
        return jsonify({
            'success': True,
            'message': 'Verification email sent to your current email address'
        })
        
    except Exception as e:
        # Rollback on error
        current_user.email_change_token = None
        current_user.new_email_pending = None
        current_user.email_change_requested_at = None
        current_user.email_change_expires_at = None
        current_user.email_change_confirmed = False  # Also reset on error
        db.session.commit()
        
        current_app.logger.error(f"Error sending verification email: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/confirm-email-change/<token>/<action>')
def confirm_email_change(token, action):
    """Handle email confirmation from link - DATABASE VERSION"""
    print("\n" + "="*60)
    print(f"CONFIRM EMAIL CHANGE CALLED")
    print(f"Token: {token}")
    print(f"Action: {action}")
    print("="*60)
    
    # Find admin by token
    admin = Admin.query.filter_by(email_change_token=token).first()
    
    if not admin:
        return "Invalid request", 404
    
    print(f"[DEBUG] Found admin: {admin.username}")
    
    # Check expiry
    now_ph = get_philippine_time_naive()
    
    if now_ph > admin.email_change_expires_at:
        print(f"[ERROR] Token expired!")
        admin.email_change_token = None
        admin.new_email_pending = None
        admin.email_change_requested_at = None
        admin.email_change_expires_at = None
        admin.email_change_confirmed = False
        db.session.commit()
        return "Request expired", 400
    
    if action == 'confirm':
        print(f"[DEBUG] ACTION: CONFIRM - UPDATING DATABASE")
        
        # STORE CONFIRMATION IN DATABASE
        admin.email_change_confirmed = True
        db.session.commit()
        
        print(f"[DEBUG] admin.email_change_confirmed = {admin.email_change_confirmed}")
        print("="*60 + "\n")
        
        return """
        <html>
        <head>
            <style>
                body { font-family: Arial; text-align: center; padding: 50px; background: #f4f6fb; }
                .card { background: white; max-width: 400px; margin: 0 auto; padding: 30px; border-radius: 18px; box-shadow: 0 10px 30px rgba(0,0,0,0.08); }
                .icon { font-size: 64px; margin-bottom: 20px; }
                h2 { color: #1f2937; }
                p { color: #4b5563; }
                .success { color: #10b981; }
            </style>
        </head>
        <body>
            <div class="card">
                <div class="icon">✅</div>
                <h2 class="success">Email Ownership Confirmed!</h2>
                <p>You have successfully confirmed ownership of your current email address.</p>
                <p>Please return to the admin panel on your laptop.</p>
                <p>The OTP code will be sent automatically.</p>
                <p>This window can be closed.</p>
            </div>
        </body>
        </html>
        """
        
    elif action == 'reject':
        print(f"[DEBUG] ACTION: REJECT - CLEARING ALL DATA")
        admin.email_change_token = None
        admin.new_email_pending = None
        admin.email_change_requested_at = None
        admin.email_change_expires_at = None
        admin.email_change_confirmed = False
        db.session.commit()
        print(f"[DEBUG] All email change data cleared for admin: {admin.username}")
        print("="*60 + "\n")
        return "Email change rejected. You can close this window.", 200
    
    return "Invalid action", 400


@admin_bp.route('/check-email-confirmation', methods=['GET'])
@admin_required
def check_email_confirmation():
    """Check if old email has been confirmed via link click"""
    print("\n" + "="*40)
    print("CHECK EMAIL CONFIRMATION CALLED")
    
    # ✅ CHECK DATABASE FOR CONFIRMATION STATUS
    admin = current_user
    
    print(f"[DEBUG] Admin: {admin.username}")
    print(f"[DEBUG] Admin email_change_confirmed: {admin.email_change_confirmed}")
    print(f"[DEBUG] Admin email_change_token: {admin.email_change_token}")
    print(f"[DEBUG] Admin new_email_pending: {admin.new_email_pending}")
    
    if admin.email_change_confirmed and admin.email_change_token and admin.new_email_pending:
        print(f"[DEBUG] Confirmation found in DATABASE!")
        print("="*40 + "\n")
        
        return jsonify({
            'confirmed': True,
            'token': admin.email_change_token,
            'new_email': admin.new_email_pending
        })
    else:
        print(f"[DEBUG] No confirmation found")
        print("="*40 + "\n")
        return jsonify({'confirmed': False})


@admin_bp.route('/send-otp-to-new-email', methods=['POST'])
@admin_required
def send_otp_to_new_email():
    """Send OTP to new email address"""
    try:
        data = request.get_json()
        email = data.get('email')
        token = data.get('token')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        admin = current_user
        
        # ✅ CHECK DATABASE FOR CONFIRMATION
        if not admin.email_change_confirmed:
            return jsonify({'error': 'Please confirm your old email first'}), 400
        
        if admin.email_change_token != token:
            return jsonify({'error': 'Invalid session'}), 400
        
        if admin.new_email_pending != email:
            return jsonify({'error': 'Email mismatch'}), 400
        
        # Generate 6-digit OTP
        import random
        import string
        otp = ''.join(random.choices(string.digits, k=6))
        
        # Store OTP in session (short-term only)
        now_ph = get_philippine_time()
        otp_expiry = now_ph + timedelta(minutes=10)
        
        session['admin_new_email_otp'] = otp
        session['admin_new_email_otp_expiry'] = otp_expiry.timestamp()
        
        # Send OTP email
        msg = Message(
            'Verify Your New Email Address - Admin Account',
            recipients=[email]
        )
        msg.html = f"""
        <div style="font-family: Arial, sans-serif; text-align: center; padding: 30px;">
            <h2>Cebu Technological University Moalboal Campus</h2>
            <h3>Admin Account Email Verification</h3>
            <p>Your verification code for your new email address is:</p>
            <h1 style="font-size: 36px; letter-spacing: 5px; color: #2563eb;">{otp}</h1>
            <p>This code will expire in 10 minutes.</p>
        </div>
        """
        
        mail.send(msg)
        
        return jsonify({'success': True})
        
    except Exception as e:
        current_app.logger.error(f"Error sending OTP: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/verify-otp-code', methods=['POST'])
@admin_required
def verify_otp_code():
    """Verify OTP code and complete email change"""
    try:
        data = request.get_json()
        entered_otp = data.get('otp')
        email = data.get('email')
        
        if not entered_otp:
            return jsonify({'success': False, 'message': 'Code required'}), 400
        
        # Get stored OTP from session
        stored_otp = session.get('admin_new_email_otp')
        stored_expiry = session.get('admin_new_email_otp_expiry')
        
        if not stored_otp:
            return jsonify({'success': False, 'message': 'No code found'}), 400
        
        if datetime.now().timestamp() > stored_expiry:
            return jsonify({'success': False, 'message': 'Code expired'}), 400
        
        if entered_otp == stored_otp:
            admin = current_user
            
            # Update email
            old_email = admin.email
            admin.email = email
            
            # Clear all email change fields
            admin.email_change_token = None
            admin.new_email_pending = None
            admin.email_change_requested_at = None
            admin.email_change_expires_at = None
            admin.email_change_confirmed = False  # Reset flag
            
            db.session.commit()
            
            # Clear session
            session.pop('admin_new_email_otp', None)
            session.pop('admin_new_email_otp_expiry', None)
            
            # Send notification
            send_admin_email_change_notification(admin, old_email, email)
            
            return jsonify({'success': True, 'message': 'Email changed successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Invalid code'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def send_admin_email_change_notification(admin, old_email, new_email):
    """Send notification to old email about email change"""
    try:
        msg = Message(
            'Your Admin Email Has Been Changed',
            recipients=[old_email]
        )
        msg.html = f"""
        <div style="font-family: Arial, sans-serif;">
            <h2>Cebu Technological University Moalboal Campus</h2>
            <h3>Admin Account Email Change</h3>
            <p>Hello <strong>{admin.username}</strong>,</p>
            <p>Your email address for your admin account has been changed.</p>
            <p><strong>Old email:</strong> {old_email}<br>
            <strong>New email:</strong> {new_email}</p>
            <p>If you did not make this change, please contact the system administrator immediately.</p>
        </div>
        """
        
        mail.send(msg)
        current_app.logger.info(f"Email change notification sent to {old_email}")
    except Exception as e:
        current_app.logger.error(f"Failed to send email change notification: {e}")


@admin_bp.route('/settings/profile/update', methods=['POST'])
@admin_required
def settings_profile_update():
    """
    Update admin profile (username and email) with verification
    """
    import traceback
    print("\n" + "="*60)
    print("PROFILE UPDATE REQUEST RECEIVED")
    print("="*60)
    
    try:
        admin = current_user
        print(f"[DEBUG] Current admin ID: {admin.id}")
        print(f"[DEBUG] Current admin username: {admin.username}")
        print(f"[DEBUG] Current admin email: {admin.email}")
        
        # Get data from JSON
        data = request.get_json()
        print(f"[DEBUG] Raw request data: {data}")
        
        new_username = data.get('username')
        new_email = data.get('email')
        old_email_verified = data.get('old_email_verified')
        new_email_verified = data.get('new_email_verified')
        
        print(f"[DEBUG] Extracted values:")
        print(f"  - new_username: {new_username}")
        print(f"  - new_email: {new_email}")
        print(f"  - old_email_verified: {old_email_verified} (type: {type(old_email_verified)})")
        print(f"  - new_email_verified: {new_email_verified} (type: {type(new_email_verified)})")
        
        changes = {}
        email_changed = False
        old_email = admin.email
        
        # Update username if changed
        if new_username and new_username != admin.username:
            print(f"[DEBUG] Username change detected: {admin.username} -> {new_username}")
            existing_admin = Admin.query.filter_by(username=new_username).first()
            if existing_admin and existing_admin.id != admin.id:
                print(f"[DEBUG] Username already taken by another admin (ID: {existing_admin.id})")
                return jsonify({
                    'success': False,
                    'message': 'Username already taken'
                }), 400
            
            admin.username = new_username
            changes['username'] = new_username
            print(f"[DEBUG] Username updated to: {admin.username}")
        else:
            print(f"[DEBUG] No username change detected")
        
        # Check if email is being changed
        if new_email and new_email != admin.email:
            email_changed = True
            print(f"[DEBUG] Email change detected: {admin.email} -> {new_email}")
            
            # Convert to string for comparison
            old_verified_str = str(old_email_verified).lower() if old_email_verified is not None else 'false'
            new_verified_str = str(new_email_verified).lower() if new_email_verified is not None else 'false'
            
            print(f"[DEBUG] Verification flags after conversion:")
            print(f"  - old_verified_str: '{old_verified_str}'")
            print(f"  - new_verified_str: '{new_verified_str}'")
            
            if old_verified_str != 'true' or new_verified_str != 'true':
                print(f"[DEBUG] VERIFICATION FAILED: Both flags must be 'true'")
                print(f"  - old_verified_str == 'true': {old_verified_str == 'true'}")
                print(f"  - new_verified_str == 'true': {new_verified_str == 'true'}")
                return jsonify({
                    'success': False,
                    'message': 'Email change requires verification of both old and new emails'
                }), 400
            
            print(f"[DEBUG] Verification passed! Proceeding with email update...")
            
            # Check if new email already exists
            existing_admin = Admin.query.filter_by(email=new_email).first()
            if existing_admin and existing_admin.id != admin.id:
                print(f"[DEBUG] New email already in use by another admin (ID: {existing_admin.id})")
                return jsonify({
                    'success': False,
                    'message': 'Email already in use'
                }), 400
            
            # ACTUALLY UPDATE THE EMAIL
            old_email_value = admin.email
            admin.email = new_email
            changes['email'] = {'old': old_email_value, 'new': new_email}
            print(f"[DEBUG] Email updated in memory to: {admin.email}")
            
            # Send notification to old email
            try:
                send_admin_email_change_notification(admin, old_email_value, new_email)
                print(f"[DEBUG] Notification email sent to old address: {old_email_value}")
            except Exception as email_err:
                print(f"[DEBUG] Failed to send notification email: {str(email_err)}")
        else:
            print(f"[DEBUG] No email change detected")
        
        # Commit changes to database
        print(f"[DEBUG] Attempting to commit to database...")
        db.session.commit()
        print(f"[DEBUG] Database commit successful!")
        
        # Verify the changes were saved
        db_admin = Admin.query.get(admin.id)
        print(f"[DEBUG] Verification from fresh DB query:")
        print(f"  - Username in DB: {db_admin.username}")
        print(f"  - Email in DB: {db_admin.email}")
        
        # Audit log
        log_audit(
            action='PROFILE_UPDATED',
            description=f"Admin user '{admin.username}' updated profile from IP: {request.remote_addr}. Changes: {changes}"
        )
        
        print(f"[DEBUG] Returning success response")
        print("="*60 + "\n")
        
        return jsonify({
            'success': True,
            'message': 'Profile updated successfully',
            'new_username': admin.username,
            'new_email': admin.email
        })
        
    except Exception as e:
        db.session.rollback()
        error_traceback = traceback.format_exc()
        print(f"[DEBUG ERROR] Exception occurred:")
        print(f"  - Error: {str(e)}")
        print(f"  - Full traceback:\n{error_traceback}")
        print("="*60 + "\n")
        
        current_app.logger.error(f"Error updating profile: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    




import secrets
import hashlib
from datetime import datetime, timedelta
import pytz
from flask import render_template, request, jsonify, url_for, redirect, flash
from flask_login import current_user, login_required
from werkzeug.security import generate_password_hash

# Philippine Timezone helper
PHILIPPINE_TZ = pytz.timezone('Asia/Manila')

def get_ph_time():
    """Get current Philippine time as naive datetime (for database)"""
    return datetime.now(PHILIPPINE_TZ).replace(tzinfo=None)

# Dictionary to store password reset tokens
password_reset_tokens = {}

@admin_bp.route('/forgot-password', methods=['POST'])
@admin_required
def forgot_password():
    """
    Send password reset email - Admin is already logged in (from settings page)
    Uses reset_token column in admins table
    """
    try:
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Request must be JSON'
            }), 400
        
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({
                'success': False,
                'message': 'Email is required'
            }), 400
        
        if email != current_user.email:
            return jsonify({
                'success': False,
                'message': 'Email does not match current user'
            }), 400
        
        # Generate reset token
        token = secrets.token_urlsafe(32)
        
        # Store token in reset_token column (NO expiry time, just like student!)
        current_user.reset_token = token
        db.session.commit()
        
        # Generate reset URL
        reset_url = url_for('admin.reset_password_page', token=token, _external=True)
        
        # Render email template
        email_html = render_template('admin_password_reset_email.html',
                                   admin_name=f"{current_user.first_name} {current_user.last_name}",
                                   admin_username=current_user.username,
                                   reset_url=reset_url)
        
        # Send email
        from flask_mail import Message
        from extensions import mail
        
        msg = Message(
            'Reset Your Admin Password - CTU COMELEC',
            recipients=[current_user.email]
        )
        msg.html = email_html
        
        mail.send(msg)
        
        # Audit log
        log_audit(
            action='PASSWORD_RESET_REQUESTED',
            description=f"Admin user '{current_user.username}' requested password reset from IP: {request.remote_addr}"
        )
        
        return jsonify({
            'success': True,
            'message': 'Password reset instructions sent to your email!'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error sending password reset: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500




@admin_bp.route('/reset-password/<token>')
def reset_password_page(token):
    """Display password reset page - USES reset_token column (like student)"""
    # Find admin by reset_token
    admin = Admin.query.filter_by(reset_token=token).first()
    
    if not admin:
        return render_template('admin_reset_password.html', 
                             error='Invalid or expired reset link. Please request a new one.')
    
    # Get current secret path for back to login link
    secret_path = AccessCode.get_secret_path()
    
    return render_template('admin_reset_password.html',
                         token=token,
                         admin_username=admin.username,
                         secret_path=secret_path)


@admin_bp.route('/reset-password/<token>', methods=['POST'])
def reset_password(token):
    """Handle password reset form submission - USES reset_token column (like student)"""
    # Find admin by reset_token
    admin = Admin.query.filter_by(reset_token=token).first()
    
    if not admin:
        return render_template('admin_reset_password.html', 
                             error='Invalid or expired reset link. Please request a new one.',
                             token=token)
    
    # Get form data
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    # Validate
    if not password or not confirm_password:
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='All fields are required.')
    
    if password != confirm_password:
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='Passwords do not match.')
    
    # Validate password strength
    if len(password) < 8:
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='Password must be at least 8 characters long.')
    
    if not any(c.isupper() for c in password):
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='Password must contain at least one uppercase letter.')
    
    if not any(c.islower() for c in password):
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='Password must contain at least one lowercase letter.')
    
    if not any(c.isdigit() for c in password):
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error='Password must contain at least one number.')
    
    try:
        # Update password
        from extensions import bcrypt
        admin.password = bcrypt.generate_password_hash(password).decode('utf-8')
        
        # Clear the reset_token (important! same as student)
        admin.reset_token = None
        db.session.commit()
        
        # Audit log
        log_audit(
            action='PASSWORD_RESET_COMPLETED',
            description=f"Admin user '{admin.username}' successfully reset password from IP: {request.remote_addr}"
        )
        
        # Get secret path for login link
        secret_path = AccessCode.get_secret_path()
        
        return render_template('admin_reset_password.html',
                             success='Password reset successfully! You can now login with your new password.',
                             secret_path=secret_path)
        
    except Exception as e:
        current_app.logger.error(f"Error resetting password: {str(e)}")
        return render_template('admin_reset_password.html',
                             token=token,
                             admin_username=admin.username,
                             error=f'Error resetting password: {str(e)}')




from datetime import datetime, timedelta
from flask import jsonify, request, render_template, flash, redirect, url_for, session
from flask_login import login_required, current_user
import secrets
import hashlib

@admin_bp.route('/trusted-devices')
@login_required
def trusted_devices():
    """View all trusted devices for current admin"""
    devices = AdminTrustedDevice.query.filter_by(
        admin_id=current_user.id
    ).order_by(AdminTrustedDevice.last_used.desc()).all()
    
    # Mark current device
    current_fingerprint = session.get('admin_device_fingerprint')
    for device in devices:
        device.is_current = (device.device_fingerprint == current_fingerprint)
    
    # 🚫 REMOVED: Page view audit log - not a data modification
    return render_template('trusted_devices.html', devices=devices)


@admin_bp.route('/trusted-devices/add', methods=['POST'])
@login_required
def add_trusted_device():
    """Add current device as trusted"""
    device_info = AdminTrustedDevice.get_device_info(request)
    
    # 🔥 FIXED: EXCLUDE IP address from fingerprint (IP changes!)
    # OLD (WRONG): fingerprint_data = f"{current_user.id}{device_info['ip_address']}{device_info['user_agent']}{device_info['browser']}{device_info['os']}"
    # NEW (CORRECT) - NO IP ADDRESS:
    fingerprint_data = f"{current_user.id}{device_info['user_agent']}{device_info['browser']}{device_info['os']}"
    device_fingerprint = hashlib.sha256(fingerprint_data.encode()).hexdigest()
    
    print(f"🔑 Add trusted device fingerprint: {device_fingerprint[:20]}...")
    
    # Check if this specific device (by fingerprint) already exists
    existing_device = AdminTrustedDevice.query.filter_by(
        admin_id=current_user.id,
        device_fingerprint=device_fingerprint
    ).first()
    
    if existing_device:
        # Update existing device
        existing_device.trusted = True
        existing_device.expires_at = datetime.utcnow() + timedelta(days=30)
        existing_device.last_used = datetime.utcnow()
        existing_device.device_name = f"{device_info['browser']} on {device_info['os']}"
        existing_device.ip_address = device_info['ip_address']
        existing_device.user_agent = device_info['user_agent']
        existing_device.browser = device_info['browser']
        existing_device.os = device_info['os']
        existing_device.device_type = device_info['device_type']
        
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (Device updated)
        log_audit(
            action='DEVICE_UPDATED',
            description=f"Admin user '{current_user.username}' updated trusted device: {existing_device.device_name} from IP: {request.remote_addr}"
        )
        
        # Set the fingerprint in session
        session['admin_device_fingerprint'] = existing_device.device_fingerprint
        
        return jsonify({
            'success': True, 
            'message': 'Device already trusted and updated!',
            'device_added': True
        })
    else:
        # Create NEW trusted device
        device = AdminTrustedDevice(
            admin_id=current_user.id,
            device_name=f"{device_info['browser']} on {device_info['os']}",
            ip_address=device_info['ip_address'],
            user_agent=device_info['user_agent'],
            browser=device_info['browser'],
            os=device_info['os'],
            device_type=device_info['device_type'],
            trusted=True,
            expires_at=datetime.utcnow() + timedelta(days=30),
            last_used=datetime.utcnow()
        )
        
        # Set the consistent fingerprint (NO IP)
        device.device_fingerprint = device_fingerprint
        
        db.session.add(device)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (Device added)
        log_audit(
            action='DEVICE_ADDED',
            description=f"Admin user '{current_user.username}' added new trusted device: {device.device_name} from IP: {request.remote_addr}"
        )
        
        # Store fingerprint in session
        session['admin_device_fingerprint'] = device.device_fingerprint
        
        return jsonify({
            'success': True, 
            'message': 'Device added to trusted devices!',
            'device_added': True
        })


@admin_bp.route('/trusted-devices/remove/<int:device_id>', methods=['POST'])
@login_required
def request_remove_device(device_id):
    """Request to remove a trusted device (sends confirmation email)"""
    print(f"=== DEBUG: Device removal requested for device ID: {device_id} ===")
    print(f"Current user ID: {current_user.id}")
    print(f"Request method: {request.method}")
    print(f"Request headers: {dict(request.headers)}")
    
    try:
        # Get JSON data if sent
        data = request.get_json(silent=True) or {}
        print(f"Request data: {data}")
        
        device = AdminTrustedDevice.query.filter_by(
            id=device_id,
            admin_id=current_user.id
        ).first()
        
        if not device:
            print(f"DEBUG: Device not found - ID: {device_id}, admin_id: {current_user.id}")
            return jsonify({
                'success': False,
                'message': 'Device not found or already removed'
            }), 404
        
        print(f"DEBUG: Device found - ID: {device.id}, Name: {device.device_name}, Trusted: {device.trusted}")
        print(f"DEBUG: Device fingerprint: {device.device_fingerprint}")
        
        # Check if device is trusted
        if not device.trusted:
            return jsonify({
                'success': False,
                'message': 'Device is not trusted'
            }), 400
        
        # Don't allow removing current device
        current_fingerprint = session.get('admin_device_fingerprint')
        print(f"DEBUG: Current session fingerprint: {current_fingerprint}")
        
        if device.device_fingerprint and current_fingerprint and device.device_fingerprint == current_fingerprint:
            print(f"DEBUG: Cannot remove current device")
            return jsonify({
                'success': False,
                'message': 'Cannot remove your current device'
            }), 400
        
        # Generate removal token
        removal_token = secrets.token_urlsafe(32)
        print(f"DEBUG: Generated token: {removal_token}")
        
        device.verification_token = removal_token
        device.verification_sent_at = datetime.utcnow()
        db.session.commit()
        print(f"DEBUG: Token saved to database for device {device.id}")
        
        # Send confirmation email
        try:
            from admin.utils import send_device_removal_confirmation
            
            # Check if email function exists
            if 'send_device_removal_confirmation' not in dir():
                print(f"DEBUG: Email function not found, creating mock response")
                # For testing, return success without email
                return jsonify({
                    'success': True,
                    'message': 'Device removal initiated (email simulation)',
                    'dev_mode': True
                })
            
            print(f"DEBUG: Attempting to send email to: {current_user.email}")
            send_device_removal_confirmation(current_user, device, removal_token)
            print(f"DEBUG: Email sent successfully")
            
            # 🚫 REMOVED: Audit log for email sending (no DB change yet)
            
            return jsonify({
                'success': True,
                'message': 'Removal confirmation email sent. Please check your inbox.'
            })
            
        except ImportError as e:
            print(f"DEBUG: Email function import error: {str(e)}")
            # For development, return success anyway
            return jsonify({
                'success': True,
                'message': 'Device removal initiated (email disabled in development)',
                'dev_mode': True
            })
            
        except Exception as e:
            print(f"DEBUG: Email sending failed: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Rollback token if email fails
            device.verification_token = None
            device.verification_sent_at = None
            db.session.commit()
            
            return jsonify({
                'success': False,
                'message': f'Failed to send email: {str(e)}'
            }), 500
            
    except Exception as e:
        print(f"DEBUG: Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'Server error occurred'
        }), 500


@admin_bp.route('/trusted-devices/confirm-remove/<token>')
def confirm_remove_device(token):
    """Confirm device removal via email link"""
    device = AdminTrustedDevice.query.filter_by(verification_token=token).first()
    
    if not device:
        return render_template('device_removal_result.html',
                             success=False,
                             message='Invalid or expired confirmation link.')
    
    # Check if token expired (15 minutes)
    if device.verification_sent_at and datetime.utcnow() > device.verification_sent_at + timedelta(minutes=15):
        device.verification_token = None
        device.verification_sent_at = None
        db.session.commit()
        
        return render_template('device_removal_result.html',
                             success=False,
                             message='Confirmation link has expired. Please try again.')
    
    # Don't allow removing current device
    current_fingerprint = session.get('admin_device_fingerprint')
    if device.device_fingerprint == current_fingerprint:
        return render_template('device_removal_result.html',
                             success=False,
                             message='Cannot remove your current device.')
    
    # Remove the device
    device_name = device.device_name
    db.session.delete(device)
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (Device actually removed - data modification)
    username = getattr(current_user, 'username', 'Unknown') if current_user.is_authenticated else 'Unknown'
    ip = request.remote_addr
    
    log_audit(
        action='DEVICE_REMOVED',
        description=f"Admin user removed trusted device '{device_name}' from IP: {ip}"
    )
    
    return render_template('device_removal_result.html',
                         success=True,
                         message='Device has been successfully removed from your trusted devices.')


@admin_bp.route('/trusted-devices/cancel-remove/<token>')
def cancel_remove_device(token):
    """Cancel device removal request"""
    device = AdminTrustedDevice.query.filter_by(verification_token=token).first()
    
    if device:
        # Clear the verification token
        device.verification_token = None
        device.verification_sent_at = None
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (Device removal cancelled)
        log_audit(
            action='DEVICE_REMOVAL_CANCELLED',
            description=f"Admin user cancelled removal for device '{device.device_name}' from IP: {request.remote_addr}"
        )
    
    return render_template('device_removal_result.html',
                         success=True,
                         message='Device removal request has been cancelled. Your device remains trusted.')


@admin_bp.route('/trusted-devices/verify/send', methods=['POST'])
@login_required
def send_device_verification():
    """Send verification email for new device"""
    device_info = AdminTrustedDevice.get_device_info(request)
    
    # Check if device already exists and is trusted
    existing = AdminTrustedDevice.query.filter_by(
        admin_id=current_user.id,
        ip_address=device_info['ip_address'],
        browser=device_info['browser'],
        trusted=True
    ).first()
    
    if existing:
        # Device already trusted, update last used
        existing.last_used = datetime.utcnow()
        existing.expires_at = datetime.utcnow() + timedelta(days=30)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (Device auto-verified)
        log_audit(
            action='DEVICE_AUTO_VERIFIED',
            description=f"Admin user '{current_user.username}' auto-verified existing device: {existing.device_name} from IP: {request.remote_addr}"
        )
        
        # Store fingerprint in session
        session['admin_device_fingerprint'] = existing.device_fingerprint
        
        return jsonify({
            'success': True,
            'trusted': True,
            'message': 'Device already trusted'
        })
    
    # Create or get unverified device
    device = AdminTrustedDevice.query.filter_by(
        admin_id=current_user.id,
        ip_address=device_info['ip_address'],
        browser=device_info['browser'],
        trusted=False
    ).first()
    
    if not device:
        device = AdminTrustedDevice(
            admin_id=current_user.id,
            device_name=f"{device_info['browser']} on {device_info['os']}",
            trusted=False,
            **device_info
        )
        device.generate_fingerprint()
        db.session.add(device)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (New device pending verification)
        log_audit(
            action='DEVICE_PENDING_VERIFICATION',
            description=f"Admin user '{current_user.username}' initiated verification for new device: {device.device_name} from IP: {request.remote_addr}"
        )
    
    # Generate verification token
    token = device.generate_verification_token()
    db.session.commit()
    
    # Send verification email
    from admin.utils import send_admin_device_verification_email
    send_admin_device_verification_email(current_user, device, token)
    
    # 🚫 REMOVED: Audit log for email sending (no DB change)
    
    return jsonify({
        'success': True,
        'trusted': False,
        'message': 'Verification email sent'
    })


@admin_bp.route('/trusted-devices/verify/<token>')
def verify_admin_device(token):
    """Verify device via email link"""
    device = AdminTrustedDevice.query.filter_by(verification_token=token).first()
    
    if not device:
        return render_template('device_verification_result.html', 
                             success=False, 
                             message='Invalid or expired verification link.')
    
    # Check if token expired (15 minutes)
    if device.verification_sent_at and datetime.utcnow() > device.verification_sent_at + timedelta(minutes=15):
        db.session.delete(device)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (Expired verification)
        log_audit(
            action='DEVICE_VERIFICATION_EXPIRED',
            description=f"Device verification expired for: {device.device_name} from IP: {request.remote_addr}"
        )
        
        return render_template('device_verification_result.html',
                             success=False,
                             message='Verification link has expired. Please try again.')
    
    # Mark device as trusted
    device.trusted = True
    device.verification_token = None
    device.expires_at = datetime.utcnow() + timedelta(days=30)
    device.last_used = datetime.utcnow()
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (Device verified)
    admin_username = getattr(device.admin, 'username', 'Unknown')
    log_audit(
        action='DEVICE_VERIFIED',
        description=f"Admin user '{admin_username}' verified new device: {device.device_name} from IP: {request.remote_addr}"
    )
    
    # Store fingerprint in session if this is the current device
    if device.ip_address == request.remote_addr:
        session['admin_device_fingerprint'] = device.device_fingerprint
    
    return render_template('device_verification_result.html',
                         success=True,
                         message='Device verified successfully! You can now use this device without 2FA.')


@admin_bp.route('/trusted-devices/check', methods=['POST'])
def check_device_trust():
    """Check if current device is trusted - used during login verification page"""
    if not current_user.is_authenticated:
        return jsonify({'trusted': False})
    
    device_info = AdminTrustedDevice.get_device_info(request)
    current_fingerprint = session.get('admin_device_fingerprint')
    
    # 🔥 FIXED: Generate fingerprint WITHOUT IP (same as verify_2fa)
    fingerprint_data = f"{current_user.id}{device_info['user_agent']}{device_info['browser']}{device_info['os']}"
    device_fingerprint = hashlib.sha256(fingerprint_data.encode()).hexdigest()
    
    # Check if device is trusted by fingerprint only (NO IP FALLBACK)
    trusted_device = AdminTrustedDevice.query.filter_by(
        admin_id=current_user.id,
        device_fingerprint=device_fingerprint,
        trusted=True
    ).first()
    
    if trusted_device and not trusted_device.is_expired():
        # Update last used
        trusted_device.last_used = datetime.utcnow()
        db.session.commit()
        # Update session fingerprint
        session['admin_device_fingerprint'] = trusted_device.device_fingerprint
        return jsonify({'trusted': True})
    
    return jsonify({'trusted': False})



from student.models import TrustedDevice 

# ---------------------- IMPORT STUDENTS ---------------------- #
@admin_bp.route("/import_students", methods=["GET", "POST"])
def import_students():
    import time
    import pandas as pd
    import numpy as np
    from sqlalchemy import text
    from datetime import datetime
    
    start_time = time.time()

    if request.method == "POST":
        file = request.files.get("excel_file")

        if not file or file.filename == "":
            flash("No file selected.", "import-danger")
            return redirect(url_for("admin.import_students"))

        temp_file_path = None
        try:
            # Save file temporarily to avoid memory issues with large Excel
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                file.save(tmp.name)
                temp_file_path = tmp.name
            
            print(f"📊 STEP 0: File saved in {time.time() - start_time:.2f}s")
            
            # STEP 1: Read Excel with optimized settings
            df = pd.read_excel(
                temp_file_path, 
                dtype={"StudentNo": str},
                usecols=["StudentNo", "LastName", "FirstName", "YearLevel"],
                engine='openpyxl'
            )
            
            # Clean up temp file
            os.unlink(temp_file_path)
            
            print(f"📊 STEP 1: Loaded Excel in {time.time() - start_time:.2f}s")
            
            # STEP 2: Clean data using vectorized operations
            # First, replace all NaN values with None (this is the key fix!)
            df = df.where(pd.notnull(df), None)
            df = df.replace([np.nan], [None])
            
            df['StudentNo'] = df['StudentNo'].astype(str).str.strip()
            df['StudentNo'] = df['StudentNo'].str.replace(r'\.0$', '', regex=True)
            df['FirstName'] = df['FirstName'].astype(str).str.strip()
            df['LastName'] = df['LastName'].astype(str).str.strip()
            
            # Handle YearLevel - convert to proper format, handling None values
            if 'YearLevel' in df.columns:
                # Define a function to safely convert year level
                def convert_year_level(val):
                    if val is None or pd.isna(val) or str(val).lower() in ['nan', 'none', '']:
                        return None
                    
                    # Convert to string and clean
                    val_str = str(val).strip()
                    
                    # Define mapping from numbers to text
                    year_mapping = {
                        '1': '1st Year',
                        '2': '2nd Year', 
                        '3': '3rd Year',
                        '4': '4th Year',
                        '1.0': '1st Year',
                        '2.0': '2nd Year',
                        '3.0': '3rd Year',
                        '4.0': '4th Year'
                    }
                    
                    # Check if it's a number
                    if val_str in year_mapping:
                        return year_mapping[val_str]
                    
                    # Handle numeric values that might be floats
                    try:
                        num_val = float(val_str)
                        if num_val.is_integer():
                            int_val = int(num_val)
                            if int_val == 1:
                                return "1st Year"
                            elif int_val == 2:
                                return "2nd Year"
                            elif int_val == 3:
                                return "3rd Year"
                            elif int_val == 4:
                                return "4th Year"
                            else:
                                return f"{int_val}th Year"
                        else:
                            return None
                    except (ValueError, TypeError):
                        # If it's already a text like "1st Year", keep as is
                        if 'st' in val_str or 'nd' in val_str or 'rd' in val_str or 'th' in val_str:
                            return val_str
                        return None
                
                # Apply the conversion function
                df['YearLevel'] = df['YearLevel'].apply(convert_year_level)
            else:
                df['YearLevel'] = None
            
            # Remove rows with invalid student numbers
            df = df[df['StudentNo'].notna() & (df['StudentNo'] != '') & (df['StudentNo'] != 'nan')]
            
            # Get unique student numbers (remove duplicates)
            df = df.drop_duplicates(subset=['StudentNo'])
            
            excel_student_nos = set(df['StudentNo'].tolist())
            excel_data = df.to_dict('records')
            
            print(f"📊 STEP 2: Cleaned data - {len(excel_data)} unique students in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 1: Use RAW SQL for bulk operations =====
            connection = db.session.connection()
            
            # STEP 3: Get existing CTU students in one query
            existing_result = connection.execute(
                text("SELECT student_number, id, first_name, last_name, year_level FROM ctu_students")
            )
            existing_ctu = {}
            for row in existing_result:
                existing_ctu[row[0]] = {
                    'id': row[1], 
                    'first_name': row[2], 
                    'last_name': row[3],
                    'year_level': row[4]
                }
            
            print(f"📊 STEP 3: Fetched {len(existing_ctu)} CTU students in {time.time() - start_time:.2f}s")
            
            # STEP 4: Get registered students and their vote status in ONE query
            registered_result = connection.execute(
                text("""
                    SELECT s.id, s.id_number, s.year_level_id,
                           CASE WHEN v.student_id IS NOT NULL THEN 1 ELSE 0 END as has_voted
                    FROM students s
                    LEFT JOIN votes v ON s.id = v.student_id
                """)
            )
            registered_map = {}
            students_with_votes = set()
            for row in registered_result:
                registered_map[row[1]] = {'id': row[0], 'year_level_id': row[2], 'has_voted': row[3]}
                if row[3]:
                    students_with_votes.add(row[0])
            
            print(f"📊 STEP 4: Fetched {len(registered_map)} registered students in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 2: Prepare bulk operations =====
            to_insert = []
            to_update = []
            
            # Get year level mapping from database
            year_levels = {yl.year_name: yl.id for yl in YearLevel.query.all()}
            
            for student in excel_data:
                student_no = student['StudentNo']
                first_name = student['FirstName']
                last_name = student['LastName']
                year_level = student.get('YearLevel')
                
                # Ensure year_level is None if it's invalid (not a string)
                if year_level and not isinstance(year_level, str):
                    year_level = None
                
                if student_no in existing_ctu:
                    # Check if any field changed
                    existing = existing_ctu[student_no]
                    if (existing['first_name'] != first_name or 
                        existing['last_name'] != last_name or
                        existing['year_level'] != year_level):
                        to_update.append({
                            'student_number': student_no,
                            'first_name': first_name,
                            'last_name': last_name,
                            'year_level': year_level
                        })
                else:
                    to_insert.append({
                        'student_number': student_no,
                        'first_name': first_name,
                        'last_name': last_name,
                        'year_level': year_level
                    })
            
            print(f"📊 STEP 5: Prepared {len(to_insert)} inserts, {len(to_update)} updates in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 3: BULK INSERT using executemany =====
            if to_insert:
                chunk_size = 1000
                for i in range(0, len(to_insert), chunk_size):
                    chunk = to_insert[i:i+chunk_size]
                    
                    # Prepare data for insertion - ensure year_level is None not nan
                    insert_data = []
                    for s in chunk:
                        insert_record = {
                            'student_number': s['student_number'],
                            'first_name': s['first_name'],
                            'last_name': s['last_name'],
                            'year_level': s['year_level'] if s['year_level'] is not None else None
                        }
                        insert_data.append(insert_record)
                    
                    # Use SQLAlchemy's bulk insert
                    db.session.bulk_insert_mappings(CtuStudent, insert_data)
                    db.session.flush()
                    
                    print(f"   Inserted chunk {i//chunk_size + 1}/{(len(to_insert)-1)//chunk_size + 1}")
            
            print(f"📊 STEP 6: Completed inserts in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 4: BULK UPDATE using SQLAlchemy =====
            if to_update:
                chunk_size = 1000
                for i in range(0, len(to_update), chunk_size):
                    chunk = to_update[i:i+chunk_size]
                    
                    for student_data in chunk:
                        # Update each student individually but in same transaction
                        update_data = {
                            'first_name': student_data['first_name'],
                            'last_name': student_data['last_name']
                        }
                        if student_data['year_level'] is not None and student_data['year_level']:
                            update_data['year_level'] = student_data['year_level']
                        else:
                            update_data['year_level'] = None
                            
                        db.session.query(CtuStudent)\
                            .filter(CtuStudent.student_number == student_data['student_number'])\
                            .update(update_data, synchronize_session=False)
                    
                    db.session.flush()
                    print(f"   Updated chunk {i//chunk_size + 1}/{(len(to_update)-1)//chunk_size + 1}")
            
            print(f"📊 STEP 7: Completed updates in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 5: Update year levels in students table =====
            year_level_updates = 0
            for student_no, ctu_data in existing_ctu.items():
                if student_no in excel_student_nos and student_no in registered_map:
                    # Find the student's year level from Excel
                    excel_row = next((s for s in excel_data if s['StudentNo'] == student_no), None)
                    if excel_row and excel_row.get('YearLevel'):
                        year_level_str = excel_row['YearLevel']
                        if year_level_str and year_level_str in year_levels:
                            new_year_level_id = year_levels[year_level_str]
                            if registered_map[student_no]['year_level_id'] != new_year_level_id:
                                # Update student's year level
                                db.session.query(Student)\
                                    .filter(Student.id == registered_map[student_no]['id'])\
                                    .update({'year_level_id': new_year_level_id}, synchronize_session=False)
                                year_level_updates += 1
            
            if year_level_updates > 0:
                db.session.flush()
                print(f"📊 STEP 7.5: Updated {year_level_updates} students' year levels in {time.time() - start_time:.2f}s")

            # ===== OPTIMIZATION 6: Handle deletions =====
            to_delete_ctu = []
            to_delete_registered = []
            
            for student_no, existing in existing_ctu.items():
                if student_no not in excel_student_nos:
                    to_delete_ctu.append(student_no)
                    
                    # Check if registered
                    if student_no in registered_map and not registered_map[student_no]['has_voted']:
                        to_delete_registered.append(registered_map[student_no]['id'])
            
            deleted_from_ctu = len(to_delete_ctu)
            deleted_from_registration = 0
            
            # Bulk delete CTU students
            if to_delete_ctu:
                chunk_size = 500
                for i in range(0, len(to_delete_ctu), chunk_size):
                    chunk = to_delete_ctu[i:i+chunk_size]
                    db.session.query(CtuStudent)\
                        .filter(CtuStudent.student_number.in_(chunk))\
                        .delete(synchronize_session=False)
                    db.session.flush()
            
            print(f"📊 STEP 8: Deleted {deleted_from_ctu} CTU students in {time.time() - start_time:.2f}s")
            
            # Bulk delete registered students (with no votes)
            if to_delete_registered:
                # First delete TrustedDevice records
                chunk_size = 500
                for i in range(0, len(to_delete_registered), chunk_size):
                    chunk = to_delete_registered[i:i+chunk_size]
                    from student.models import TrustedDevice
                    db.session.query(TrustedDevice)\
                        .filter(TrustedDevice.student_id.in_(chunk))\
                        .delete(synchronize_session=False)
                    db.session.flush()
                
                # Then delete students
                for i in range(0, len(to_delete_registered), chunk_size):
                    chunk = to_delete_registered[i:i+chunk_size]
                    db.session.query(Student)\
                        .filter(Student.id.in_(chunk))\
                        .delete(synchronize_session=False)
                    db.session.flush()
                
                deleted_from_registration = len(to_delete_registered)
            
            print(f"📊 STEP 9: Deleted {deleted_from_registration} registered students in {time.time() - start_time:.2f}s")
            
            # ===== OPTIMIZATION 7: Commit once at the end =====
            db.session.commit()
            
            total_time = time.time() - start_time
            print(f"✅ TOTAL TIME: {total_time:.2f} seconds for {len(excel_data)} students")
            
            # Audit log
            username = getattr(current_user, 'username', 'Unknown') if current_user.is_authenticated else 'Unknown'
            ip = request.remote_addr
            filename = file.filename if file else 'Unknown'
            
            log_audit(
                action='IMPORT_STUDENTS',
                description=f"Admin user '{username}' imported students from '{filename}' from IP: {ip} | Imported: {len(to_insert)}, Updated: {len(to_update)}, Year Level Updates: {year_level_updates}, Deleted from CTU: {deleted_from_ctu}, Deleted registered: {deleted_from_registration}, Time: {total_time:.2f}s"
            )

            flash(
                f"✅ Sync complete in {total_time:.1f}s!\n"
                f"📥 Imported: {len(to_insert)}\n"
                f"🔄 Updated: {len(to_update)}\n"
                f"📅 Year Levels Updated: {year_level_updates}\n"
                f"🗑️ Removed from CTU list: {deleted_from_ctu}\n"
                f"🚫 Removed registered students: {deleted_from_registration}",
                "import-success"
            )
            
            return redirect(url_for("admin.import_students"))

        except Exception as e:
            db.session.rollback()
            print(f"❌ ERROR: {str(e)}")
            print(f"❌ ERROR after {time.time() - start_time:.2f} seconds")
            import traceback
            traceback.print_exc()
            
            # Clean up temp file if it exists
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            
            flash(f"Error importing file: {str(e)}", "import-danger")
            return redirect(url_for("admin.import_students"))

    # ------------------ GET ------------------
    page = request.args.get("page", 1, type=int)
    search_query = request.args.get("q", "", type=str)

    students_query = CtuStudent.query

    if search_query:
        search = f"%{search_query.strip()}%"
        students_query = students_query.filter(
            or_(
                CtuStudent.student_number.ilike(search),
                CtuStudent.first_name.ilike(search),
                CtuStudent.last_name.ilike(search)
            )
        )

    students = students_query.order_by(CtuStudent.last_name.asc()) \
        .paginate(page=page, per_page=20, error_out=False)

    total_students = CtuStudent.query.count()

    return render_template(
        "import_students.html",
        students=students,
        total_students=total_students
    )


# Add this import with your other imports at the top of admin/routes.py
from admin.utils import sync_registered_students_with_ctu
@admin_bp.route('/sync-registered-students', methods=['POST'])
@admin_required
def sync_registered_students():
    """Manual sync endpoint to update registered students' year levels from CTU list"""
    try:
        from sqlalchemy import text
        
        # Get year level mapping
        year_levels = {yl.year_name: yl.id for yl in YearLevel.query.all()}
        
        # Get all registered students with their ID numbers
        registered_students = Student.query.filter(Student.id_number.isnot(None)).all()
        
        updates = 0
        kept_with_votes = 0
        
        for student in registered_students:
            # Find this student in CTU list
            ctu_student = CtuStudent.query.filter_by(student_number=student.id_number).first()
            
            if ctu_student and ctu_student.year_level:
                # Check if student has voted
                has_voted = Vote.query.filter_by(student_id=student.id).first() is not None
                
                if has_voted:
                    kept_with_votes += 1
                
                # Update year level if it exists in our mapping
                if ctu_student.year_level in year_levels:
                    new_year_level_id = year_levels[ctu_student.year_level]
                    if student.year_level_id != new_year_level_id:
                        student.year_level_id = new_year_level_id
                        updates += 1
        
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='SYNC_REGISTERED_STUDENTS',
            description=f"Admin user '{username}' manually synced registered students from IP: {ip} | Year Level Updates: {updates}, Students with votes: {kept_with_votes}"
        )
        
        return jsonify({
            'success': True,
            'updates': updates,
            'kept_with_votes': kept_with_votes
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route("/import_students_table")
def import_students_table():
    from flask import request, render_template
    from sqlalchemy import or_
    from admin.models import CtuStudent

    try:
        page = request.args.get("page", 1, type=int)
        search_query = request.args.get("q", "", type=str)

        students_query = CtuStudent.query

        if search_query:
            search = f"%{search_query.strip()}%"
            students_query = students_query.filter(
                or_(
                    CtuStudent.student_number.ilike(search),
                    CtuStudent.first_name.ilike(search),
                    CtuStudent.last_name.ilike(search)
                )
            )

        # Prevent paginate error if page > total_pages
        students = students_query.order_by(CtuStudent.last_name.asc()) \
                                 .paginate(page=page, per_page=20, error_out=False)

        # Ensure string comparison
        student_numbers = [str(s.student_number) for s in students.items]
        if student_numbers:
            registered_students = Student.query.filter(
                Student.id_number.in_(student_numbers)
            ).all()
            registered_numbers = set(str(s.id_number) for s in registered_students)
        else:
            registered_numbers = set()

        # 🚫 REMOVED: IMPORT_STUDENTS_TABLE_VIEW audit log (AJAX table refresh - not a data modification)

        return render_template(
            "partials/_students_table.html",
            students=students,
            registered_numbers=registered_numbers
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        
        # ✅ KEEP THIS AUDIT LOG (Error occurred - but only log once, not on every refresh)
        # Only log if this is a POST request or first load, not on every AJAX refresh
        if request.method == "POST" or request.args.get("error_logged") != "true":
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            
            log_audit(
                action='IMPORT_STUDENTS_TABLE_ERROR',
                description=f"Admin user '{username}' encountered error viewing import students table from IP: {ip} | Error: {str(e)[:100]}..."
            )
        
        # Return a simple HTML message instead of plain text, for AJAX
        return render_template(
            "partials/_students_table.html",
            students=None,
            registered_numbers=set()
        )
    

@admin_bp.route('/students')
@admin_required
def manage_students():
    # Get year filter from session (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range for elections filtering
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    # -------------------- Fetch Departments & Courses -------------------- #
    departments = Department.query.order_by(Department.name).all()
    courses = Course.query.order_by(Course.course_name).all()

    departments_data = [{"id": d.id, "name": d.name} for d in departments]
    courses_data = [{"id": c.id, "name": c.course_name} for c in courses]

    # -------------------- Fetch Elections with scope - FILTERED BY YEAR -------------------- #
    election_query = Election.query.order_by(Election.start_date.desc())
    
    # Apply year filter to elections ONLY
    if start_date and end_date:
        election_query = election_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    elections = election_query.all()
    elections_data = [{
        "id": e.id, 
        "title": e.title,
        "scope": e.scope,
        "year_levels": e.year_levels  # Include year_levels for display
    } for e in elections]

    # -------------------- Render Template -------------------- #
    return render_template(
        'manage_students.html',
        departments=departments_data,
        courses=courses_data,
        elections=elections_data,
        current_year=year  # Changed from current_sy to current_year
    )


from sqlalchemy import or_, func

# ---------------------- AJAX STUDENT DATA (with voting status) ---------------------- #
@admin_bp.route('/students/data')
@admin_required
def students_data():
    filter_type = request.args.get('filter_type', 'all')
    filter_id = request.args.get('filter_id')
    search = request.args.get('search', '')
    election_id = request.args.get('election_id')  # Selected election
    page = int(request.args.get('page', 1))
    per_page = 10

    query = Student.query

    # Apply filters
    if filter_type == 'department' and filter_id:
        query = query.filter(Student.department_id == int(filter_id))

    if filter_type == 'course' and filter_id:
        query = query.filter(Student.course_id == int(filter_id))

    # Apply search
    if search:
        query = query.filter(
            or_(
                Student.id_number.ilike(f'%{search}%'),
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                func.coalesce(Student.course, '').ilike(f'%{search}%')
            )
        )

    # Get election details if selected
    election = None
    if election_id and election_id != "all":
        election = Election.query.get(int(election_id))
        
        # 🎯 Filter students based on election's year levels (but not by school year)
        if election and election.scope == 'campus' and election.year_levels:
            # Join with year_level to filter by year
            if election.year_levels != 'all':
                # Get list of allowed year levels
                allowed_years = election.year_levels.split(',')
                
                # Filter students whose year_level_id matches allowed years
                query = query.filter(Student.year_level_id.in_(allowed_years))
            
        # 🎯 For department elections, filter by department
        elif election and election.scope == 'department' and election.department_id:
            query = query.filter(Student.department_id == election.department_id)

    # Pagination
    pagination = query.order_by(Student.last_name).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    # Convert students to simple dicts
    students = []
    for s in pagination.items:
        # Check voting status if election_id is provided
        has_voted = False
        if election_id and election_id != "all":
            has_voted = Vote.query.filter_by(student_id=s.id, election_id=int(election_id)).first() is not None

        # Get year level name if available
        year_level_name = ""
        if s.year_level:
            year_level_name = s.year_level.year_name

        students.append({
            "id": s.id,
            "id_number": s.id_number,
            "first_name": s.first_name,
            "last_name": s.last_name,
            "course": s.course,
            "year_level": year_level_name,
            "year_level_id": s.year_level_id,
            "has_voted": has_voted
        })

    # Log the filter info for debugging
    current_app.logger.info(f"Students data - Election: {election_id}, Total students: {len(students)}")

    # 🚫 REMOVED: STUDENTS_DATA_VIEW audit log (AJAX data fetch - not a data modification)

    return jsonify({
        "students": students,
        "total_pages": pagination.pages,
        "current_page": pagination.page
    })
   

@admin_bp.route('/students/export-excel')
@admin_required
def export_students_excel():
    # ------------------- GET FILTERS ------------------- #
    filter_type = request.args.get('filter_type', 'all')
    filter_id = request.args.get('filter_id')
    search = request.args.get('search', '')
    election_id = request.args.get('election_id')

    # ------------------- QUERY STUDENTS ------------------- #
    query = Student.query

    if filter_type == 'department' and filter_id:
        query = query.filter(Student.department_id == int(filter_id))
    if filter_type == 'course' and filter_id:
        query = query.filter(Student.course_id == int(filter_id))
    if search:
        query = query.filter(
            or_(
                Student.id_number.ilike(f'%{search}%'),
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                Student.course.ilike(f'%{search}%')
            )
        )

    # Get election details if selected
    election = None
    if election_id and election_id != "all":
        election = Election.query.get(int(election_id))
        
        # 🎯 NEW: Filter students based on election's year levels
        if election and election.scope == 'campus' and election.year_levels:
            if election.year_levels != 'all':
                # Filter by allowed year levels
                allowed_years = election.year_levels.split(',')
                query = query.filter(Student.year_level_id.in_(allowed_years))
        
        # 🎯 NEW: For department elections, filter by department
        elif election and election.scope == 'department' and election.department_id:
            query = query.filter(Student.department_id == election.department_id)

    students = query.order_by(Student.last_name).all()

    # ------------------- FETCH SELECTED ELECTION ------------------- #
    election = None
    if election_id and election_id != "all":
        election = Election.query.get(int(election_id))

    # ------------------- CREATE EXCEL ------------------- #
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # ------------------- HEADER STYLES ------------------- #
    mustard_fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # FIXED: Add proper text alignment for data cells
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # FIXED: Add indent to create padding/gap
    left_alignment_with_indent = Alignment(horizontal='left', vertical='center', indent=2)

    max_col = 6  # A–F

    # ------------------- HEADER ROWS ------------------- #
    # Row 1: Election Title with Year Info
    election_title = f"{election.title if election else 'All Elections'}"
    if election and election.scope == 'campus' and election.year_levels:
        if election.year_levels == 'all':
            election_title += " (All Years)"
        else:
            election_title += f" (Year {election.year_levels})"
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value=election_title)
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = mustard_fill
    ws.cell(row=1, column=1).alignment = center_alignment

    # Row 2: Department / All
    department_text = "-"
    if filter_type == "department" and filter_id:
        dept_obj = Department.query.get(int(filter_id))
        if dept_obj:
            department_text = dept_obj.name
    elif election and election.department:
        department_text = election.department
    elif election and election.scope == 'campus':
        department_text = "Campus-Wide"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value=f"Department: {department_text}")
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).fill = mustard_fill
    ws.cell(row=2, column=1).alignment = center_alignment

    # Row 3: Course / All
    course_text = "-"
    if filter_type == "course" and filter_id:
        course_obj = Course.query.get(int(filter_id))
        if course_obj:
            course_text = course_obj.course_name
    elif election and election.course_rel:
        course_text = election.course_rel.course_name

    # Add year level info to row 3
    if election and election.scope == 'campus' and election.year_levels:
        if election.year_levels == 'all':
            course_text += " | Target: All Year Levels"
        else:
            course_text += f" | Target: Year {election.year_levels}"

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws.cell(row=3, column=1, value=f"Course: {course_text}")
    ws.cell(row=3, column=1).font = header_font
    ws.cell(row=3, column=1).fill = mustard_fill
    ws.cell(row=3, column=1).alignment = center_alignment

    # ------------------- COLUMN HEADERS ------------------- #
    headers = ["Student ID", "First Name", "Last Name", "Course", "Year Level", "Status"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2980b9', end_color='2980b9', fill_type='solid')
        cell.alignment = center_alignment

    # ------------------- ADD STUDENTS ------------------- #
    # FIXED: Store column data to calculate proper widths
    column_data = {i: [] for i in range(1, max_col + 1)}
    
    for idx, s in enumerate(students, start=5):
        has_voted = False
        if election:
            has_voted = Vote.query.filter_by(student_id=s.id, election_id=election.id).first() is not None
        status_text = "Voted" if has_voted else "Not Voted"

        # Get year level name
        year_level_name = ""
        if s.year_level:
            year_level_name = s.year_level.year_name

        # FIXED: Clean up student ID - remove any extra spaces
        student_id = str(s.id_number).strip() if s.id_number else ""
        
        row_values = [
            student_id,
            s.first_name.strip() if s.first_name else "",
            s.last_name.strip() if s.last_name else "",
            s.course.strip() if s.course else "",
            year_level_name,
            status_text
        ]

        for col_num, value in enumerate(row_values, start=1):
            cell = ws.cell(row=idx, column=col_num, value=value)
            
            # FIXED: Use left alignment with indent for all data cells
            if col_num == 6:  # Status column
                cell.alignment = center_alignment
                if has_voted:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006')
            else:
                # FIXED: Add indent (padding) to all data cells
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
            
            # Store data for width calculation
            column_data[col_num].append(str(value) if value else "")

    # ------------------- AUTO FIT COLUMNS WITH EXTRA GAP ------------------- #
    from openpyxl.utils import get_column_letter
    
    # FIXED: Calculate column widths with proper padding
    for col_num in range(1, max_col + 1):
        # Get column letter
        col_letter = get_column_letter(col_num)
        
        # Find the maximum length in this column (including header)
        max_length = len(str(headers[col_num - 1]))  # Start with header length
        
        for value in column_data[col_num]:
            if value:
                # FIXED: Use proper length calculation without counting extra spaces
                value_length = len(str(value))
                if value_length > max_length:
                    max_length = value_length
        
        # FIXED: Set width with proper padding (max_length + 3 for better spacing)
        # Add 3-5 characters of padding for comfortable reading
        adjusted_width = max_length + 5
        
        # Set a minimum width of 12 for Student ID and other columns
        if col_num == 1:  # Student ID column
            adjusted_width = max(adjusted_width, 15)  # Minimum 15 for student ID
        elif col_num == 4:  # Course column
            adjusted_width = max(adjusted_width, 20)  # Minimum 20 for course names
        elif col_num == 5:  # Year Level column
            adjusted_width = max(adjusted_width, 12)  # Minimum 12 for year level
        
        # Cap at maximum 50 to avoid overly wide columns
        adjusted_width = min(adjusted_width, 50)
        
        ws.column_dimensions[col_letter].width = adjusted_width

    # ✅ KEEP THIS AUDIT LOG (EXPORT - creates file, counts as action)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    student_count = len(students)
    election_name = election.title if election else 'All Elections'
    
    log_audit(
        action='EXPORT_STUDENTS_EXCEL',
        description=f"Admin user '{username}' exported {student_count} students to Excel from IP: {ip} | Election: {election_name}, Filter: {filter_type}, Search: '{search}'"
    )

    # ------------------- RETURN EXCEL ------------------- #
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )

    
# ---------------------- DELETE STUDENT ---------------------- #
@admin_bp.route('/students/delete/<int:id>', methods=['POST'])
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    
    # Store student info before deletion for audit log
    student_name = f"{student.first_name} {student.last_name}"
    student_id_number = student.id_number
    
    db.session.delete(student)
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (DELETE - data modification)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    
    log_audit(
        action='DELETE_STUDENT',
        description=f"Admin user '{username}' deleted student: {student_name} (ID: {student_id_number}) from IP: {ip}"
    )
    
    return jsonify({"success": True})


# ==================== ALL REGISTERED STUDENTS PAGE ====================
@admin_bp.route('/all-registered-students')
@admin_required
def all_registered_students():
    """Display all registered students"""
    
    # Get year filter from session (set by dashboard) - for consistency across pages
    year = session.get('admin_current_year')
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    students = Student.query.order_by(Student.last_name).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get registered student numbers
    ctu_students = CtuStudent.query.all()
    registered_numbers = {s.student_number for s in ctu_students if s.student_number}
    
    # Get filter dropdown data
    departments = Department.query.all()
    courses = Course.query.all()
    year_levels = YearLevel.query.all()
    
    # 🚫 REMOVED: Page view audit log (not a data modification)
    
    return render_template(
        'all_registered_students.html',
        students=students,
        departments=[{"id": d.id, "name": d.name} for d in departments],
        courses=[{"id": c.id, "name": c.course_name, "department_id": c.department_id} for c in courses],
        year_levels=[{"id": y.id, "year_name": y.year_name} for y in year_levels],
        registered_numbers=registered_numbers,
        current_year=year  # Changed from current_sy to current_year
    )


# ==================== AJAX DATA ====================
@admin_bp.route('/all-students-data')
@admin_required
def all_students_data():
    """AJAX endpoint for filtered students"""
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    department_id = request.args.get('department_id', '')
    course_id = request.args.get('course_id', '')
    year_level_id = request.args.get('year_level_id', '')
    search = request.args.get('search', '')
    
    query = Student.query
    
    if department_id:
        query = query.filter(Student.department_id == int(department_id))
    if course_id:
        query = query.filter(Student.course_id == int(course_id))
    if year_level_id:
        query = query.filter(Student.year_level_id == int(year_level_id))
    if search:
        query = query.filter(
            or_(
                Student.id_number.ilike(f'%{search}%'),
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%')
            )
        )
    
    # Get total count for the badge
    total_count = query.count()
    
    students = query.order_by(Student.last_name).paginate(page=page, per_page=per_page)
    
    ctu_students = CtuStudent.query.all()
    registered_numbers = {s.student_number for s in ctu_students if s.student_number}
    
    html = render_template(
        'partials/students_table.html',
        students=students,
        registered_numbers=registered_numbers
    )
    
    # 🚫 REMOVED: AJAX data fetch audit log (not a data modification)
    
    return jsonify({
        'html': html,
        'current_page': students.page,
        'total_pages': students.pages,
        'total_count': total_count  # ← ADD THIS LINE
    })


# ==================== DELETE STUDENT ====================
@admin_bp.route('/delete-all-student/<int:id>', methods=['POST'])
@admin_required
def delete_all_student(id):
    """Delete a student permanently - with vote anonymization (same as request deletion)"""
    
    student = Student.query.get_or_404(id)
    
    # Store student info before deletion for audit log
    student_name = f"{student.first_name} {student.last_name}"
    student_id_number = student.id_number
    
    try:
        # MANUALLY DELETE ALL RELATED RECORDS FIRST
        from student.models import Vote, TrustedDevice, DeletionRequest, QualifiedCandidate, PendingCandidate
        
        # Delete in correct order to avoid foreign key issues
        print(f"Processing student deletion for ID: {id}")
        
        # 1. Delete pending candidates
        pending = PendingCandidate.query.filter_by(student_id=id).all()
        for p in pending:
            db.session.delete(p)
        print(f"Deleted {len(pending)} pending candidates")
        
        # 2. Delete qualified candidates
        qualified = QualifiedCandidate.query.filter_by(student_id=id).all()
        for q in qualified:
            db.session.delete(q)
        print(f"Deleted {len(qualified)} qualified candidates")
        
        # 3. Delete deletion requests
        requests = DeletionRequest.query.filter_by(student_id=id).all()
        for r in requests:
            db.session.delete(r)
        print(f"Deleted {len(requests)} deletion requests")
        
        # 4. Delete trusted devices
        devices = TrustedDevice.query.filter_by(student_id=id).all()
        for d in devices:
            db.session.delete(d)
        print(f"Deleted {len(devices)} trusted devices")
        
        # 5. ANONYMIZE VOTES (instead of deleting them) - Same as request deletion
        votes = Vote.query.filter_by(student_id=id).all()
        vote_count = len(votes)
        for vote in votes:
            vote.original_student_id = id  # Store original student ID for audit
            vote.anonymized_at = datetime.utcnow()  # Record when anonymized
            vote.student_id = None  # Remove link to student
        print(f"Anonymized {vote_count} votes (student_id set to NULL, original_student_id preserved)")
        
        # 6. Finally delete the student
        db.session.delete(student)
        
        # Commit all changes
        db.session.commit()
        print(f"Successfully deleted student ID: {id}")
        
        # ✅ KEEP THIS AUDIT LOG (PERMANENT DELETE - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='DELETE_STUDENT_PERMANENT',
            description=f"Admin user '{username}' permanently deleted student: {student_name} (ID: {student_id_number}) from IP: {ip} | Anonymized {vote_count} votes, removed related records: {len(pending)} pending, {len(qualified)} qualified, {len(requests)} requests, {len(devices)} devices"
        )
        
        return jsonify({"success": True, "message": f"Student deleted successfully. {vote_count} votes were anonymized."})
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR deleting student {id}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "success": False, 
            "message": str(e)
        }), 500

    
from student.models import DeletionRequest
@admin_bp.route('/deletion-requests')
def account_deletion_requests():
    """Render the account deletion requests page"""
    # Get filters from session (set by dashboard)
    start_date = session.get('admin_start_date')
    end_date = session.get('admin_end_date')
    year = session.get('admin_current_year')  
    
    # Pass filters to template
    return render_template(
        'account_deletion_requests.html',
        start_date=start_date,
        end_date=end_date,
        current_year=year  
    )


@admin_bp.route('/deletion-requests/data')
def get_deletion_requests_data():
    """Get paginated deletion requests data including both pending (from deletion_requests) 
    and processed requests (from deletion_request_audit)"""
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', 'all')
    search = request.args.get('search', '')
    date = request.args.get('date', '')
    
    # Get GLOBAL filters from session (set by dashboard)
    start_date_filter = session.get('admin_start_date')
    end_date_filter = session.get('admin_end_date')
    year = session.get('admin_current_year')  
    
    per_page = 10
    
    # Parse date range from session
    start_date_obj = None
    end_date_obj = None
    if start_date_filter and end_date_filter:
        try:
            start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError):
            pass
    
    # If no explicit date range, try year filter
    if not start_date_obj and year:
        try:
            year_int = int(year)
            start_date_obj = datetime(year_int, 1, 1)
            end_date_obj = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            pass
    
    # Handle different status filters
    if status == 'pending':
        # Only from deletion_requests table (pending requests)
        query = DeletionRequest.query.filter_by(status='pending')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            query = query.filter(
                DeletionRequest.request_date >= start_date_obj,
                DeletionRequest.request_date <= end_date_obj
            )
        
        if search:
            query = query.join(Student).filter(
                db.or_(
                    Student.first_name.ilike(f'%{search}%'),
                    Student.last_name.ilike(f'%{search}%')
                )
            )
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            query = query.filter(
                db.func.date(DeletionRequest.request_date) == date_obj.date()
            )
        
        paginated = query.order_by(DeletionRequest.request_date.desc()).paginate(page=page, per_page=per_page)
        
        requests_data = []
        for req in paginated.items:
            requests_data.append({
                'id': req.id,
                'student_name': f"{req.student.first_name} {req.student.last_name}",
                'reason': req.reason,
                'request_date': req.request_date.isoformat(),
                'status': req.status,
                'from_audit': False
            })
        
        return jsonify({
            'requests': requests_data,
            'total_pages': paginated.pages,
            'current_page': page
        })
    
    elif status == 'approved':
        # ✅ FIXED: Get from deletion_request_audit table with status 'approved'
        query = DeletionRequestAudit.query.filter_by(status='approved')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            query = query.filter(
                DeletionRequestAudit.request_date >= start_date_obj,
                DeletionRequestAudit.request_date <= end_date_obj
            )
        
        if search:
            query = query.filter(DeletionRequestAudit.student_name.ilike(f'%{search}%'))
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            query = query.filter(
                db.func.date(DeletionRequestAudit.request_date) == date_obj.date()
            )
        
        paginated = query.order_by(DeletionRequestAudit.processed_date.desc()).paginate(page=page, per_page=per_page)
        
        requests_data = []
        for audit in paginated.items:
            requests_data.append({
                'id': audit.id,
                'original_request_id': audit.original_request_id,
                'student_name': audit.student_name,
                'reason': audit.reason,
                'request_date': audit.request_date.isoformat(),
                'status': audit.status,
                'processed_date': audit.processed_date.isoformat(),
                'processed_by': audit.processed_by_username,
                'votes_anonymized': audit.votes_anonymized,
                'admin_notes': audit.admin_notes,
                'from_audit': True
            })
        
        return jsonify({
            'requests': requests_data,
            'total_pages': paginated.pages,
            'current_page': page
        })
    
    elif status == 'rejected':
        # ✅ FIXED: Get from deletion_request_audit table with status 'rejected'
        query = DeletionRequestAudit.query.filter_by(status='rejected')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            query = query.filter(
                DeletionRequestAudit.request_date >= start_date_obj,
                DeletionRequestAudit.request_date <= end_date_obj
            )
        
        if search:
            query = query.filter(DeletionRequestAudit.student_name.ilike(f'%{search}%'))
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            query = query.filter(
                db.func.date(DeletionRequestAudit.request_date) == date_obj.date()
            )
        
        paginated = query.order_by(DeletionRequestAudit.processed_date.desc()).paginate(page=page, per_page=per_page)
        
        requests_data = []
        for audit in paginated.items:
            requests_data.append({
                'id': audit.id,
                'original_request_id': audit.original_request_id,
                'student_name': audit.student_name,
                'reason': audit.reason,
                'request_date': audit.request_date.isoformat(),
                'status': audit.status,
                'processed_date': audit.processed_date.isoformat(),
                'processed_by': audit.processed_by_username,
                'votes_anonymized': audit.votes_anonymized,
                'admin_notes': audit.admin_notes,
                'from_audit': True
            })
        
        return jsonify({
            'requests': requests_data,
            'total_pages': paginated.pages,
            'current_page': page
        })
    
    else:  # status == 'all' - combine BOTH tables (pending + approved + rejected)
        all_requests = []
        
        # 1. Get pending requests from deletion_requests table
        pending_query = DeletionRequest.query.filter_by(status='pending')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            pending_query = pending_query.filter(
                DeletionRequest.request_date >= start_date_obj,
                DeletionRequest.request_date <= end_date_obj
            )
        
        if search:
            pending_query = pending_query.join(Student).filter(
                db.or_(
                    Student.first_name.ilike(f'%{search}%'),
                    Student.last_name.ilike(f'%{search}%')
                )
            )
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            pending_query = pending_query.filter(
                db.func.date(DeletionRequest.request_date) == date_obj.date()
            )
        
        pending_requests = pending_query.all()
        
        # Format pending requests
        for req in pending_requests:
            all_requests.append({
                'id': req.id,
                'student_name': f"{req.student.first_name} {req.student.last_name}",
                'reason': req.reason,
                'request_date': req.request_date,
                'status': req.status,
                'from_audit': False,
                'sort_date': req.request_date
            })
        
        # 2. Get approved requests from audit table
        approved_query = DeletionRequestAudit.query.filter_by(status='approved')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            approved_query = approved_query.filter(
                DeletionRequestAudit.request_date >= start_date_obj,
                DeletionRequestAudit.request_date <= end_date_obj
            )
        
        if search:
            approved_query = approved_query.filter(DeletionRequestAudit.student_name.ilike(f'%{search}%'))
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            approved_query = approved_query.filter(
                db.func.date(DeletionRequestAudit.request_date) == date_obj.date()
            )
        
        approved_requests = approved_query.all()
        
        # Format approved requests
        for audit in approved_requests:
            all_requests.append({
                'id': audit.id,
                'original_request_id': audit.original_request_id,
                'student_name': audit.student_name,
                'reason': audit.reason,
                'request_date': audit.request_date,
                'status': audit.status,
                'processed_date': audit.processed_date,
                'processed_by': audit.processed_by_username,
                'votes_anonymized': audit.votes_anonymized,
                'admin_notes': audit.admin_notes,
                'from_audit': True,
                'sort_date': audit.request_date
            })
        
        # 3. Get rejected requests from audit table
        rejected_query = DeletionRequestAudit.query.filter_by(status='rejected')
        
        # Apply GLOBAL date range filter from dashboard
        if start_date_obj and end_date_obj:
            rejected_query = rejected_query.filter(
                DeletionRequestAudit.request_date >= start_date_obj,
                DeletionRequestAudit.request_date <= end_date_obj
            )
        
        if search:
            rejected_query = rejected_query.filter(DeletionRequestAudit.student_name.ilike(f'%{search}%'))
        if date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            rejected_query = rejected_query.filter(
                db.func.date(DeletionRequestAudit.request_date) == date_obj.date()
            )
        
        rejected_requests = rejected_query.all()
        
        # Format rejected requests
        for audit in rejected_requests:
            all_requests.append({
                'id': audit.id,
                'original_request_id': audit.original_request_id,
                'student_name': audit.student_name,
                'reason': audit.reason,
                'request_date': audit.request_date,
                'status': audit.status,
                'processed_date': audit.processed_date,
                'processed_by': audit.processed_by_username,
                'votes_anonymized': audit.votes_anonymized,
                'admin_notes': audit.admin_notes,
                'from_audit': True,
                'sort_date': audit.request_date
            })
        
        # Sort all requests by date (newest first)
        all_requests.sort(key=lambda x: x['sort_date'], reverse=True)
        
        # Paginate manually
        total_items = len(all_requests)
        total_pages = (total_items + per_page - 1) // per_page
        
        start = (page - 1) * per_page
        end = start + per_page
        paginated_items = all_requests[start:end]
        
        # Remove sort_date and convert datetime to string for JSON
        for item in paginated_items:
            if 'sort_date' in item:
                del item['sort_date']
            if isinstance(item['request_date'], datetime):
                item['request_date'] = item['request_date'].isoformat()
            if 'processed_date' in item and item['processed_date'] and isinstance(item['processed_date'], datetime):
                item['processed_date'] = item['processed_date'].isoformat()
        
        return jsonify({
            'requests': paginated_items,
            'total_pages': total_pages,
            'current_page': page
        })


@admin_bp.route('/deletion-requests/stats')
def get_deletion_requests_stats():
    """Get statistics for deletion requests including audit logs"""
    
    # Get GLOBAL filters from session (set by dashboard)
    start_date_filter = session.get('admin_start_date')
    end_date_filter = session.get('admin_end_date')
    year = session.get('admin_current_year')  
    
    # Parse date range from session
    start_date_obj = None
    end_date_obj = None
    if start_date_filter and end_date_filter:
        try:
            start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError):
            pass
    
    # If no explicit date range, try year filter
    if not start_date_obj and year:
        try:
            year_int = int(year)
            start_date_obj = datetime(year_int, 1, 1)
            end_date_obj = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            pass
    
    # Current pending requests with global filter
    pending_query = DeletionRequest.query.filter_by(status='pending')
    if start_date_obj and end_date_obj:
        pending_query = pending_query.filter(
            DeletionRequest.request_date >= start_date_obj,
            DeletionRequest.request_date <= end_date_obj
        )
    pending = pending_query.count()
    
    # ✅ FIXED: From audit table with global filter
    approved_query = DeletionRequestAudit.query.filter_by(status='approved')
    rejected_query = DeletionRequestAudit.query.filter_by(status='rejected')
    
    if start_date_obj and end_date_obj:
        approved_query = approved_query.filter(
            DeletionRequestAudit.request_date >= start_date_obj,
            DeletionRequestAudit.request_date <= end_date_obj
        )
        rejected_query = rejected_query.filter(
            DeletionRequestAudit.request_date >= start_date_obj,
            DeletionRequestAudit.request_date <= end_date_obj
        )
    
    approved = approved_query.count()
    rejected = rejected_query.count()
    
    # Total including both tables
    total = pending + approved + rejected
    
    return jsonify({
        'total': total,
        'pending': pending,
        'approved': approved,
        'rejected': rejected
    })


@admin_bp.route('/deletion-requests/<int:request_id>')
@admin_required
def get_deletion_request(request_id):
    """Get details of a specific deletion request - fetches from audit table for processed requests"""
    try:
        # First try to get from deletion_requests table (pending requests)
        req = DeletionRequest.query.get(request_id)
        
        if req:
            # This is a pending request - student still exists
            student_name = "Unknown"
            student_id_number = "N/A"
            
            if req.student:
                student_name = f"{req.student.first_name} {req.student.last_name}"
                student_id_number = req.student.id_number if req.student.id_number else "N/A"
            
            return jsonify({
                'success': True,
                'from_audit': False,
                'id': req.id,
                'student_name': student_name,
                'student_id_number': student_id_number,
                'reason': req.reason,
                'request_date': req.request_date.isoformat() if req.request_date else None,
                'status': req.status,
                'admin_notes': req.admin_notes,
                'processed_by_name': req.processed_by.username if req.processed_by else None,
                'processed_date': req.processed_date.isoformat() if req.processed_date else None
            })
        
        # ✅ FIXED: If not found in deletion_requests, check the audit table (processed requests)
        # Try to find by original_request_id first, then by id
        audit = DeletionRequestAudit.query.filter_by(original_request_id=request_id).first()
        
        if not audit:
            # Try by audit table's own ID
            audit = DeletionRequestAudit.query.get(request_id)
        
        if audit:
            # This is a processed request - get data from audit table
            return jsonify({
                'success': True,
                'from_audit': True,
                'id': audit.id,
                'original_request_id': audit.original_request_id,
                'student_name': audit.student_name,
                'student_id_number': audit.student_id_number,
                'reason': audit.reason,
                'request_date': audit.request_date.isoformat() if audit.request_date else None,
                'status': audit.status,
                'admin_notes': audit.admin_notes,
                'processed_by_name': audit.processed_by_username,
                'processed_date': audit.processed_date.isoformat() if audit.processed_date else None,
                'votes_anonymized': audit.votes_anonymized,
                'action_taken': audit.action_taken
            })
        
        # Not found in either table
        return jsonify({
            'success': False,
            'error': 'Deletion request not found'
        }), 404
        
    except Exception as e:
        current_app.logger.error(f"Error getting deletion request {request_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@admin_bp.route('/debug-audit-table')
@admin_required
def debug_audit_table():
    """Debug endpoint to check audit table contents"""
    from admin.models import DeletionRequestAudit
    
    all_records = DeletionRequestAudit.query.all()
    approved_records = DeletionRequestAudit.query.filter_by(status='approved').all()
    
    result = {
        'total_records': len(all_records),
        'approved_count': len(approved_records),
        'records': []
    }
    
    for r in all_records:
        result['records'].append({
            'id': r.id,
            'student_name': r.student_name,
            'status': r.status,
            'action_taken': r.action_taken,
            'processed_date': r.processed_date.isoformat() if r.processed_date else None
        })
    
    return jsonify(result)

@admin_bp.route('/deletion-requests/<int:request_id>/process', methods=['POST'])
@admin_required
def process_deletion_request(request_id):
    """Approve or reject a deletion request - COMPLETELY DELETE STUDENT"""
    
    print(f"\n🔥 PROCESSING DELETION REQUEST #{request_id}")
    
    try:
        data = request.get_json()
        action = data.get('action')
        admin_notes = data.get('admin_notes', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({'success': False, 'message': 'Invalid action'}), 400
        
        req = DeletionRequest.query.get_or_404(request_id)
        student = req.student
        
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
            
        student_name = f"{student.first_name} {student.last_name}"
        student_id = student.id
        student_id_number = student.id_number
        
        # Store data for audit log (using existing table - NO NEW COLUMNS NEEDED)
        audit_data = {
            'original_request_id': req.id,
            'student_id': student_id,
            'student_name': student_name,
            'student_id_number': student_id_number,
            'reason': req.reason,
            'request_date': req.request_date,
            'processed_date': datetime.utcnow(),
            'processed_by': current_user.id,
            'processed_by_username': current_user.username,
            'admin_notes': admin_notes,
            'action_taken': action,
            'status': 'approved' if action == 'approve' else 'rejected',  # 🔥 FIXED: Use full word 'approved'
            'votes_anonymized': 0  # Will update below
        }
        
        vote_count = 0
        
        if action == 'approve':
            print(f"✅ APPROVING - Completely deleting student ID: {student_id}")
            
            # 1. Anonymize votes (keep them but disconnect from student)
            votes = Vote.query.filter_by(student_id=student_id).all()
            vote_count = len(votes)
            for vote in votes:
                vote.original_student_id = student_id
                vote.anonymized_at = datetime.utcnow()
                vote.student_id = None  # Disconnect from student
            
            # Update vote count in audit data
            audit_data['votes_anonymized'] = vote_count
            
            # 2. Delete related records (these must be deleted before student)
            pending = PendingCandidate.query.filter_by(student_id=student_id).all()
            for p in pending:
                db.session.delete(p)
            
            qualified = QualifiedCandidate.query.filter_by(student_id=student_id).all()
            for q in qualified:
                db.session.delete(q)
            
            devices = TrustedDevice.query.filter_by(student_id=student_id).all()
            for d in devices:
                db.session.delete(d)
            
            # 3. Create audit log entry BEFORE deleting student
            audit_entry = DeletionRequestAudit(**audit_data)
            db.session.add(audit_entry)
            
            # 4. Update the deletion request status
            req.status = 'approved'
            req.processed_date = datetime.utcnow()
            req.processed_by = current_user.id
            req.admin_notes = admin_notes
            
            # 5. FLUSH to ensure audit and request are saved before deletion
            db.session.flush()
            
            # 6. COMPLETELY DELETE the student from database
            db.session.delete(student)
            
            db.session.commit()
            print(f"✅ Student {student_name} (ID: {student_id}) COMPLETELY DELETED from database")
            print(f"✅ {vote_count} votes anonymized and preserved")
            print(f"✅ Audit log saved in deletion_request_audit table")
            
        else:  # reject
            print(f"🚫 REJECTING - Keeping student account")
            
            # Create audit log for rejection
            audit_data['votes_anonymized'] = 0
            audit_entry = DeletionRequestAudit(**audit_data)
            db.session.add(audit_entry)
            
            # Update the deletion request status (keep student)
            req.status = 'rejected'
            req.processed_date = datetime.utcnow()
            req.processed_by = current_user.id
            req.admin_notes = admin_notes
            
            db.session.commit()
            print(f"✅ Request rejected, student account preserved")
            print(f"✅ Audit log saved in deletion_request_audit table")
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500



from student.models import QualifiedCandidate

@admin_bp.route('/convert-to-candidates')
@login_required
def convert_to_candidates():
    """Page for admin to select students and qualify them as candidates"""
    from sqlalchemy import or_
    from student.models import QualifiedCandidate
    
    # Get filter parameters - only search now
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    per_page = 15
    
    # Base query - get all students
    query = Student.query
    
    # Apply search filter only
    if search:
        query = query.filter(
            or_(
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                Student.id_number.ilike(f'%{search}%')
            )
        )
    
    # Order by ID
    query = query.order_by(Student.id_number)
    
    # Paginate
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    students = paginated.items
    
    # Get qualified students
    qualified = QualifiedCandidate.query.all()
    qualified_student_ids = [q.student_id for q in qualified]
    qualified_count = len(qualified_student_ids)
    
    # Get filter dropdown data (still needed for the template but we'll hide them)
    departments = Department.query.all()
    courses = Course.query.all()
    year_levels = YearLevel.query.all()
    
    return render_template('convert_to_candidates.html',
                         students=students,
                         departments=departments,
                         courses=courses,
                         year_levels=year_levels,
                         qualified_student_ids=qualified_student_ids,
                         qualified_count=qualified_count,
                         current_page=page,
                         total_pages=paginated.pages,
                         total_students=paginated.total,
                         search=search)


@admin_bp.route('/get-qualified-students-data')
@login_required
def get_qualified_students_data():
    """AJAX endpoint for qualified students with real-time search (no filters)"""
    from sqlalchemy import or_
    from student.models import QualifiedCandidate
    
    page = request.args.get('page', 1, type=int)
    per_page = 15
    search = request.args.get('search', '').strip()
    
    # Base query
    query = Student.query
    
    # Apply search filter only
    if search:
        query = query.filter(
            or_(
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                Student.id_number.ilike(f'%{search}%')
            )
        )
    
    # Order by ID
    query = query.order_by(Student.id_number)
    
    # Get total count for stats
    total_students = query.count()
    
    # Paginate
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    students = paginated.items
    
    # Get qualified students
    qualified = QualifiedCandidate.query.all()
    qualified_student_ids = [q.student_id for q in qualified]
    qualified_count = len(qualified_student_ids)
    
    # Generate HTML for table body
    html = ''
    for student in students:
        full_name = f"{student.first_name} {student.last_name}"
        is_qualified = student.id in qualified_student_ids
        year_name = student.year_level.year_name if student.year_level else 'N/A'
        
        if is_qualified:
            status_html = '<span class="candidate-badge"><i class="fa-solid fa-star"></i> Qualified</span>'
            action_html = f'<button class="action-btn remove-btn" onclick="removeQualification({student.id})"><i class="fa-solid fa-times"></i> Remove</button>'
        else:
            status_html = '<span class="not-candidate-badge"><i class="fa-solid fa-user"></i> Not Qualified</span>'
            action_html = f'<button class="action-btn qualify-btn" onclick="qualifyStudent({student.id})"><i class="fa-solid fa-check"></i> Qualify</button>'
        
        html += f'''
        <tr data-id="{student.id}">
            <td>{student.id_number}</td>
            <td>{full_name}</td>
            <td>{student.course or 'N/A'}</td>
            <td>{year_name}</td>
            <td>{status_html}</td>
            <td>{action_html}</td>
        </tr>
        '''
    
    # Empty state
    if not students:
        html = '''
        <tr>
            <td colspan="6" class="empty-state">
                <i class="fa-solid fa-users-slash"></i>
                <h3>No Students Found</h3>
                <p>No students match your search criteria.</p>
            </td>
        </tr>
        '''
    
    return jsonify({
        'html': html,
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'total_students': total_students,
        'qualified_count': qualified_count
    })


@admin_bp.route('/qualify-student', methods=['POST'])
@login_required
def qualify_student():
    """Mark a student as qualified candidate"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'success': False, 'message': 'Student ID required'}), 400
        
        # Check if student exists
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        # Check if already qualified
        existing = QualifiedCandidate.query.filter_by(student_id=student_id).first()
        if existing:
            return jsonify({'success': False, 'message': 'Student is already qualified'}), 400
        
        # Create qualification record
        qualification = QualifiedCandidate(
            student_id=student_id,
            status='pending'
        )
        
        db.session.add(qualification)
        db.session.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='QUALIFY_STUDENT',
            description=f"Admin user '{username}' qualified student {student.first_name} {student.last_name} (ID: {student.id_number}) as candidate from IP: {ip}"
        )
        
        return jsonify({
            'success': True, 
            'message': 'Student qualified successfully!'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error qualifying student: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/remove-qualification', methods=['POST'])
@login_required
def remove_qualification():
    """Remove qualification from a student"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'success': False, 'message': 'Student ID required'}), 400
        
        # Find and delete qualification
        qualification = QualifiedCandidate.query.filter_by(student_id=student_id).first()
        if not qualification:
            return jsonify({'success': False, 'message': 'Student is not qualified'}), 404
        
        student = Student.query.get(student_id)
        
        db.session.delete(qualification)
        db.session.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='REMOVE_QUALIFICATION',
            description=f"Admin user '{username}' removed candidate qualification for student {student.first_name} {student.last_name} (ID: {student.id_number}) from IP: {ip}"
        )
        
        return jsonify({
            'success': True, 
            'message': 'Qualification removed successfully!'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error removing qualification: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    


@admin_bp.route('/pending-candidates')
@login_required
def pending_candidates():
    """View and manage pending candidate applications"""
    from student.models import PendingCandidate
    from sqlalchemy import or_
    
    # Get GLOBAL filters from session (set by dashboard)
    start_date_filter = session.get('admin_start_date')
    end_date_filter = session.get('admin_end_date')
    year = session.get('admin_current_year') 
    
    # Parse date range from session
    start_date = None
    end_date = None
    if start_date_filter and end_date_filter:
        try:
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError):
            pass
    
    # If no explicit date range, try year filter
    if not start_date and year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            pass
    
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', 'pending')
    search = request.args.get('search', '').strip()
    per_page = 10
    
    # Base query
    query = PendingCandidate.query
    
    # Apply GLOBAL date range filter (from dashboard) - JOIN ONLY ONCE
    if start_date and end_date:
        query = query.join(Election, PendingCandidate.election_id == Election.id).filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    # Filter by status
    if status != 'all':
        query = query.filter(PendingCandidate.status == status)
    
    # Search filter - FIXED: Use ONE join to elections, not multiple
    if search:
        # Join only once, then use that join for filtering
        query = query.join(Position, PendingCandidate.position_id == Position.id)
        
        # Only join elections if not already joined
        if not (start_date and end_date):
            query = query.join(Election, PendingCandidate.election_id == Election.id)
        
        # Join student for search
        query = query.outerjoin(Student, PendingCandidate.student_id == Student.id)
        
        query = query.filter(
            or_(
                PendingCandidate.first_name.ilike(f'%{search}%'),
                PendingCandidate.last_name.ilike(f'%{search}%'),
                Position.name.ilike(f'%{search}%'),
                Election.title.ilike(f'%{search}%'),
                Student.id_number.ilike(f'%{search}%')
            )
        )
    
    # Order by most recent first
    query = query.order_by(PendingCandidate.applied_at.desc())
    
    # Paginate
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    applications = paginated.items
    
    # Get statistics with the SAME date range filter
    stats_query = PendingCandidate.query
    
    # Apply date filter to stats (same logic as main query) - JOIN ONLY ONCE
    if start_date and end_date:
        stats_query = stats_query.join(Election, PendingCandidate.election_id == Election.id).filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    # Get counts with NO status filter (for overall stats)
    stats = {
        'pending': stats_query.filter(PendingCandidate.status == 'pending').count(),
        'approved': stats_query.filter(PendingCandidate.status == 'approved').count(),
        'rejected': stats_query.filter(PendingCandidate.status == 'rejected').count(),
        'total': stats_query.count()
    }
    
    return render_template('pending_candidates.html',
                         applications=applications,
                         pagination=paginated,
                         stats=stats,
                         status_filter=status,
                         search=search,
                         current_year=year)


@admin_bp.route('/pending-candidates/stats')
@login_required
def get_pending_candidates_stats():
    """Get statistics for pending candidates with filters applied"""
    from student.models import PendingCandidate
    
    # Get GLOBAL filters from session (set by dashboard)
    start_date_filter = session.get('admin_start_date')
    end_date_filter = session.get('admin_end_date')
    year = session.get('admin_current_year') 
    
    # Parse date range from session
    start_date = None
    end_date = None
    if start_date_filter and end_date_filter:
        try:
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError):
            pass
    
    # If no explicit date range, try year filter
    if not start_date and year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            pass
    
    # Base query
    query = PendingCandidate.query
    
    # Apply date range filter
    if start_date and end_date:
        query = query.join(Election, PendingCandidate.election_id == Election.id).filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    stats = {
        'pending': query.filter(PendingCandidate.status == 'pending').count(),
        'approved': query.filter(PendingCandidate.status == 'approved').count(),
        'rejected': query.filter(PendingCandidate.status == 'rejected').count(),
        'total': query.count()
    }
    
    print(f"Stats endpoint - Date range: {start_date} to {end_date}")
    print(f"Stats endpoint - Counts: {stats}")
    
    return jsonify(stats)


@admin_bp.route('/pending-candidate/<int:id>')
@login_required
def get_pending_candidate(id):
    """Get pending candidate details for modal view"""
    from student.models import PendingCandidate
    
    try:
        pending = PendingCandidate.query.get_or_404(id)
        student = Student.query.get(pending.student_id) if pending.student_id else None
        
        # Get course and year level names
        course_name = None
        year_name = None
        
        if student:
            if student.course_id:
                course = Course.query.get(student.course_id)
                course_name = course.course_name if course else None
            if student.year_level_id:
                year = YearLevel.query.get(student.year_level_id)
                year_name = year.year_name if year else None
        
        return jsonify({
            'success': True,
            'application': {
                'id': pending.id,
                'first_name': pending.first_name,
                'last_name': pending.last_name,
                'student_id': student.id_number if student else 'N/A',
                'course': course_name or (student.course if student else 'N/A'),
                'year_level': year_name or 'N/A',
                'party_list': pending.party_list,
                'platform': pending.platform,
                'position': pending.position.name if pending.position else 'N/A',
                'election': pending.election.title if pending.election else 'N/A',
                'election_id': pending.election_id,
                'scope': pending.scope,
                'applied_at': pending.applied_at.strftime('%Y-%m-%d %H:%M') if pending.applied_at else 'N/A',
                'status': pending.status
            }
        })
    except Exception as e:
        print(f"Error in get_pending_candidate: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500



@admin_bp.route('/approve-pending/<int:id>', methods=['POST'])
@login_required
def approve_pending(id):
    """Approve a pending candidate application and move to candidates table"""
    from student.models import PendingCandidate
    from datetime import datetime
    
    try:
        pending = PendingCandidate.query.get_or_404(id)
        
        if pending.status != 'pending':
            return jsonify({'success': False, 'message': 'This application is no longer pending'}), 400
        
        # Check if candidate already exists with same name in the SAME election only
        existing = Candidate.query.filter_by(
            first_name=pending.first_name,
            last_name=pending.last_name,
            election_id=pending.election_id
        ).first()
        
        if existing:
            pending.status = 'rejected'
            pending.rejection_reason = f'Candidate with same name already exists in this election: {pending.election.title if pending.election else "Unknown"}'
            pending.reviewed_at = datetime.utcnow()
            db.session.commit()
            
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            
            log_audit(
                action='AUTO_REJECT_CANDIDATE',
                description=f"Admin user '{username}' auto-rejected duplicate candidate application for {pending.first_name} {pending.last_name} in election {pending.election.title if pending.election else 'Unknown'} from IP: {ip}"
            )
            
            return jsonify({
                'success': False, 
                'message': f'A candidate with the name {pending.first_name} {pending.last_name} already exists in this election. Cannot approve.'
            }), 400
        
        # Create new candidate with year level
        new_candidate = Candidate(
            first_name=pending.first_name,
            last_name=pending.last_name,
            party_list=pending.party_list,
            platform=pending.platform,
            department_id=pending.department_id,
            course_id=pending.course_id,
            year_level_id=pending.year_level_id,
            position_id=pending.position_id,
            election_id=pending.election_id,
            scope=pending.scope,
            photo=pending.photo
        )
        
        db.session.add(new_candidate)
        
        # Update pending record
        pending.status = 'approved'
        pending.reviewed_at = datetime.utcnow()
        if hasattr(current_user, 'id'):
            pending.reviewed_by = current_user.id
        
        db.session.commit()
        
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='APPROVE_CANDIDATE',
            description=f"Admin user '{username}' approved candidate application for {pending.first_name} {pending.last_name} in election {pending.election.title if pending.election else 'Unknown'} from IP: {ip}"
        )
        
        return jsonify({
            'success': True, 
            'message': f'Candidate {pending.first_name} {pending.last_name} approved successfully for {pending.election.title if pending.election else "election"}!'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in approve_pending: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/reject-pending/<int:id>', methods=['POST'])
@login_required
def reject_pending(id):
    """Reject a pending candidate application"""
    from student.models import PendingCandidate
    from datetime import datetime
    
    try:
        data = request.get_json()
        reason = data.get('reason', 'Application rejected')
        
        if not reason:
            return jsonify({'success': False, 'message': 'Rejection reason is required'}), 400
        
        pending = PendingCandidate.query.get_or_404(id)
        
        if pending.status != 'pending':
            return jsonify({'success': False, 'message': 'This application is no longer pending'}), 400
        
        pending.status = 'rejected'
        pending.rejection_reason = reason
        pending.reviewed_at = datetime.utcnow()
        if hasattr(current_user, 'id'):
            pending.reviewed_by = current_user.id
        
        db.session.commit()
        
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='REJECT_CANDIDATE',
            description=f"Admin user '{username}' rejected candidate application for {pending.first_name} {pending.last_name} in election {pending.election.title if pending.election else 'Unknown'} from IP: {ip} | Reason: {reason[:100]}..."
        )
        
        return jsonify({'success': True, 'message': 'Application rejected'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in reject_pending: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500



@admin_bp.route('/bulk-approve-pending', methods=['POST'])
@login_required
def bulk_approve_pending():
    """Bulk approve multiple pending candidate applications"""
    from student.models import PendingCandidate
    from datetime import datetime
    
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'message': 'No applications selected'}), 400
        
        approved_count = 0
        errors = []
        
        for app_id in ids:
            try:
                pending = PendingCandidate.query.get(app_id)
                if not pending:
                    errors.append(f'Application {app_id} not found')
                    continue
                
                if pending.status != 'pending':
                    errors.append(f'Application {pending.first_name} {pending.last_name} is no longer pending')
                    continue
                
                # Check if candidate already exists
                existing = Candidate.query.filter_by(
                    first_name=pending.first_name,
                    last_name=pending.last_name,
                    election_id=pending.election_id
                ).first()
                
                if existing:
                    pending.status = 'rejected'
                    pending.rejection_reason = f'Bulk approve failed: Candidate with same name already exists in this election'
                    pending.reviewed_at = datetime.utcnow()
                    db.session.commit()
                    errors.append(f'{pending.first_name} {pending.last_name}: Duplicate candidate exists')
                    continue
                
                # Create new candidate
                new_candidate = Candidate(
                    first_name=pending.first_name,
                    last_name=pending.last_name,
                    party_list=pending.party_list,
                    platform=pending.platform,
                    department_id=pending.department_id,
                    course_id=pending.course_id,
                    year_level_id=pending.year_level_id,
                    position_id=pending.position_id,
                    election_id=pending.election_id,
                    scope=pending.scope,
                    photo=pending.photo
                )
                
                db.session.add(new_candidate)
                
                # Update pending record
                pending.status = 'approved'
                pending.reviewed_at = datetime.utcnow()
                if hasattr(current_user, 'id'):
                    pending.reviewed_by = current_user.id
                
                approved_count += 1
                
            except Exception as e:
                errors.append(f'Error processing application {app_id}: {str(e)}')
        
        db.session.commit()
        
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='BULK_APPROVE_CANDIDATES',
            description=f"Admin user '{username}' bulk approved {approved_count} candidate applications from IP: {ip}"
        )
        
        return jsonify({
            'success': True,
            'approved_count': approved_count,
            'errors': errors if errors else None,
            'message': f'Successfully approved {approved_count} application(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in bulk_approve_pending: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/bulk-reject-pending', methods=['POST'])
@login_required
def bulk_reject_pending():
    """Bulk reject multiple pending candidate applications"""
    from student.models import PendingCandidate
    from datetime import datetime
    
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        reason = data.get('reason', 'Bulk rejection')
        
        if not ids:
            return jsonify({'success': False, 'message': 'No applications selected'}), 400
        
        if not reason:
            return jsonify({'success': False, 'message': 'Rejection reason is required'}), 400
        
        rejected_count = 0
        errors = []
        
        for app_id in ids:
            try:
                pending = PendingCandidate.query.get(app_id)
                if not pending:
                    errors.append(f'Application {app_id} not found')
                    continue
                
                if pending.status != 'pending':
                    errors.append(f'Application {pending.first_name} {pending.last_name} is no longer pending')
                    continue
                
                pending.status = 'rejected'
                pending.rejection_reason = reason
                pending.reviewed_at = datetime.utcnow()
                if hasattr(current_user, 'id'):
                    pending.reviewed_by = current_user.id
                
                rejected_count += 1
                
            except Exception as e:
                errors.append(f'Error processing application {app_id}: {str(e)}')
        
        db.session.commit()
        
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='BULK_REJECT_CANDIDATES',
            description=f"Admin user '{username}' bulk rejected {rejected_count} candidate applications from IP: {ip} | Reason: {reason[:100]}..."
        )
        
        return jsonify({
            'success': True,
            'rejected_count': rejected_count,
            'errors': errors if errors else None,
            'message': f'Successfully rejected {rejected_count} application(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in bulk_reject_pending: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


        
# ---------------------- Departments & Courses ---------------------- #
@admin_bp.route('/departments')
@admin_required
def manage_departments():
    connection = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB
    )
    cursor = connection.cursor(dictionary=True)

    cursor.execute("SELECT * FROM departments ORDER BY name")
    departments = cursor.fetchall()

    cursor.execute("""
        SELECT c.*, d.name AS department_name
        FROM courses c
        JOIN departments d ON c.department_id = d.id
        ORDER BY d.name, c.course_name
    """)
    courses = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template('manage_departments.html', departments=departments, courses=courses)


@admin_bp.route('/departments/add', methods=['POST'])
@admin_required
def add_department():
    name = request.form['name'].strip()
    
    connection = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB
    )
    cursor = connection.cursor(dictionary=True)
    
    # Check for duplicate department name
    cursor.execute("SELECT id FROM departments WHERE name = %s", (name,))
    existing = cursor.fetchone()
    
    if existing:
        cursor.close()
        connection.close()
        flash(f'❌ Department "{name}" already exists!', 'error')
        return redirect(url_for('admin.manage_departments'))
    
    try:
        cursor.execute("INSERT INTO departments (name) VALUES (%s)", (name,))
        connection.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='ADD_DEPARTMENT',
            description=f"Admin user '{username}' added new department: '{name}' from IP: {ip}"
        )
        
        flash(f'✅ Department added successfully!', 'success')
    except Exception as e:
        flash(f'❌ Error adding department: {str(e)}', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin.manage_departments'))


@admin_bp.route('/courses/add', methods=['POST'])
@admin_required
def add_course():
    course_name = request.form['course_name'].strip()
    department_id = request.form['department_id']
    
    if not department_id:
        flash('⚠️ Please select a department', 'warning')
        return redirect(url_for('admin.manage_departments'))

    connection = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB
    )
    cursor = connection.cursor(dictionary=True)
    
    # Get department name for message
    cursor.execute("SELECT name FROM departments WHERE id = %s", (department_id,))
    dept = cursor.fetchone()
    dept_name = dept['name'] if dept else 'Unknown'
    
    # Check for duplicate course in the same department
    cursor.execute(
        "SELECT id FROM courses WHERE course_name = %s AND department_id = %s",
        (course_name, department_id)
    )
    existing = cursor.fetchone()
    
    if existing:
        cursor.close()
        connection.close()
        flash(f'❌ Course "{course_name}" already exists in {dept_name} department!', 'error')
        return redirect(url_for('admin.manage_departments'))
    
    try:
        cursor.execute(
            "INSERT INTO courses (course_name, department_id) VALUES (%s, %s)",
            (course_name, department_id)
        )
        connection.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='ADD_COURSE',
            description=f"Admin user '{username}' added new course: '{course_name}' to department: '{dept_name}' (ID: {department_id}) from IP: {ip}"
        )
        
        flash(f'✅ Course added successfully!', 'success')
    except Exception as e:
        flash(f'❌ Error adding course: {str(e)}', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin.manage_departments'))


@admin_bp.route('/departments/delete-multiple', methods=['POST'])
@admin_required
def delete_multiple_departments():
    ids = request.form.getlist('department_ids')
    
    if not ids:
        flash('⚠️ No departments selected for deletion.', 'warning')
        return redirect(url_for('admin.manage_departments'))
    
    connection = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB
    )
    cursor = connection.cursor(dictionary=True)
    
    try:
        # Check if departments have courses
        format_strings = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            SELECT COUNT(*) as course_count 
            FROM courses 
            WHERE department_id IN ({format_strings})
        """, tuple(ids))
        result = cursor.fetchone()
        course_count = result['course_count'] if result else 0
        
        if course_count > 0:
            flash(f'❌ Cannot delete departments that have courses. Please delete courses first.', 'error')
            cursor.close()
            connection.close()
            return redirect(url_for('admin.manage_departments'))
        
        # Get department names before deletion
        cursor.execute(f"SELECT name FROM departments WHERE id IN ({format_strings})", tuple(ids))
        departments_to_delete = cursor.fetchall()
        
        if not departments_to_delete:
            flash('⚠️ No matching departments found to delete.', 'warning')
            cursor.close()
            connection.close()
            return redirect(url_for('admin.manage_departments'))
        
        # Delete departments
        cursor.execute(f"DELETE FROM departments WHERE id IN ({format_strings})", tuple(ids))
        connection.commit()
        rows_deleted = cursor.rowcount
        
        if rows_deleted > 0:
            # Audit log
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            
            log_audit(
                action='DELETE_MULTIPLE_DEPARTMENTS',
                description=f"Admin user '{username}' deleted {len(departments_to_delete)} department(s) from IP: {ip}"
            )
            
            flash(f'✅ {len(departments_to_delete)} department(s) deleted successfully!', 'success')
        else:
            flash('⚠️ No departments were deleted.', 'warning')
            
    except Exception as e:
        flash(f'❌ Error deleting departments: {str(e)}', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin.manage_departments'))


@admin_bp.route('/courses/delete-multiple', methods=['POST'])
@admin_required
def delete_multiple_courses():
    ids = request.form.getlist('course_ids')
    
    if not ids:
        flash('⚠️ No courses selected for deletion.', 'warning')
        return redirect(url_for('admin.manage_departments'))
    
    connection = mysql.connector.connect(
        host=MYSQL_HOST,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DB
    )
    cursor = connection.cursor(dictionary=True)
    
    try:
        format_strings = ','.join(['%s'] * len(ids))
        
        # Check if courses have students
        cursor.execute(f"""
            SELECT COUNT(*) as student_count 
            FROM students 
            WHERE course_id IN ({format_strings})
        """, tuple(ids))
        result = cursor.fetchone()
        student_count = result['student_count'] if result else 0
        
        if student_count > 0:
            flash(f'❌ Cannot delete courses that have students assigned. Please reassign students first.', 'error')
            cursor.close()
            connection.close()
            return redirect(url_for('admin.manage_departments'))
        
        # Get course names before deletion
        cursor.execute(f"""
            SELECT c.course_name, d.name as dept_name
            FROM courses c
            JOIN departments d ON c.department_id = d.id
            WHERE c.id IN ({format_strings})
        """, tuple(ids))
        courses_to_delete = cursor.fetchall()
        
        if not courses_to_delete:
            flash('⚠️ No matching courses found to delete.', 'warning')
            cursor.close()
            connection.close()
            return redirect(url_for('admin.manage_departments'))
        
        # Delete courses
        cursor.execute(f"DELETE FROM courses WHERE id IN ({format_strings})", tuple(ids))
        connection.commit()
        rows_deleted = cursor.rowcount
        
        if rows_deleted > 0:
            # Audit log
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            
            log_audit(
                action='DELETE_MULTIPLE_COURSES',
                description=f"Admin user '{username}' deleted {len(courses_to_delete)} course(s) from IP: {ip}"
            )
            
            flash(f'✅ {len(courses_to_delete)} course(s) deleted successfully!', 'success')
        else:
            flash('⚠️ No courses were deleted.', 'warning')
            
    except Exception as e:
        flash(f'❌ Error deleting courses: {str(e)}', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin.manage_departments'))


@admin_bp.route('/departments/update', methods=['POST'])
@admin_required
def update_department():
    try:
        dept_id = request.form['id']
        new_name = request.form['name'].strip()
        
        connection = mysql.connector.connect(
            host=MYSQL_HOST,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            database=MYSQL_DB
        )
        cursor = connection.cursor(dictionary=True)
        
        # Check if department exists
        cursor.execute("SELECT name FROM departments WHERE id = %s", (dept_id,))
        old_name_result = cursor.fetchone()
        
        if not old_name_result:
            return jsonify({'success': False, 'message': 'Department not found'}), 404
        
        old_name = old_name_result['name']
        
        # Check for duplicate name
        if old_name != new_name:
            cursor.execute("SELECT id FROM departments WHERE name = %s AND id != %s", (new_name, dept_id))
            duplicate = cursor.fetchone()
            if duplicate:
                return jsonify({'success': False, 'message': f'Department "{new_name}" already exists'}), 400
        
        # Update department
        cursor.execute(
            "UPDATE departments SET name = %s WHERE id = %s",
            (new_name, dept_id)
        )
        connection.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='UPDATE_DEPARTMENT',
            description=f"Admin user '{username}' updated department from '{old_name}' to '{new_name}' (ID: {dept_id}) from IP: {ip}"
        )
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'message': '✅ Department updated successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'❌ Error: {str(e)}'}), 500


@admin_bp.route('/courses/update', methods=['POST'])
@admin_required
def update_course():
    try:
        course_id = request.form['id']
        new_name = request.form['course_name'].strip()
        new_department_id = request.form['department_id']
        
        connection = mysql.connector.connect(
            host=MYSQL_HOST,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            database=MYSQL_DB
        )
        cursor = connection.cursor(dictionary=True)
        
        # Get old data
        cursor.execute("""
            SELECT c.course_name, c.department_id, d.name as dept_name 
            FROM courses c
            JOIN departments d ON c.department_id = d.id
            WHERE c.id = %s
        """, (course_id,))
        old_data = cursor.fetchone()
        
        if not old_data:
            return jsonify({'success': False, 'message': 'Course not found'}), 404
        
        # Check for duplicate
        if old_data['course_name'] != new_name or old_data['department_id'] != int(new_department_id):
            cursor.execute("""
                SELECT id FROM courses 
                WHERE course_name = %s AND department_id = %s AND id != %s
            """, (new_name, new_department_id, course_id))
            duplicate = cursor.fetchone()
            if duplicate:
                return jsonify({'success': False, 'message': f'Course "{new_name}" already exists in this department'}), 400
        
        # Update course
        cursor.execute(
            "UPDATE courses SET course_name = %s, department_id = %s WHERE id = %s",
            (new_name, new_department_id, course_id)
        )
        connection.commit()
        
        # Get new department name
        cursor.execute("SELECT name FROM departments WHERE id = %s", (new_department_id,))
        new_dept = cursor.fetchone()
        new_dept_name = new_dept['name'] if new_dept else 'Unknown'
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='UPDATE_COURSE',
            description=f"Admin user '{username}' updated course from '{old_data['course_name']}' ({old_data['dept_name']}) to '{new_name}' ({new_dept_name}) from IP: {ip}"
        )
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True, 
            'message': '✅ Course updated successfully!',
            'department_name': new_dept_name
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'❌ Error: {str(e)}'}), 500


# ---------------------- MANAGE CANDIDATES ---------------------- #
@admin_bp.route('/candidates', methods=['GET', 'POST'])
@admin_required
def manage_candidates():
    # Get year filter from session (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    # Get positions, departments, year levels (these are not filtered by year)
    positions = Position.query.all()
    departments = Department.query.order_by(Department.name).all()
    year_levels = YearLevel.query.all()
    
    # Get current time for filtering elections in dropdown
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)
    
    # ===== FOR MAIN DISPLAY: Get ALL elections (for filtering candidates) =====
    all_elections_query = Election.query.order_by(Election.start_date.desc())
    if start_date and end_date:
        all_elections_query = all_elections_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    all_elections = all_elections_query.all()
    
    # ===== FOR DROPDOWN (Add/Edit): Only upcoming and active elections =====
    dropdown_elections = []
    for e in all_elections:
        # Make dates timezone-aware for comparison
        start_date_e = e.start_date
        end_date_e = e.end_date
        
        if start_date_e.tzinfo is None:
            start_date_e = tz.localize(start_date_e)
        if end_date_e.tzinfo is None:
            end_date_e = tz.localize(end_date_e)
        
        # Include if election is upcoming (start_date > now) OR active (start_date <= now <= end_date)
        if start_date_e > now or (start_date_e <= now <= end_date_e):
            dropdown_elections.append(e)

    # ================= FILTER =================
    selected_scope = request.args.get('scope', default=None)
    department_id = request.args.get('department_id', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = 10
    selected_department = None

    query = Candidate.query

    # Filter by year through elections (using ALL elections)
    if start_date and end_date:
        # Get election IDs within the year
        election_ids = [e.id for e in all_elections]
        if election_ids:
            query = query.filter(Candidate.election_id.in_(election_ids))
        else:
            # No elections in this year, return empty result
            query = query.filter(False)

    # Filter by scope
    if selected_scope:
        query = query.filter(Candidate.scope == selected_scope)

    # Filter by department
    if selected_scope == 'department' and department_id:
        selected_department = Department.query.get(department_id)
        if selected_department:
            query = query.filter(Candidate.department_id == department_id)

    # ---------------- PAGINATE ----------------
    candidates_pagination = query.order_by(Candidate.id.desc()).paginate(page=page, per_page=per_page)
    candidates = candidates_pagination.items
    # ==========================================

    # ---------- ADD CANDIDATE ----------
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Get candidate type
        candidate_type = request.form.get('candidate_type', 'student')
        
        if candidate_type == 'studio':
            # ===== STUDIO CANDIDATE =====
            studio_name = request.form.get('studio_name', '').strip()
            position_id = request.form.get('position_id')
            election_id = request.form.get('election_id')
            scope = request.form.get('scope')
            
            # If scope is missing (disabled input doesn't submit), default to 'campus'
            if not scope:
                scope = 'campus'
            
            # Get election to verify
            election = Election.query.get(election_id)
            if not election:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Selected election does not exist.'})
                flash('Selected election does not exist.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Validate required fields for studio
            if not studio_name:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Studio name is required.'})
                flash('Studio name is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not position_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Position is required.'})
                flash('Position is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not election_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Election is required.'})
                flash('Election is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # For studio candidates, scope should always be 'campus'
            if scope != 'campus':
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Studio candidates can only be added to campus-wide elections.'})
                flash('Studio candidates can only be added to campus-wide elections.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Check for duplicate studio candidate in the SAME election
            existing_candidate = Candidate.query.filter_by(
                studio_name=studio_name,
                election_id=election_id,
                candidate_type='studio'
            ).first()
            
            if existing_candidate:
                error_msg = f'A studio candidate with the name "{studio_name}" already exists in this election: {election.title}'
                if is_ajax:
                    return jsonify({'success': False, 'message': error_msg}), 400
                flash(error_msg, 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Save photo if uploaded
            photo_file = request.files.get('photo')
            photo_filename = None
            photo_folder = os.path.join(current_app.root_path, 'admin', 'static', 'images')
            os.makedirs(photo_folder, exist_ok=True)
            
            if photo_file and photo_file.filename:
                filename = secure_filename(photo_file.filename)
                name, ext = os.path.splitext(filename)
                photo_filename = f"{name}_{int(time.time())}{ext}"
                photo_file.save(os.path.join(photo_folder, photo_filename))
            
            # Create studio candidate - ONLY these fields are populated
            new_candidate = Candidate(
                candidate_type='studio',
                studio_name=studio_name,
                first_name=None,
                last_name=None,
                party_list=None,
                platform=None,
                department_id=None,
                course_id=None,
                position_id=position_id,
                election_id=election_id,
                scope=scope,
                year_level_id=None,
                photo=photo_filename
            )
            
            db.session.add(new_candidate)
            db.session.commit()
            
            # Audit log for studio candidate
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            position_name = new_candidate.position.name if new_candidate.position else 'N/A'
            election_title = new_candidate.election.title if new_candidate.election else 'N/A'
            
            year_info = f" | Year: {year}" if year else ""
            
            log_audit(
                action='CREATE_STUDIO_CANDIDATE',
                description=f"Admin user '{username}' added studio candidate: {studio_name} from IP: {ip} | Position: {position_name} | Election: {election_title} ({scope}){year_info}"
            )
            
            # Return JSON for AJAX requests
            if is_ajax:
                return jsonify({
                    'success': True,
                    'message': 'Studio candidate added successfully!',
                    'id': new_candidate.id,
                    'studio_name': new_candidate.studio_name,
                    'candidate_type': 'studio',
                    'position': new_candidate.position.name,
                    'position_id': new_candidate.position_id,
                    'election_id': new_candidate.election_id,
                    'scope': new_candidate.scope,
                    'photo': url_for('admin.static', filename='images/' + new_candidate.photo) if new_candidate.photo else None
                })
            
            flash('Studio candidate added successfully!', 'success')
            return redirect(url_for('admin.manage_candidates'))
        
        else:
            # ===== STUDENT CANDIDATE =====
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            party_list = request.form.get('party_list')
            platform = request.form.get('platform')
            department_id_form = request.form.get('department_id', type=int)
            course_id = request.form.get('course_id', type=int)
            position_id = request.form.get('position_id')
            election_id = request.form.get('election_id')
            scope = request.form.get('scope')
            year_level_id = request.form.get('year_level_id', type=int)
            
            # Get election to verify
            election = Election.query.get(election_id)
            if not election:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Selected election does not exist.'})
                flash('Selected election does not exist.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Validate required fields for student
            if not first_name:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'First name is required.'})
                flash('First name is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not last_name:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Last name is required.'})
                flash('Last name is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not position_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Position is required.'})
                flash('Position is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not election_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Election is required.'})
                flash('Election is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not scope:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Scope is required.'})
                flash('Scope is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not year_level_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Year level is required.'})
                flash('Year level is required.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Validate department and course (required for student candidates)
            if not department_id_form:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Department is required for student candidates.'})
                flash('Department is required for student candidates.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            if not course_id:
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Course is required for student candidates.'})
                flash('Course is required for student candidates.', 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # ===== VALIDATE STUDENT EXISTS IN STUDENT TABLE =====
            # Check if student exists with matching first name, last name, department, course, and year level
            existing_student = Student.query.filter(
                Student.first_name.ilike(first_name),
                Student.last_name.ilike(last_name),
                Student.department_id == department_id_form,
                Student.course_id == course_id,
                Student.year_level_id == year_level_id
            ).first()
            
            if not existing_student:
                error_msg = f'Student "{first_name} {last_name}" not found in {department_id_form} department, course ID {course_id}, year level {year_level_id}. Please verify the student information.'
                if is_ajax:
                    return jsonify({'success': False, 'message': error_msg}), 400
                flash(error_msg, 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Check for duplicate candidate in the SAME election
            existing_candidate = Candidate.query.filter_by(
                first_name=first_name,
                last_name=last_name,
                election_id=election_id,
                candidate_type='student'
            ).first()
            
            if existing_candidate:
                error_msg = f'A candidate with the name {first_name} {last_name} already exists in this election: {election.title}'
                if is_ajax:
                    return jsonify({'success': False, 'message': error_msg}), 400
                flash(error_msg, 'danger')
                return redirect(url_for('admin.manage_candidates'))
            
            # Save photo if uploaded
            photo_file = request.files.get('photo')
            photo_filename = None
            photo_folder = os.path.join(current_app.root_path, 'admin', 'static', 'images')
            os.makedirs(photo_folder, exist_ok=True)
            
            if photo_file and photo_file.filename:
                filename = secure_filename(photo_file.filename)
                name, ext = os.path.splitext(filename)
                photo_filename = f"{name}_{int(time.time())}{ext}"
                photo_file.save(os.path.join(photo_folder, photo_filename))
            
            # Create student candidate
            new_candidate = Candidate(
                candidate_type='student',
                first_name=first_name,
                last_name=last_name,
                party_list=party_list if party_list else None,
                platform=platform if platform else None,
                department_id=department_id_form,
                course_id=course_id,
                position_id=position_id,
                election_id=election_id,
                scope=scope,
                year_level_id=year_level_id,
                studio_name=None,
                photo=photo_filename
            )
            
            db.session.add(new_candidate)
            db.session.commit()
            
            # Audit log
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            department_name = new_candidate.department.name if new_candidate.department else 'N/A'
            position_name = new_candidate.position.name if new_candidate.position else 'N/A'
            election_title = new_candidate.election.title if new_candidate.election else 'N/A'
            party_list_name = new_candidate.party_list if new_candidate.party_list else 'Independent'
            year_level_name = new_candidate.year_level.year_name if new_candidate.year_level else 'N/A'
            platform_display = new_candidate.platform[:50] + '...' if new_candidate.platform and len(new_candidate.platform) > 50 else (new_candidate.platform or 'None')
            
            year_info = f" | Year: {year}" if year else ""
            
            log_audit(
                action='CREATE_CANDIDATE',
                description=f"Admin user '{username}' added candidate: {first_name} {last_name} from IP: {ip} | Party: {party_list_name} | Position: {position_name} | Department: {department_name} | Year Level: {year_level_name} | Platform: {platform_display} | Election: {election_title} ({scope}){year_info}"
            )
            
            # Return JSON for AJAX requests
            if is_ajax:
                return jsonify({
                    'success': True,
                    'message': 'Candidate added successfully!',
                    'id': new_candidate.id,
                    'first_name': new_candidate.first_name,
                    'last_name': new_candidate.last_name,
                    'candidate_type': 'student',
                    'party_list': new_candidate.party_list,
                    'platform': new_candidate.platform,
                    'department': new_candidate.department.name if new_candidate.department else '',
                    'department_id': new_candidate.department_id,
                    'course_id': new_candidate.course_id,
                    'year_level_id': new_candidate.year_level_id,
                    'year_level': new_candidate.year_level.year_name if new_candidate.year_level else '',
                    'position': new_candidate.position.name,
                    'position_id': new_candidate.position_id,
                    'election_id': new_candidate.election_id,
                    'scope': new_candidate.scope,
                    'photo': url_for('admin.static', filename='images/' + new_candidate.photo) if new_candidate.photo else None
                })
            
            flash('Candidate added successfully!', 'success')
            return redirect(url_for('admin.manage_candidates'))
    
    # ===== FOR DROPDOWN IN MODALS: Only upcoming and active elections =====
    campus_elections = [e for e in dropdown_elections if e.scope == 'campus']
    department_elections = [e for e in dropdown_elections if e.scope == 'department']
    
    return render_template(
        'manage_candidates.html',
        candidates=candidates,
        candidates_pagination=candidates_pagination,
        positions=positions,
        departments=departments,
        year_levels=year_levels,
        elections=dropdown_elections,
        campus_elections=campus_elections,
        department_elections=department_elections,
        selected_department=selected_department,
        selected_scope=selected_scope,
        current_year=year
    )


         
@admin_bp.route('/candidates/filter', methods=['GET'])
@admin_required
def filter_candidates():
    """AJAX endpoint for filtering candidates"""
    # Get year filter from session
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    selected_scope = request.args.get('scope', default=None)
    department_id = request.args.get('department_id', type=int)
    search = request.args.get('search', default='')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    query = Candidate.query
    
    # Filter by year through elections
    if start_date and end_date:
        # Get elections within the year
        election_query = Election.query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
        election_ids = [e.id for e in election_query.all()]
        if election_ids:
            query = query.filter(Candidate.election_id.in_(election_ids))
        else:
            # No elections in this year, return empty result
            candidates_data = []
            return jsonify({
                'candidates': candidates_data,
                'pagination': {
                    'current_page': 1,
                    'total_pages': 1,
                    'total_items': 0,
                    'has_prev': False,
                    'has_next': False,
                    'prev_page': None,
                    'next_page': None
                },
                'current_year': year
            })
    
    # Filter by scope
    if selected_scope:
        query = query.filter(Candidate.scope == selected_scope)
    
    # Filter by department
    if selected_scope == 'department' and department_id:
        query = query.filter(Candidate.department_id == department_id)
    
    # Search functionality
    if search:
        search_term = f'%{search}%'
        query = query.join(Position).join(Election, isouter=True).filter(
            db.or_(
                Candidate.first_name.ilike(search_term),
                Candidate.last_name.ilike(search_term),
                Candidate.studio_name.ilike(search_term),
                Candidate.party_list.ilike(search_term),
                Candidate.platform.ilike(search_term),
                Position.name.ilike(search_term),
                Election.title.ilike(search_term)
            )
        )
    
    # Paginate
    pagination = query.order_by(Candidate.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    candidates = pagination.items
    
    # Format candidates for JSON response
    candidates_data = []
    for c in candidates:
        # Determine display name based on candidate type
        if c.candidate_type == 'studio':
            display_name = c.studio_name
            first_name = None
            last_name = None
        else:
            display_name = f"{c.first_name} {c.last_name}"
            first_name = c.first_name
            last_name = c.last_name
        
        candidates_data.append({
            'id': c.id,
            'first_name': first_name,
            'last_name': last_name,
            'studio_name': c.studio_name,
            'candidate_type': c.candidate_type,
            'display_name': display_name,
            'party_list': c.party_list,
            'platform': c.platform,
            'department': c.department.name if c.department else '',
            'department_id': c.department_id,
            'course_id': c.course_id,
            'year_level_id': c.year_level_id,
            'year_level': c.year_level.year_name if c.year_level else '',
            'position': c.position.name if c.position else '',
            'position_id': c.position_id,
            'election_id': c.election_id,
            'election_title': c.election.title if c.election else '',
            'scope': c.scope,
            'photo': url_for('admin.static', filename='images/' + c.photo) if c.photo else None
        })
    
    return jsonify({
        'candidates': candidates_data,
        'pagination': {
            'current_page': pagination.page,
            'total_pages': pagination.pages,
            'total_items': pagination.total,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_page': pagination.prev_num if pagination.has_prev else None,
            'next_page': pagination.next_num if pagination.has_next else None
        },
        'current_year': year
    })


@admin_bp.route('/candidates/edit/<int:id>', methods=['POST'])
@admin_required
def update_candidate(id):
    candidate = Candidate.query.get_or_404(id)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Store old values
    old_candidate_type = candidate.candidate_type
    old_first_name = candidate.first_name
    old_last_name = candidate.last_name
    old_studio_name = candidate.studio_name
    old_party_list = candidate.party_list
    old_platform = candidate.platform
    old_position = candidate.position.name if candidate.position else 'N/A'
    old_department = candidate.department.name if candidate.department else 'N/A'
    old_year_level = candidate.year_level.year_name if candidate.year_level else 'N/A'
    old_scope = candidate.scope
    
    # Get candidate type from form
    candidate_type = request.form.get('candidate_type', 'student')
    
    if candidate_type == 'studio':
        # ===== UPDATE STUDIO CANDIDATE =====
        studio_name = request.form.get('studio_name', '').strip()
        party_list = request.form.get('party_list')
        platform = request.form.get('platform')
        candidate.position_id = request.form.get('position_id')
        candidate.election_id = request.form.get('election_id')
        candidate.scope = request.form.get('scope')
        candidate.department_id = request.form.get('department_id', type=int)
        candidate.course_id = request.form.get('course_id', type=int)
        
        # Update studio-specific fields
        candidate.studio_name = studio_name
        candidate.party_list = party_list if party_list else None
        candidate.platform = platform if platform else None
        candidate.candidate_type = 'studio'
        
        # Clear student fields
        candidate.first_name = None
        candidate.last_name = None
        candidate.year_level_id = None
        
    else:
        # ===== UPDATE STUDENT CANDIDATE =====
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        party_list = request.form.get('party_list')
        platform = request.form.get('platform')
        candidate.position_id = request.form.get('position_id')
        candidate.election_id = request.form.get('election_id')
        candidate.scope = request.form.get('scope')
        candidate.department_id = request.form.get('department_id', type=int)
        candidate.course_id = request.form.get('course_id', type=int)
        candidate.year_level_id = request.form.get('year_level_id', type=int)
        
        # Update student-specific fields
        candidate.first_name = first_name
        candidate.last_name = last_name
        candidate.party_list = party_list if party_list else None
        candidate.platform = platform if platform else None
        candidate.candidate_type = 'student'
        
        # Clear studio fields
        candidate.studio_name = None
    
    # Get election to verify
    election = Election.query.get(candidate.election_id)
    
    # Handle photo upload
    photo_file = request.files.get('photo')
    if photo_file and photo_file.filename:
        filename = secure_filename(photo_file.filename)
        name, ext = os.path.splitext(filename)
        filename = f"{name}_{int(time.time())}{ext}"
        
        photo_folder = os.path.join(current_app.root_path, 'admin', 'static', 'images')
        os.makedirs(photo_folder, exist_ok=True)
        
        # Delete old photo if exists
        if candidate.photo:
            old_photo_path = os.path.join(photo_folder, candidate.photo)
            if os.path.exists(old_photo_path):
                try:
                    os.remove(old_photo_path)
                except Exception as e:
                    print(f"Error deleting old photo: {e}")
        
        photo_file.save(os.path.join(photo_folder, filename))
        candidate.photo = filename
    
    db.session.commit()
    
    # Format platform for audit log (truncate if too long)
    old_platform_display = old_platform[:50] + '...' if old_platform and len(old_platform) > 50 else (old_platform or 'None')
    new_platform_display = candidate.platform[:50] + '...' if candidate.platform and len(candidate.platform) > 50 else (candidate.platform or 'None')
    
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    new_department = candidate.department.name if candidate.department else 'N/A'
    new_position = candidate.position.name if candidate.position else 'N/A'
    new_party_list = candidate.party_list if candidate.party_list else 'Independent'
    new_year_level = candidate.year_level.year_name if candidate.year_level else 'N/A'
    
    # Build candidate name for audit log
    if candidate.candidate_type == 'studio':
        old_name_display = old_studio_name or 'Unknown Studio'
        new_name_display = candidate.studio_name or 'Unknown Studio'
        action_type = 'UPDATE_STUDIO_CANDIDATE'
    else:
        old_name_display = f"{old_first_name or ''} {old_last_name or ''}".strip() or 'Unknown Student'
        new_name_display = f"{candidate.first_name or ''} {candidate.last_name or ''}".strip() or 'Unknown Student'
        action_type = 'UPDATE_CANDIDATE'
    
    log_audit(
        action=action_type,
        description=f"Admin user '{username}' updated candidate from IP: {ip} | {old_name_display} → {new_name_display} | Scope: {old_scope} → {candidate.scope} | Party: {old_party_list or 'Independent'} → {new_party_list} | Platform: {old_platform_display} → {new_platform_display} | Position: {old_position} → {new_position} | Department: {old_department} → {new_department} | Year Level: {old_year_level} → {new_year_level}"
    )
    
    if is_ajax:
        # Prepare response data based on candidate type
        if candidate.candidate_type == 'studio':
            response_data = {
                'success': True,
                'message': 'Studio candidate updated successfully!',
                'candidate': {
                    'id': candidate.id,
                    'studio_name': candidate.studio_name,
                    'candidate_type': 'studio',
                    'party_list': candidate.party_list,
                    'platform': candidate.platform,
                    'department': candidate.department.name if candidate.department else '',
                    'department_id': candidate.department_id,
                    'course_id': candidate.course_id,
                    'position': candidate.position.name if candidate.position else '',
                    'position_id': candidate.position_id,
                    'election_id': candidate.election_id,
                    'scope': candidate.scope,
                    'photo': url_for('admin.static', filename='images/' + candidate.photo) if candidate.photo else None
                }
            }
        else:
            response_data = {
                'success': True,
                'message': 'Candidate updated successfully!',
                'candidate': {
                    'id': candidate.id,
                    'first_name': candidate.first_name,
                    'last_name': candidate.last_name,
                    'candidate_type': 'student',
                    'party_list': candidate.party_list,
                    'platform': candidate.platform,
                    'department': candidate.department.name if candidate.department else '',
                    'department_id': candidate.department_id,
                    'course_id': candidate.course_id,
                    'year_level_id': candidate.year_level_id,
                    'year_level': candidate.year_level.year_name if candidate.year_level else '',
                    'position': candidate.position.name if candidate.position else '',
                    'position_id': candidate.position_id,
                    'election_id': candidate.election_id,
                    'scope': candidate.scope,
                    'photo': url_for('admin.static', filename='images/' + candidate.photo) if candidate.photo else None
                }
            }
        return jsonify(response_data)
    
    flash('Candidate updated successfully!', 'success')
    return redirect(url_for('admin.manage_candidates'))


@admin_bp.route('/candidates/delete/<int:id>', methods=['POST'])
@admin_required
def delete_candidate(id):
    candidate = Candidate.query.get_or_404(id)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Store candidate info for audit log before deletion
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    
    if candidate.candidate_type == 'studio':
        candidate_name = candidate.studio_name or 'Unknown Studio'
        action_type = 'DELETE_STUDIO_CANDIDATE'
    else:
        candidate_name = f"{candidate.first_name} {candidate.last_name}"
        action_type = 'DELETE_CANDIDATE'
    
    party_list = candidate.party_list if candidate.party_list else 'Independent'
    position_name = candidate.position.name if candidate.position else 'N/A'
    department_name = candidate.department.name if candidate.department else 'N/A'
    year_level_name = candidate.year_level.year_name if candidate.year_level else 'N/A'
    election_title = candidate.election.title if candidate.election else 'N/A'
    
    try:
        # FIRST: Manually delete related tally_votes records
        from admin.models import TallyVote
        
        # Delete all tally votes for this candidate
        tally_votes = TallyVote.query.filter_by(candidate_id=id).all()
        for tv in tally_votes:
            db.session.delete(tv)
        
        # Now delete the candidate
        db.session.delete(candidate)
        db.session.commit()
        
        # Audit log
        log_audit(
            action=action_type,
            description=f"Admin user '{username}' deleted candidate from IP: {ip} | Candidate: {candidate_name} | Party: {party_list} | Position: {position_name} | Department: {department_name} | Year Level: {year_level_name} | Election: {election_title}"
        )
        
        # Always return JSON for AJAX requests
        if is_ajax:
            return jsonify({
                'success': True,
                'message': 'Candidate deleted successfully!'
            })
        
        flash('Candidate deleted successfully!', 'success')
        return redirect(url_for('admin.manage_candidates'))
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting candidate: {str(e)}")
        import traceback
        traceback.print_exc()
        
        if is_ajax:
            return jsonify({
                'success': False,
                'message': f'Error deleting candidate: {str(e)}'
            }), 500
        flash(f'Error deleting candidate: {str(e)}', 'error')
        return redirect(url_for('admin.manage_candidates'))


@admin_bp.route('/courses/by_department/<int:department_id>')
@admin_required
def get_courses_by_department(department_id):
    """Get all courses for a specific department"""
    try:
        courses = Course.query.filter_by(department_id=department_id).order_by(Course.course_name).all()
        
        courses_data = [{
            'id': c.id,
            'course_name': c.course_name,
            'course_code': c.course_code
        } for c in courses]
        
        return jsonify({
            'success': True,
            'courses': courses_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@admin_bp.route('/validate-student', methods=['GET'])
@admin_required
def validate_student():
    """Validate if a student exists in the student table before adding as candidate"""
    first_name = request.args.get('first_name', '').strip()
    last_name = request.args.get('last_name', '').strip()
    course_id = request.args.get('course_id', type=int)
    
    if not first_name or not last_name:
        return jsonify({
            'exists': False,
            'message': 'Please provide both first name and last name.'
        })
    
    # Build query to find matching students
    query = Student.query.filter(
        Student.first_name.ilike(first_name),
        Student.last_name.ilike(last_name)
    )
    
    # If course_id is provided, filter by course
    if course_id:
        query = query.filter(Student.course_id == course_id)
    
    students = query.all()
    
    if students:
        # Student exists
        student = students[0]  # Take the first match
        return jsonify({
            'exists': True,
            'message': f'Student found: {student.first_name} {student.last_name}',
            'student': {
                'id': student.id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'course_id': student.course_id,
                'course_name': student.course_rel.course_name if student.course_rel else None,
                'department_id': student.department_id,
                'department_name': student.department.name if student.department else None,
                'year_level_id': student.year_level_id,
                'year_level': student.year_level.year_name if student.year_level else None
            }
        })
    else:
        # Student not found - try to find similar names for suggestions
        from sqlalchemy import or_
        
        # Find similar students (either first name or last name matches partially)
        similar = Student.query.filter(
            or_(
                Student.first_name.ilike(f'%{first_name}%'),
                Student.last_name.ilike(f'%{last_name}%')
            )
        ).limit(5).all()
        
        suggestions = []
        for s in similar:
            suggestions.append({
                'first_name': s.first_name,
                'last_name': s.last_name,
                'course_name': s.course_rel.course_name if s.course_rel else None
            })
        
        return jsonify({
            'exists': False,
            'message': f'No student found with name "{first_name} {last_name}". The student must register first.',
            'suggestions': suggestions
        })

        
# ---------------------- Manage Positions ---------------------- #
@admin_bp.route('/manage_positions', methods=['GET', 'POST'])
@admin_required
def manage_positions():
    if request.method == 'POST':
        position_name = request.form.get('position_name', '').strip()
        position_color = request.form.get('position_color', '#3498db').strip()
        
        if not position_color.startswith('#'):
            position_color = f'#{position_color}'
        
        if position_name:
            existing = Position.query.filter_by(name=position_name).first()
            if existing:
                # 🚫 REMOVED: ADD_POSITION_DUPLICATE audit log (no data modification)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({"success": False, "message": f'Position "{position_name}" already exists!'})
                flash(f'Position "{position_name}" already exists!', 'warning')
            else:
                new_position = Position(name=position_name, color=position_color)
                db.session.add(new_position)
                db.session.commit()
                
                # ✅ KEEP THIS AUDIT LOG (ADD POSITION - data modification)
                username = getattr(current_user, 'username', 'Unknown')
                ip = request.remote_addr
                log_audit(
                    action='ADD_POSITION',
                    description=f"Admin user '{username}' added new position: '{position_name}' with color: {position_color} from IP: {ip}"
                )
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({"success": True, "message": f'Position "{position_name}" added successfully!'})
                flash(f'Position "{position_name}" added successfully!', 'success')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"success": True, "message": f'Position "{position_name}" added successfully!'})
        return redirect(url_for('admin.manage_positions'))
    
    positions = Position.query.all()
    
    # Return JSON for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        positions_data = [{"id": p.id, "name": p.name, "color": p.color} for p in positions]
        
        # 🚫 REMOVED: GET_POSITIONS_DATA audit log (AJAX data fetch - not a data modification)
        return jsonify(positions_data)
    
    # 🚫 REMOVED: MANAGE_POSITIONS_VIEW audit log (page view - not a data modification)
    
    return render_template('manage_positions.html', positions=positions)


# Get positions data (AJAX endpoint)
@admin_bp.route('/manage_positions/data')
@admin_required
def get_positions_data():
    positions = Position.query.all()
    positions_data = [{"id": p.id, "name": p.name, "color": p.color} for p in positions]
    
    # 🚫 REMOVED: GET_POSITIONS_DATA_ENDPOINT audit log (AJAX data fetch - not a data modification)
    
    return jsonify(positions_data)


# Update position
@admin_bp.route('/manage_positions/<int:position_id>', methods=['PUT'])
@admin_required
def update_position(position_id):
    position = Position.query.get_or_404(position_id)
    old_name = position.name
    old_color = position.color
    data = request.get_json()
    
    if not data or 'position_name' not in data:
        return jsonify({"success": False, "message": "Position name is required"}), 400
    
    position_name = data['position_name'].strip()
    position_color = data.get('position_color', '#3498db').strip()
    
    if not position_color.startswith('#'):
        position_color = f'#{position_color}'
    
    if not position_name:
        return jsonify({"success": False, "message": "Position name cannot be empty"}), 400
    
    # Check if position name already exists (excluding current position)
    existing = Position.query.filter(Position.name == position_name, Position.id != position_id).first()
    if existing:
        # 🚫 REMOVED: UPDATE_POSITION_DUPLICATE audit log (no data modification happened)
        return jsonify({"success": False, "message": f'Position "{position_name}" already exists!'})
    
    position.name = position_name
    position.color = position_color
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (UPDATE POSITION - data modification)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    log_audit(
        action='UPDATE_POSITION',
        description=f"Admin user '{username}' updated position from '{old_name}' (color: {old_color}) to '{position_name}' (color: {position_color}) (ID: {position_id}) from IP: {ip}"
    )
    
    return jsonify({"success": True, "message": f'Position "{position_name}" updated successfully!'})


# Delete position
@admin_bp.route('/manage_positions/<int:position_id>', methods=['DELETE'])
@admin_required
def delete_position(position_id):
    position = Position.query.get_or_404(position_id)
    position_name = position.name
    
    # Check if position is being used by any candidate
    candidates_using = Candidate.query.filter_by(position_id=position_id).first()
    if candidates_using:
        # 🚫 REMOVED: DELETE_POSITION_IN_USE audit log (no data modification happened)
        return jsonify({"success": False, "message": f'Cannot delete position "{position.name}" because it is being used by candidates.'})
    
    db.session.delete(position)
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (DELETE POSITION - data modification)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    log_audit(
        action='DELETE_POSITION',
        description=f"Admin user '{username}' deleted position: '{position_name}' (ID: {position_id}) from IP: {ip}"
    )
    
    return jsonify({"success": True, "message": f'Position "{position_name}" deleted successfully!'})


from admin.models import ElectionPosition
@admin_bp.route('/configure-election-positions/<int:election_id>', methods=['GET', 'POST'])
@admin_required
def configure_election_positions(election_id):
    """Configure which positions are in an election and their vote limits"""
    # Get the election first
    election = Election.query.get_or_404(election_id)
    
    # Get all positions
    all_positions = Position.query.order_by(Position.name).all()
    
    # Get all courses for dropdown
    all_courses = Course.query.order_by(Course.course_name).all()
    
    # Get all program types (Day/Night)
    from student.models import ProgramType
    program_types = ProgramType.query.order_by(ProgramType.name).all()
    
    # Year level options
    year_levels = [
        {'value': 1, 'label': '1st Year'},
        {'value': 2, 'label': '2nd Year'},
        {'value': 3, 'label': '3rd Year'},
        {'value': 4, 'label': '4th Year'}
    ]
    
    # Get currently configured positions for this election
    configured_positions = ElectionPosition.query.filter_by(election_id=election_id).all()
    configured_position_ids = [ep.position_id for ep in configured_positions]
    configured_positions_dict = {ep.position_id: ep.max_votes for ep in configured_positions}
    
    # Get course restrictions for configured positions
    position_courses = {ep.position_id: ep.course_id for ep in configured_positions if ep.course_id}
    
    # Get program type restrictions for configured positions
    position_program_types = {ep.position_id: ep.program_type_id for ep in configured_positions if ep.program_type_id}
    
    # Get year level restrictions for configured positions
    position_year_levels = {ep.position_id: ep.year_level for ep in configured_positions if ep.year_level}
    
    if request.method == 'POST':
        # ========== DEBUGGING: Print everything ==========
        import sys
        print("=" * 60, file=sys.stderr)
        print("CONFIGURE POSITIONS - POST REQUEST RECEIVED", file=sys.stderr)
        print("=" * 60, file=sys.stderr)
        print(f"Election ID: {election_id}", file=sys.stderr)
        print(f"Election Title: {election.title}", file=sys.stderr)
        print(f"Form keys: {list(request.form.keys())}", file=sys.stderr)
        print(f"Selected positions: {request.form.getlist('positions')}", file=sys.stderr)
        
        # Get form data
        selected_positions = request.form.getlist('positions')
        
        # Print each position's data
        for pos_id in selected_positions:
            max_votes = request.form.get(f'max_votes_{pos_id}')
            course_id = request.form.get(f'course_{pos_id}')
            program_type_id = request.form.get(f'program_type_{pos_id}')
            year_level = request.form.get(f'year_level_{pos_id}')
            print(f"Position {pos_id}: max_votes={max_votes}, course_id={course_id}, program_type_id={program_type_id}, year_level={year_level}", file=sys.stderr)
        
        print("=" * 60, file=sys.stderr)
        # ========== END DEBUGGING ==========
        
        # Validate: At least one position must be selected
        if not selected_positions:
            flash('Please select at least one position for this election.', 'error')
            return redirect(url_for('admin.configure_election_positions', election_id=election_id))
        
        try:
            # Delete existing configurations
            ElectionPosition.query.filter_by(election_id=election_id).delete()
            
            # Add new configurations
            display_order = 0
            for position_id_str in selected_positions:
                position_id = int(position_id_str)
                max_votes = request.form.get(f'max_votes_{position_id}', type=int, default=1)
                
                # Validate max votes range
                if max_votes < 1:
                    max_votes = 1
                elif max_votes > 50:
                    flash(f'Maximum votes cannot exceed 50 for a position. Please check your configuration.', 'error')
                    return redirect(url_for('admin.configure_election_positions', election_id=election_id))
                
                # Get course restriction
                course_id = request.form.get(f'course_{position_id}', type=int)
                
                # Get program type restriction
                program_type_id = request.form.get(f'program_type_{position_id}', type=int)
                
                # Get year level restriction
                year_level_str = request.form.get(f'year_level_{position_id}')
                year_level = int(year_level_str) if year_level_str and year_level_str != '' else None
                
                ep = ElectionPosition(
                    election_id=election_id,
                    position_id=position_id,
                    max_votes=max_votes,
                    min_votes=1,  # Default minimum
                    course_id=course_id if course_id else None,
                    program_type_id=program_type_id if program_type_id else None,
                    year_level=year_level,
                    department_id=None,  # Always None - department restriction removed
                    display_order=display_order
                )
                db.session.add(ep)
                display_order += 1
            
            db.session.commit()
            
            # Audit log
            username = getattr(current_user, 'username', 'Unknown')
            ip = request.remote_addr
            
            # Get position names for better audit description
            position_names = []
            for pid in selected_positions:
                position = Position.query.get(int(pid))
                if position:
                    position_names.append(position.name)
            
            position_summary = ', '.join(position_names[:3])
            if len(position_names) > 3:
                position_summary += f" and {len(position_names) - 3} more"
            
            log_audit(
                action='CONFIGURE_ELECTION_POSITIONS',
                description=f"Admin user '{username}' configured {len(selected_positions)} positions for election '{election.title}' (ID: {election_id}) from IP: {ip} | Positions: {position_summary}"
            )
            
            flash('Election positions configured successfully!', 'success')
            print("SUCCESS: Configuration saved!", file=sys.stderr)
            
        except Exception as e:
            db.session.rollback()
            print(f"ERROR: {str(e)}", file=sys.stderr)
            flash(f'Error configuring positions: {str(e)}', 'error')
        
        # Redirect back to the same page
        return redirect(url_for('admin.configure_election_positions', election_id=election_id))
    
    return render_template(
        'configure_election_positions.html',
        election=election,
        all_positions=all_positions,
        all_courses=all_courses,
        program_types=program_types,
        year_levels=year_levels,
        configured_position_ids=configured_position_ids,
        configured_positions=configured_positions_dict,
        position_courses=position_courses,
        position_program_types=position_program_types,
        position_year_levels=position_year_levels
    )
    


# admin/routes.py - Update your create_election route

@admin_bp.route('/create-election', methods=['GET', 'POST'])
@admin_required
def create_election():
    """
    IMPROVED CREATE ELECTION ROUTE
    - Handles both campus and department elections
    - Adds year level filtering for campus elections
    - Redirects to position configuration after creation
    - Filters election lists by selected year from dashboard
    """
    # Clear any existing flash messages from other pages
    session.pop('_flashes', None)
    
    # Get all departments
    departments = Department.query.order_by(Department.name).all()

    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        scope = request.form.get('scope', '').strip().lower()
        department_id_str = request.form.get('department_id')
        description = request.form.get('description', '').strip()
        start_date_str = request.form.get('start_date', '').strip()
        end_date_str = request.form.get('end_date', '').strip()
        year_levels = request.form.getlist('year_levels')
        
        # Validation
        if not all([title, scope, start_date_str, end_date_str]):
            flash('All required fields must be filled.', 'election-error')
            return redirect(url_for('admin.create_election'))

        if scope not in ['campus', 'department']:
            flash('Invalid election scope. Must be campus or department.', 'election-error')
            return redirect(url_for('admin.create_election'))

        try:
            start_date = tz.localize(datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M'))
            end_date = tz.localize(datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M'))
        except ValueError:
            flash('Invalid date format.', 'election-error')
            return redirect(url_for('admin.create_election'))

        if end_date <= start_date:
            flash('End date must be later than start date.', 'election-error')
            return redirect(url_for('admin.create_election'))

        department_id = None
        department_name = None
        
        if scope == 'department':
            if not department_id_str:
                flash('Department is required for Department Elections.', 'election-error')
                return redirect(url_for('admin.create_election'))
            
            department_id = int(department_id_str)
            dept_obj = Department.query.get(department_id)
            if not dept_obj:
                flash('Selected department does not exist.', 'election-error')
                return redirect(url_for('admin.create_election'))
            department_name = dept_obj.name
        
        year_levels_str = 'all'
        if scope == 'campus' and year_levels:
            year_levels.sort()
            year_levels_str = ','.join(year_levels)
        
        election_type = 'SSG' if scope == 'campus' else 'Department'

        new_election = Election(
            title=title,
            election_type=election_type,
            scope=scope,
            department_id=department_id,
            department=department_name,
            year_levels=year_levels_str,
            description=description,
            start_date=start_date,
            end_date=end_date
        )
        
        db.session.add(new_election)
        db.session.commit()
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        year_levels_display = 'All Years' if year_levels_str == 'all' else f"Year(s) {year_levels_str}"
        
        log_audit(
            action='CREATE_ELECTION',
            description=f"Admin user '{username}' created new {scope} election: '{title}' from IP: {ip} | Department: {department_name or 'N/A'}, Target: {year_levels_display}, Start: {start_date.strftime('%Y-%m-%d %H:%M')}, End: {end_date.strftime('%Y-%m-%d %H:%M')}"
        )
        
        flash('Election created successfully! Now configure positions and vote limits.', 'election-success')
        return redirect(url_for('admin.configure_election_positions', election_id=new_election.id))

    # ========== GET request: fetch elections filtered by selected year ==========
    # Get year filter from session (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    # Build election query with year filter
    election_query = Election.query.order_by(Election.start_date.desc())
    
    if start_date and end_date:
        election_query = election_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    # Get all elections filtered by year
    elections_all = election_query.all()

    # Ensure datetime are timezone-aware
    for e in elections_all:
        if e.start_date.tzinfo is None:
            e.start_date = tz.localize(e.start_date)
        if e.end_date.tzinfo is None:
            e.end_date = tz.localize(e.end_date)

    # Filter elections by status (upcoming, active, ended)
    upcoming_elections = [e for e in elections_all if e.start_date > now]
    active_elections = [e for e in elections_all if e.start_date <= now <= e.end_date]
    # Ended elections are the rest (elections_all minus upcoming and active)

    return render_template(
        'create_election.html',
        departments=departments,
        upcoming=upcoming_elections,
        active=active_elections,
        elections_all=elections_all,
        current_year=year,  # Pass current year to template for display if needed
        now=now
    )


# Keep the old route for backward compatibility
@admin_bp.route('/create-department-election', methods=['GET', 'POST'])
@admin_required
def create_department_election():
    """Legacy route - redirects to new unified create election page"""
    return redirect(url_for('admin.create_election'))


# ----------- ANNOUNCEMENTS ROUTE -----------
@admin_bp.route('/announcements', methods=['GET', 'POST'])
@login_required
def announcements():
    # Get year filter from session (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    departments = Department.query.all()  # For dropdown
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)  # Get current datetime

    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        date = request.form.get('date')
        department_id = request.form.get('department')

        if department_id == "all":
            department_id = None  # None means visible to all

        new_announcement = Announcement(
            title=title,
            content=content,
            date=date,
            department_id=department_id
        )
        db.session.add(new_announcement)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (CREATE action - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        # Get department name for audit log
        department_name = "All Departments"
        if department_id:
            dept = Department.query.get(department_id)
            department_name = dept.name if dept else f"ID: {department_id}"
        
        log_audit(
            action='CREATE_ANNOUNCEMENT',
            description=f"Admin user '{username}' created new announcement: '{title}' from IP: {ip} | Target: {department_name}, Date: {date}"
        )

        # Tag the flash message for announcements page
        flash('announcements_page:Announcement created successfully!', 'success')
        return redirect(url_for('admin.announcements'))

    # GET: fetch announcements to display with year filter
    announcements_query = Announcement.query
    
    # Apply year filter from dashboard
    if start_date and end_date:
        announcements_query = announcements_query.filter(
            Announcement.date >= start_date,
            Announcement.date <= end_date
        )
    
    announcements_list = announcements_query.order_by(Announcement.created_at.desc()).all()
    
    # 🚫 REMOVED: ANNOUNCEMENTS_VIEW audit log (not a data modification)
    # Viewing the page should not be logged

    return render_template(
        'announcements.html', 
        departments=departments, 
        announcements=announcements_list,
        now=now,  # Pass current datetime to template
        current_year=year,  # Changed from current_sy to current_year
        start_date=start_date,
        end_date=end_date
    )


# ----------- UPDATE ANNOUNCEMENT -----------
@admin_bp.route('/update-announcement/<int:announcement_id>', methods=['POST'])
@login_required
def update_announcement(announcement_id):
    announcement = Announcement.query.get_or_404(announcement_id)
    
    title = request.form.get('title')
    content = request.form.get('content')
    date = request.form.get('date')
    department_id = request.form.get('department')
    
    if department_id == "all":
        department_id = None
    else:
        department_id = int(department_id)
    
    # Store old values for audit log
    old_title = announcement.title
    old_department_id = announcement.department_id
    
    announcement.title = title
    announcement.content = content
    announcement.date = datetime.strptime(date, '%Y-%m-%d').date()
    announcement.department_id = department_id
    
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (UPDATE action - data modification)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    
    # Get department names for audit log
    old_dept_name = "All Departments"
    if old_department_id:
        old_dept = Department.query.get(old_department_id)
        old_dept_name = old_dept.name if old_dept else f"ID: {old_department_id}"
    
    new_dept_name = "All Departments"
    if department_id:
        new_dept = Department.query.get(department_id)
        new_dept_name = new_dept.name if new_dept else f"ID: {department_id}"
    
    log_audit(
        action='UPDATE_ANNOUNCEMENT',
        description=f"Admin user '{username}' updated announcement: '{old_title}' → '{title}' (ID: {announcement_id}) from IP: {ip} | Department: {old_dept_name} → {new_dept_name}"
    )
    
    # Tag the flash message for announcements page
    flash('announcements_page:Announcement updated successfully!', 'success')
    return redirect(url_for('admin.announcements'))


# ----------- DELETE ANNOUNCEMENT -----------
@admin_bp.route('/delete-announcement/<int:announcement_id>', methods=['POST'])
@login_required
def delete_announcement(announcement_id):
    announcement = Announcement.query.get_or_404(announcement_id)
    title = announcement.title
    department_id = announcement.department_id
    
    # Get department name for audit log
    department_name = "All Departments"
    if department_id:
        dept = Department.query.get(department_id)
        department_name = dept.name if dept else f"ID: {department_id}"
    
    db.session.delete(announcement)
    db.session.commit()
    
    # ✅ KEEP THIS AUDIT LOG (DELETE action - data modification)
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    
    log_audit(
        action='DELETE_ANNOUNCEMENT',
        description=f"Admin user '{username}' deleted announcement: '{title}' (ID: {announcement_id}) from IP: {ip} | Was targeted to: {department_name}"
    )
    
    # Tag the flash message for announcements page
    flash('announcements_page:Announcement deleted successfully!', 'success')
    return redirect(url_for('admin.announcements'))


from student.models import ContactInfo, HelpPageContent

# ----------- GET ANNOUNCEMENT FOR EDIT -----------
@admin_bp.route('/get-announcement/<int:announcement_id>', methods=['GET'])
@login_required
def get_announcement(announcement_id):
    try:
        announcement = Announcement.query.get_or_404(announcement_id)
        
        # Format the date properly
        date_str = announcement.date.strftime('%Y-%m-%d') if announcement.date else ''
        
        return jsonify({
            'success': True,
            'announcement': {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'date': date_str,
                'department_id': announcement.department_id
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ----------- HELP PAGE MANAGEMENT -----------
@admin_bp.route('/help-settings', methods=['GET', 'POST'])
@login_required
def help_settings():
    contact_info = ContactInfo.get_settings()
    help_content = HelpPageContent.get_content()
    
    if request.method == 'POST':
        # Store old values for audit log
        old_email = contact_info.email
        old_phone = contact_info.phone
        old_committee_name = contact_info.committee_name
        
        # Update contact info
        contact_info.email = request.form.get('email')
        contact_info.phone = request.form.get('phone')
        contact_info.committee_name = request.form.get('committee_name')
        contact_info.additional_info = request.form.get('additional_info')
        contact_info.updated_by = current_user.id
        
        # Update help content
        old_common_issues = help_content.common_issues
        help_content.common_issues = request.form.get('common_issues')
        help_content.updated_by = current_user.id
        
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (UPDATE action - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        log_audit(
            action='UPDATE_HELP_SETTINGS',
            description=f"Admin user '{username}' updated help page settings from IP: {ip} | Email: {old_email} → {contact_info.email}, Committee: {old_committee_name} → {contact_info.committee_name}"
        )
        
        flash('announcements_page:Help page settings updated successfully!', 'success')
        return redirect(url_for('admin.help_settings'))
    
    # 🚫 REMOVED: VIEW audit log for help settings page
    
    return render_template('help_settings.html', 
                         contact=contact_info, 
                         help_content=help_content)


# ----------- GET HELP SETTINGS FOR AJAX -----------
@admin_bp.route('/get-help-settings')
@login_required
def get_help_settings():
    contact_info = ContactInfo.get_settings()
    help_content = HelpPageContent.get_content()
    
    # 🚫 REMOVED: VIEW audit log for AJAX endpoint
    
    return jsonify({
        'success': True,
        'contact': {
            'email': contact_info.email,
            'phone': contact_info.phone,
            'committee_name': contact_info.committee_name,
            'additional_info': contact_info.additional_info
        },
        'help_content': {
            'common_issues': help_content.common_issues
        }
    })


# ----------- GUIDELINES MANAGEMENT -----------
@admin_bp.route('/guidelines-settings', methods=['GET', 'POST'])
@login_required
def guidelines_settings():
    from student.models import GuidelinesContent
    
    guidelines_content = GuidelinesContent.get_content()
    
    if request.method == 'POST':
        # Store old values for audit log
        old_purpose = guidelines_content.purpose
        old_voting_rules = guidelines_content.voting_rules
        
        # Update guidelines content
        guidelines_content.purpose = request.form.get('purpose')
        guidelines_content.voting_rules = request.form.get('voting_rules')
        guidelines_content.how_to_vote = request.form.get('how_to_vote')
        guidelines_content.privacy_security = request.form.get('privacy_security')
        guidelines_content.important_reminders = request.form.get('important_reminders')
        guidelines_content.fingerprint_info = request.form.get('fingerprint_info')
        
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (UPDATE action - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        log_audit(
            action='UPDATE_GUIDELINES',
            description=f"Admin user '{username}' updated voting guidelines from IP: {ip}"
        )
        
        flash('announcements_page:Guidelines updated successfully!', 'success')
        return redirect(url_for('admin.guidelines_settings'))
    
    # 🚫 REMOVED: VIEW audit log for guidelines page
    
    return render_template('guidelines_settings.html', content=guidelines_content)


from student.models import GuidelinesContent
# ----------- GET GUIDELINES FOR AJAX -----------
@admin_bp.route('/get-guidelines')
@login_required
def get_guidelines():
    from student.models import GuidelinesContent
    
    guidelines_content = GuidelinesContent.get_content()
    
    # 🚫 REMOVED: VIEW audit log for AJAX endpoint
    
    return jsonify({
        'success': True,
        'content': {
            'purpose': guidelines_content.purpose,
            'voting_rules': guidelines_content.voting_rules,
            'how_to_vote': guidelines_content.how_to_vote,
            'privacy_security': guidelines_content.privacy_security,
            'important_reminders': guidelines_content.important_reminders,
            'fingerprint_info': guidelines_content.fingerprint_info
        }
    })


from admin.models import PdfResult
@admin_bp.route('/results')
@admin_required
def results_page():
    # Get year filter from session (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)
    
    # Filter elections by year if set
    election_query = Election.query.order_by(Election.end_date.desc())
    
    if start_date and end_date:
        # Convert to timezone-aware for comparison
        start_date_aware = tz.localize(start_date) if start_date.tzinfo is None else start_date
        end_date_aware = tz.localize(end_date) if end_date.tzinfo is None else end_date
        
        election_query = election_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    elections = election_query.all()
    
    upcoming, active, completed = [], [], []
    
    for election in elections:
        # Create timezone-aware copies for comparison
        start_date_e = election.start_date
        end_date_e = election.end_date
        
        # Convert to timezone-aware if naive
        if start_date_e.tzinfo is None:
            start_date_e = tz.localize(start_date_e)
        if end_date_e.tzinfo is None:
            end_date_e = tz.localize(end_date_e)
        
        # Add timezone-aware attributes to election object for template use
        election.tz_start = start_date_e
        election.tz_end = end_date_e
        
        # ============ NEW: Check if result is published (has PDF) ============
        from admin.models import PdfResult  # Make sure this import is at the top of your file
        election.has_published_result = PdfResult.query.filter_by(election_id=election.id).first() is not None
        # ====================================================================
        
        # Categorize based on the dates
        if end_date_e < now:
            completed.append(election)
        elif start_date_e <= now <= end_date_e:
            active.append(election)
        else:
            upcoming.append(election)
    
    # Get total elections count (unfiltered) for context
    total_elections = Election.query.count()
    
    return render_template(
        'admin_results.html',
        upcoming_elections=upcoming,
        active_elections=active,
        completed_elections=completed,
        now=now,
        current_year=year,
        total_filtered=len(elections),
        total_elections=total_elections
    )


@admin_bp.route('/results/<int:election_id>')
@admin_required
def election_results(election_id):
    """OPTIMIZED: Uses finder_hashes for live results, TallyVote for official results"""
    import json
    from collections import defaultdict
    from datetime import datetime as dt
    
    election = Election.query.get_or_404(election_id)
    
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)
    
    # Handle timezone for start_date and end_date
    if election.start_date.tzinfo is None:
        election.start_date = tz.localize(election.start_date)
    if election.end_date.tzinfo is None:
        election.end_date = tz.localize(election.end_date)
    
    if election.end_date < now:
        status = "Completed"
    elif election.start_date <= now <= election.end_date:
        status = "Active"
    else:
        status = "Upcoming"
    
    # Check if election has been officially tallied
    is_tallied = False
    tally_timestamp = None
    
    if TALLY_VOTE_AVAILABLE:
        tally_record = TallyVote.query.filter_by(election_id=election_id).first()
        is_tallied = tally_record is not None
        if is_tallied:
            latest_tally = TallyVote.query.filter_by(
                election_id=election_id
            ).order_by(TallyVote.tally_timestamp.desc()).first()
            if latest_tally and latest_tally.tally_timestamp:
                # Convert UTC timestamp to Manila timezone
                tally_utc = latest_tally.tally_timestamp
                # If the timestamp is naive (no timezone), assume it's UTC
                if tally_utc.tzinfo is None:
                    tally_utc = pytz.UTC.localize(tally_utc)
                # Convert to Manila timezone
                tally_timestamp = tally_utc.astimezone(tz)
    
    # Get all candidates
    candidates = Candidate.query.filter_by(election_id=election_id).all()
    
    # GET POSITION LIMITS AND RESTRICTIONS
    position_limits = {}
    election_positions = ElectionPosition.query.filter_by(election_id=election_id).all()
    
    # Get max votes for each position
    for ep in election_positions:
        position_limits[ep.position_id] = ep.max_votes
    
    # GET ALL VOTES
    all_votes = Vote.query.filter_by(election_id=election_id).all()
    
    # ===== GET VOTE COUNTS =====
    if is_tallied and TALLY_VOTE_AVAILABLE:
        print("📊 Using official tally results")
        tally_records = TallyVote.query.filter_by(election_id=election_id).all()
        vote_counts = {t.candidate_id: t.vote_count for t in tally_records}
        for candidate in candidates:
            if candidate.id not in vote_counts:
                vote_counts[candidate.id] = 0
    else:
        print("📊 Using finder_hashes for live results")
        vote_counts = get_admin_live_vote_counts(election_id)
    
    # ===== HELPER FUNCTION: Get vote timeline for a candidate =====
    def get_candidate_vote_timeline(candidate_id, election_id):
        """Get all timestamps when votes were cast for a candidate"""
        all_votes = Vote.query.filter_by(election_id=election_id).all()
        timestamps = []
        
        for vote in all_votes:
            candidate_ids = vote.voted_candidate_ids
            if candidate_id in candidate_ids:
                # Use cast_timestamp (actual voting time) as primary
                if vote.cast_timestamp:
                    timestamps.append(vote.cast_timestamp)
                elif vote.created_at:
                    timestamps.append(vote.created_at)
        
        timestamps.sort()
        return timestamps
    
    # ===== CALCULATE VOTERS PER POSITION =====
    position_voter_counts = {}
    candidate_position_map = {c.id: c.position_id for c in candidates}
    
    for ep in election_positions:
        position_id = ep.position_id
        position_name = Position.query.get(position_id).name if Position.query.get(position_id) else "Unknown"
        
        voters_for_position = set()
        
        for vote in all_votes:
            if not vote.finder_hash:
                continue
                
            try:
                finder_data = json.loads(vote.finder_hash)
                candidate_ids = []
                
                if isinstance(finder_data, dict):
                    if 'hashes' in finder_data and isinstance(finder_data['hashes'], list):
                        for item in finder_data['hashes']:
                            if isinstance(item, dict) and 'candidate_id' in item:
                                candidate_ids.append(item['candidate_id'])
                    elif 'candidate_ids' in finder_data:
                        candidate_ids = finder_data['candidate_ids']
                elif isinstance(finder_data, list):
                    for item in finder_data:
                        if isinstance(item, dict) and 'candidate_id' in item:
                            candidate_ids.append(item['candidate_id'])
                
                # Check if ANY candidate in this vote belongs to this position
                for cid in candidate_ids:
                    cand_position_id = candidate_position_map.get(cid)
                    if cand_position_id == position_id:
                        # Count this voter for this position
                        if vote.student_id:
                            voters_for_position.add(vote.student_id)
                        else:
                            voters_for_position.add(f"anon_{vote.id}")
                        break  # Once we found this voter voted in this position, move to next vote
                        
            except (json.JSONDecodeError, Exception) as e:
                print(f"Error parsing vote {vote.id}: {e}")
                continue
        
        position_voter_counts[position_id] = len(voters_for_position)
        print(f"📊 Position {position_name} (ID: {position_id}): Voters = {len(voters_for_position)}")
    
    # ===== GROUP CANDIDATES BY POSITION =====
    candidates_by_position = {}
    
    for candidate in candidates:
        position_id = candidate.position_id
        position_name = candidate.position.name if candidate.position else "Unknown"
        
        if position_name not in candidates_by_position:
            max_votes = position_limits.get(position_id, 1)
            voter_count = position_voter_counts.get(position_id, 0)
            
            candidates_by_position[position_name] = {
                'id': position_id,
                'name': position_name,
                'max_votes': max_votes,
                'voter_count': voter_count,
                'candidates': []
            }
        
        vote_count = vote_counts.get(candidate.id, 0)
        
        candidates_by_position[position_name]['candidates'].append({
            'id': candidate.id,
            'first_name': candidate.first_name or "",
            'last_name': candidate.last_name or "",
            'studio_name': candidate.studio_name or "",
            'candidate_type': candidate.candidate_type or "student",
            'photo': candidate.photo or "",
            'party_list': candidate.party_list or "",
            'position': position_name,
            'position_id': position_id,
            'vote_count': vote_count,
            'voter_percentage': 0,
            'is_winner': False,
            'tie_broken': False,
            'reached_at': None
        })
    
    # ===== CALCULATE PERCENTAGES AND DETERMINE WINNERS WITH TIE-BREAKING =====
    for position_name, pos_data in candidates_by_position.items():
        voter_count = pos_data['voter_count']
        max_winners = pos_data['max_votes']
        
        print(f"\n📊 {position_name}: Max winners = {max_winners}, Voters = {voter_count}")
        
        # Calculate percentages
        if not voter_count or voter_count == 0:
            print(f"   ⚠️ WARNING: voter_count is 0 for {position_name} - setting percentages to 0")
            for candidate in pos_data['candidates']:
                candidate['voter_percentage'] = 0
        else:
            for candidate in pos_data['candidates']:
                percentage = (candidate['vote_count'] / voter_count) * 100
                candidate['voter_percentage'] = round(percentage, 2)
                print(f"   {candidate['studio_name'] if candidate['candidate_type'] == 'studio' else candidate['first_name']}: {candidate['vote_count']} votes = {candidate['voter_percentage']}%")
        
        # Sort candidates by vote count (descending)
        pos_data['candidates'].sort(key=lambda x: x['vote_count'], reverse=True)
        
        # Check for ties at the winner threshold
        tie_vote_count = None
        if len(pos_data['candidates']) > max_winners:
            last_winner_index = max_winners - 1
            last_winner_votes = pos_data['candidates'][last_winner_index]['vote_count']
            next_candidate_votes = pos_data['candidates'][max_winners]['vote_count']
            
            if last_winner_votes == next_candidate_votes and last_winner_votes > 0:
                tie_vote_count = last_winner_votes
                print(f"⚠️ TIE DETECTED for {position_name} at {tie_vote_count} votes!")
        
        if tie_vote_count:
            # Get all candidates tied for the last winner position
            tied_candidates = [c for c in pos_data['candidates'] 
                              if c['vote_count'] == tie_vote_count]
            
            # Get non-tied winners (those with more votes)
            non_tied_winners = [c for c in pos_data['candidates'] 
                               if c['vote_count'] > tie_vote_count]
            
            # For each tied candidate, find when they reached the tie_vote_count
            for candidate in tied_candidates:
                timestamps = get_candidate_vote_timeline(candidate['id'], election_id)
                if len(timestamps) >= tie_vote_count:
                    candidate['reached_at'] = timestamps[tie_vote_count - 1]
                elif timestamps:
                    candidate['reached_at'] = timestamps[-1]
                else:
                    candidate['reached_at'] = None
                
                candidate_display = candidate['studio_name'] if candidate['candidate_type'] == 'studio' else f"{candidate['first_name']} {candidate['last_name']}"
                print(f"   {candidate_display}: Reached {tie_vote_count} votes at {candidate['reached_at']}")
            
            # Sort tied candidates by who reached first (earliest timestamp wins)
            tied_candidates.sort(key=lambda x: x.get('reached_at') or dt.max)
            
            # Determine how many winners needed from tied group
            winners_needed = max_winners - len(non_tied_winners)
            
            # Mark all candidates as winner or loser
            for candidate in pos_data['candidates']:
                if candidate['vote_count'] > tie_vote_count:
                    candidate['is_winner'] = True
                    candidate['tie_broken'] = False
                elif candidate['vote_count'] == tie_vote_count:
                    # Check if this candidate is among the earliest to reach
                    if candidate in tied_candidates[:winners_needed]:
                        candidate['is_winner'] = True
                        candidate['tie_broken'] = True
                        candidate['tie_broken_by'] = 'first_to_reach'
                        candidate_display = candidate['studio_name'] if candidate['candidate_type'] == 'studio' else f"{candidate['first_name']} {candidate['last_name']}"
                        print(f"  ✓ WINNER (tie-breaker): {candidate_display} - {candidate['vote_count']} votes (reached first at {candidate['reached_at']})")
                    else:
                        candidate['is_winner'] = False
                        candidate['tie_broken'] = True
                else:
                    candidate['is_winner'] = False
                    candidate['tie_broken'] = False
            
            # Also mark non-tied winners
            for candidate in non_tied_winners:
                candidate_display = candidate['studio_name'] if candidate['candidate_type'] == 'studio' else f"{candidate['first_name']} {candidate['last_name']}"
                print(f"  ✓ WINNER: {candidate_display} - {candidate['vote_count']} votes ({candidate['voter_percentage']}%)")
        
        else:
            # No tie, normal winner selection
            for i, candidate in enumerate(pos_data['candidates']):
                if i < max_winners and candidate['vote_count'] > 0:
                    candidate['is_winner'] = True
                    candidate_display = candidate['studio_name'] if candidate['candidate_type'] == 'studio' else f"{candidate['first_name']} {candidate['last_name']}"
                    print(f"  ✓ WINNER: {candidate_display} - {candidate['vote_count']} votes ({candidate['voter_percentage']}%)")
                else:
                    candidate['is_winner'] = False
    
    # ===== SORT CANDIDATES_BY_POSITION BY POSITION ID =====
    sorted_candidates_by_position = {}
    sorted_position_items = sorted(candidates_by_position.items(), key=lambda x: x[1].get('id', 999))
    for position_name, pos_data in sorted_position_items:
        sorted_candidates_by_position[position_name] = pos_data
    
    # ===== BUILD WINNERS BY POSITION (MULTI-WINNER SUPPORT) =====
    winners_by_position = {}
    for position_name, pos_data in sorted_candidates_by_position.items():
        winners = [c for c in pos_data['candidates'] if c.get('is_winner')]
        if winners:
            winners_by_position[position_name] = winners
            winner_names = []
            for w in winners:
                if w['candidate_type'] == 'studio':
                    winner_names.append(w['studio_name'])
                else:
                    winner_names.append(f"{w['first_name']} {w['last_name']}")
            print(f"🏆 {position_name}: {len(winners)} winner(s) - {winner_names}")
    
    # ===== CREATE FLAT CANDIDATE RESULTS LIST (SORTED BY POSITION ID) =====
    flat_candidate_results = []
    for position_name, pos_data in sorted_candidates_by_position.items():
        sorted_candidates = sorted(pos_data['candidates'], key=lambda x: x['vote_count'], reverse=True)
        for candidate in sorted_candidates:
            flat_candidate_results.append(candidate)
    
    # ===== CALCULATE OVERALL STATISTICS =====
    total_votes_cast = len(all_votes)
    
    if election.department_id:
        total_eligible_voters = Student.query.filter_by(department_id=election.department_id).count()
    else:
        total_eligible_voters = Student.query.count()
    
    if election.scope == 'campus' and election.year_levels and election.year_levels != 'all':
        allowed_years = election.year_levels.split(',')
        total_eligible_voters = Student.query.filter(Student.year_level_id.in_(allowed_years)).count()
    
    voter_turnout = round((total_votes_cast / total_eligible_voters * 100), 2) if total_eligible_voters and total_eligible_voters > 0 else 0
    students_not_voted = (total_eligible_voters or 0) - total_votes_cast
    if students_not_voted < 0:
        students_not_voted = 0
    
    # ===== BUILD POSITION ELIGIBLE VOTERS DICTIONARY =====
    position_eligible_voters = {}
    for ep in election_positions:
        position_id = ep.position_id
        position_name = Position.query.get(position_id).name if Position.query.get(position_id) else "Unknown"
        position_eligible_voters[position_id] = {
            'count': position_voter_counts.get(position_id, 0),
            'name': position_name,
            'max_votes': position_limits.get(position_id, 1)
        }
    
    # ===== FINAL DEBUG OUTPUT =====
    print("\n" + "="*60)
    print("FINAL SORTED RESULTS BY POSITION ID:")
    for position_name, pos_data in sorted_candidates_by_position.items():
        print(f"\n📌 {position_name} (ID: {pos_data.get('id')}):")
        print(f"   Voter Count: {pos_data.get('voter_count')}")
        print(f"   Winners: {len([c for c in pos_data['candidates'] if c.get('is_winner')])}")
        for candidate in pos_data['candidates']:
            winner_status = "✓ WINNER" if candidate.get('is_winner') else "  Candidate"
            if candidate.get('tie_broken'):
                winner_status += " (tie-breaker)"
            candidate_display = candidate['studio_name'] if candidate['candidate_type'] == 'studio' else f"{candidate['first_name']} {candidate['last_name']}"
            print(f"   {winner_status}: {candidate_display} - {candidate['vote_count']} votes ({candidate['voter_percentage']}%)")
    print("="*60 + "\n")
    
    # Register helper functions for template
    app = current_app._get_current_object()
    app.jinja_env.globals.update(get_position_color=get_position_color)
    
    # Audit log
    username = getattr(current_user, 'username', 'Unknown')
    ip = request.remote_addr
    
    log_audit(
        action='ELECTION_RESULTS_VIEW',
        description=f"Admin user '{username}' viewed results for election: '{election.title}' (ID: {election_id}) from IP: {ip} | Status: {status}, Tallied: {is_tallied}"
    )
    
    return render_template(
        'election_results_detail.html',
        election=election,
        candidate_results=flat_candidate_results,
        candidate_results_by_position=sorted_candidates_by_position,
        winners_by_position=winners_by_position,
        total_voters=total_votes_cast,
        total_eligible_voters=total_eligible_voters or 0,
        voter_turnout=voter_turnout,
        students_not_voted=students_not_voted,
        status=status,
        now=now,
        is_tallied=is_tallied,
        tally_timestamp=tally_timestamp,
        position_eligible_voters=position_eligible_voters
    )


def get_admin_live_vote_counts(election_id):
    """
    ULTRA FAST: Get vote counts using finder_hashes for admin live results
    NO DECRYPTION! Takes 2-3 seconds even for thousands of voters
    """
    from collections import defaultdict
    
    vote_counts = defaultdict(int)
    
    # Stream votes in chunks to avoid memory issues
    batch_size = 1000
    offset = 0
    
    while True:
        batch = Vote.query.filter_by(election_id=election_id)\
                          .with_entities(Vote.finder_hash)\
                          .offset(offset)\
                          .limit(batch_size)\
                          .all()
        
        if not batch:
            break
        
        # Process each vote in the batch
        for vote in batch:
            if not vote.finder_hash:
                continue
                
            try:
                finder_data = json.loads(vote.finder_hash)
                
                # Extract candidate IDs based on format
                if isinstance(finder_data, dict):
                    # New format with 'hashes' array
                    if 'hashes' in finder_data and isinstance(finder_data['hashes'], list):
                        for item in finder_data['hashes']:
                            if isinstance(item, dict) and 'candidate_id' in item:
                                vote_counts[item['candidate_id']] += 1
                    
                    # Alternative format with candidate_ids directly
                    elif 'candidate_ids' in finder_data and isinstance(finder_data['candidate_ids'], list):
                        for cid in finder_data['candidate_ids']:
                            vote_counts[cid] += 1
                
                elif isinstance(finder_data, list):
                    # Old format
                    for item in finder_data:
                        if isinstance(item, dict) and 'candidate_id' in item:
                            vote_counts[item['candidate_id']] += 1
                    
            except json.JSONDecodeError:
                continue
            except Exception:
                continue
        
        offset += batch_size
    
    return dict(vote_counts)
    


@admin_bp.route('/results/<int:election_id>/check-new-votes')
@admin_required
def check_new_votes(election_id):
    """Check if there are new votes since last view - tracks per-candidate changes"""
    try:
        import json
        from collections import defaultdict
        
        election = Election.query.get_or_404(election_id)
        
        # Get current total voters (unique students who voted)
        unique_voters = db.session.query(Vote.student_id).filter_by(
            election_id=election_id
        ).distinct().count()
        
        # Get candidates for this election
        candidates = Candidate.query.filter_by(election_id=election_id).all()
        
        # Get current vote counts using finder_hash
        current_vote_counts = get_admin_live_vote_counts(election_id)
        
        # Get previous vote counts from session or use defaults
        previous_vote_counts = session.get(f'prev_vote_counts_{election_id}', {})
        
        # ===== GET ALL VOTES TO CALCULATE POSITION VOTER COUNTS =====
        all_votes = Vote.query.filter_by(election_id=election_id).all()
        
        # Build candidate position map
        candidate_position_map = {c.id: c.position_id for c in candidates}
        
        # Calculate ACTUAL voters per position (students who voted in that position)
        position_voter_counts = defaultdict(int)
        position_voters_set = defaultdict(set)
        
        for vote in all_votes:
            if not vote.finder_hash:
                continue
                
            try:
                finder_data = json.loads(vote.finder_hash)
                candidate_ids = []
                
                if isinstance(finder_data, dict):
                    if 'hashes' in finder_data and isinstance(finder_data['hashes'], list):
                        for item in finder_data['hashes']:
                            if isinstance(item, dict) and 'candidate_id' in item:
                                candidate_ids.append(item['candidate_id'])
                    elif 'candidate_ids' in finder_data:
                        candidate_ids = finder_data['candidate_ids']
                elif isinstance(finder_data, list):
                    for item in finder_data:
                        if isinstance(item, dict) and 'candidate_id' in item:
                            candidate_ids.append(item['candidate_id'])
                
                # Track which positions this voter voted in
                voter_positions = set()
                for cid in candidate_ids:
                    pos_id = candidate_position_map.get(cid)
                    if pos_id:
                        voter_positions.add(pos_id)
                
                # Add voter to position sets
                voter_key = vote.student_id if vote.student_id else f"anon_{vote.id}"
                for pos_id in voter_positions:
                    position_voters_set[pos_id].add(voter_key)
                    
            except (json.JSONDecodeError, Exception):
                continue
        
        # Convert sets to counts
        for pos_id, voters in position_voters_set.items():
            position_voter_counts[pos_id] = len(voters)
        
        # Calculate which candidates got new votes
        candidates_with_new_votes = []
        for candidate in candidates:
            current = current_vote_counts.get(candidate.id, 0)
            previous = previous_vote_counts.get(candidate.id, 0)
            
            if current > previous:
                candidates_with_new_votes.append({
                    'id': candidate.id,
                    'name': f"{candidate.first_name} {candidate.last_name}",
                    'new_votes': current - previous,
                    'position': candidate.position.name if candidate.position else "N/A"
                })
        
        # Store current counts for next comparison
        session[f'prev_vote_counts_{election_id}'] = current_vote_counts
        
        # Build candidate results with CORRECT percentages using ACTUAL position voters
        candidate_results = []
        for candidate in candidates:
            vote_count = current_vote_counts.get(candidate.id, 0)
            # FIXED: Use ACTUAL voters for this position, not total eligible
            actual_voters = position_voter_counts.get(candidate.position_id, 1)
            
            # Calculate percentage based on ACTUAL voters who voted in this position
            voter_percentage = round((vote_count / actual_voters * 100), 2) if actual_voters > 0 else 0
            
            candidate_results.append({
                'id': candidate.id,
                'vote_count': vote_count,
                'voter_percentage': voter_percentage
            })
        
        # Get total eligible voters for overall stats
        if election.department_id:
            total_eligible_voters = Student.query.filter_by(department_id=election.department_id).count()
        else:
            total_eligible_voters = Student.query.count()
        
        if election.scope == 'campus' and election.year_levels and election.year_levels != 'all':
            allowed_years = election.year_levels.split(',')
            total_eligible_voters = Student.query.filter(Student.year_level_id.in_(allowed_years)).count()
        
        voter_turnout = round((unique_voters / total_eligible_voters * 100), 2) if total_eligible_voters > 0 else 0
        students_not_voted = max(0, total_eligible_voters - unique_voters)
        
        total_new_votes = sum([c['new_votes'] for c in candidates_with_new_votes])
        
        return jsonify({
            'success': True,
            'total_voters': unique_voters,
            'voter_turnout': voter_turnout,
            'students_not_voted': students_not_voted,
            'candidate_results': candidate_results,
            'candidates_with_new_votes': candidates_with_new_votes,
            'total_new_votes': total_new_votes,
            'position_voter_counts': dict(position_voter_counts)  # Send actual voter counts
        })
        
    except Exception as e:
        print(f"Error checking new votes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
from admin.models import PdfResult

# Configure upload folder
UPLOAD_FOLDER = 'admin/static/uploads/pdf_results'
ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin_bp.route('/results/<int:election_id>/upload-pdf', methods=['POST'])
@admin_required
def upload_pdf_result(election_id):
    """Upload a PDF result file for an election"""
    try:
        election = Election.query.get_or_404(election_id)
        
        if 'pdf_file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['pdf_file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Only PDF files are allowed'}), 400
        
        # Secure filename and create unique name
        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"election_{election_id}_{timestamp}_{original_filename}"
        
        # Ensure upload directory exists
        upload_dir = os.path.join(current_app.root_path, 'static/uploads/pdf_results')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Get description from form
        description = request.form.get('description', '')
        
        # Create database record
        pdf_result = PdfResult(
            election_id=election_id,
            filename=original_filename,
            file_path=f'uploads/pdf_results/{filename}',
            file_size=file_size,
            uploaded_by=current_user.id,
            description=description
        )
        
        db.session.add(pdf_result)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (UPLOAD - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        log_audit(
            action='PDF_UPLOAD',
            description=f"Admin '{username}' from IP: {ip} uploaded PDF result '{original_filename}' for election '{election.title}' (ID: {election_id}) | Size: {file_size} bytes"
        )
        
        return jsonify({
            'success': True,
            'message': 'PDF uploaded successfully',
            'pdf_id': pdf_result.id
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error uploading PDF: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/pdf-result/<int:pdf_id>/delete', methods=['POST'])
@admin_required
def delete_pdf_result(pdf_id):
    """Delete a PDF result"""
    try:
        pdf_result = PdfResult.query.get_or_404(pdf_id)
        election_title = pdf_result.election.title if pdf_result.election else 'Unknown'
        
        # Delete file from filesystem
        file_path = os.path.join(current_app.root_path, 'static', pdf_result.file_path)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Delete database record
        db.session.delete(pdf_result)
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (DELETE - data modification)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        log_audit(
            action='PDF_DELETE',
            description=f"Admin '{username}' from IP: {ip} deleted PDF result '{pdf_result.filename}' for election '{election_title}' (ID: {pdf_result.election_id})"
        )
        
        return jsonify({'success': True, 'message': 'PDF deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting PDF: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/pdf-result/<int:pdf_id>/download')
@admin_required
def download_pdf_result(pdf_id):
    """Download a PDF result"""
    try:
        pdf_result = PdfResult.query.get_or_404(pdf_id)
        file_path = os.path.join(current_app.root_path, 'static', pdf_result.file_path)
        
        if not os.path.exists(file_path):
            flash('PDF file not found on server.', 'error')
            return redirect(request.referrer or url_for('admin.results_page'))
        
        # 🚫 REMOVED: PDF_DOWNLOAD audit log (not a data modification)
        # Downloading doesn't change database state
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=pdf_result.filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Error downloading PDF: {str(e)}")
        flash('Error downloading PDF file.', 'error')
        return redirect(request.referrer or url_for('admin.results_page'))



@admin_bp.route('/results/<int:election_id>/tally', methods=['POST'])
@admin_required
def tally_election_results(election_id):
    if not TALLY_VOTE_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'Tally system not available. TallyVote model is missing.'
        }), 500
    
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)
    
    election = Election.query.get_or_404(election_id)
    
    if election.end_date.tzinfo is None:
        election_end = tz.localize(election.end_date)
    else:
        election_end = election.end_date
    
    force_tally = request.json.get('force', False) if request.is_json else False
    if not force_tally and election_end > now:
        return jsonify({
            'success': False,
            'message': f'Election is active until {election_end.strftime("%b %d, %Y %I:%M %p")}.',
            'suggestion': 'Add "force": true to tally early.'
        }), 400
    
    try:
        candidates = Candidate.query.filter_by(election_id=election_id).all()
        if not candidates:
            return jsonify({
                'success': False,
                'message': 'No candidates found for this election.'
            }), 400
        
        candidate_results = []
        total_votes_counted = 0
        candidate_vote_map = {}
        
        for candidate in candidates:
            vote_count = count_votes_for_candidate(candidate.id, election_id)
            total_votes_counted += vote_count
            candidate_vote_map[candidate.id] = vote_count
            
            candidate_results.append({
                'candidate_id': candidate.id,
                'name': f"{candidate.first_name} {candidate.last_name}",
                'position': candidate.position.name if candidate.position else "N/A",
                'vote_count': vote_count
            })
        
        # Count unique voters
        unique_voters = db.session.query(Vote.student_id).filter_by(
            election_id=election_id
        ).distinct().count()
        
        total_votes_in_db = Vote.query.filter_by(election_id=election_id).count()
        
        tally_timestamp = datetime.utcnow()
        
        # ===== ONLY SAVE TO TallyVote TABLE (official results) =====
        TallyVote.query.filter_by(election_id=election_id).delete()
        
        for candidate_id, vote_count in candidate_vote_map.items():
            tally = TallyVote(
                election_id=election_id,
                candidate_id=candidate_id,
                vote_count=vote_count,
                tally_timestamp=tally_timestamp
            )
            db.session.add(tally)
        
        # ===== COMMIT CHANGES =====
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        force_text = " (forced)" if force_tally else ""
        
        log_audit(
            action='TALLY_ELECTION',
            description=f"Admin user '{username}' tallied results for election: '{election.title}' (ID: {election_id}) from IP: {ip}{force_text} | Candidates: {len(candidates)}, Unique voters: {unique_voters}, Total votes: {total_votes_counted}"
        )
        
        return jsonify({
            'success': True,
            'message': f'Tally completed: {unique_voters} voters, {total_votes_counted} votes for {len(candidates)} candidates.',
            'data': {
                'unique_voters': unique_voters,
                'total_votes': total_votes_counted,
                'votes_in_db': total_votes_in_db,
                'candidates_tallied': len(candidates),
                'tally_timestamp': tally_timestamp.isoformat(),
                'election_title': election.title,
                'results': candidate_results
            }
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        
        # ✅ KEEP THIS AUDIT LOG
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='TALLY_ELECTION_FAILED',
            description=f"Admin user '{username}' failed to tally results for election: '{election.title}' (ID: {election_id}) from IP: {ip} | Error: {str(e)[:200]}"
        )
        
        return jsonify({
            'success': False,
            'message': f'Tally failed: {str(e)}'
        }), 500


@admin_bp.route('/results/<int:election_id>/tally', methods=['DELETE'])
@admin_required
def clear_tally_results(election_id):
    if not TALLY_VOTE_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'Tally system not available.'
        }), 500
    
    election = Election.query.get_or_404(election_id)
    
    try:
        count = TallyVote.query.filter_by(election_id=election_id).count()
        TallyVote.query.filter_by(election_id=election_id).delete()
        db.session.commit()
        
        # ✅ KEEP THIS AUDIT LOG (CLEAR TALLY - data deletion)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='CLEAR_TALLY',
            description=f"Admin user '{username}' cleared tally results for election: '{election.title}' (ID: {election_id}) from IP: {ip} | Records deleted: {count}"
        )
        
        return jsonify({
            'success': True,
            'message': f'Removed {count} tally records for {election.title}.',
            'data': {
                'records_deleted': count,
                'election_id': election_id,
                'election_title': election.title
            }
        })
        
    except Exception as e:
        db.session.rollback()
        
        # ✅ KEEP THIS AUDIT LOG (CLEAR TALLY FAILED - still important)
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='CLEAR_TALLY_FAILED',
            description=f"Admin user '{username}' failed to clear tally results for election: '{election.title}' (ID: {election_id}) from IP: {ip} | Error: {str(e)[:200]}"
        )
        
        return jsonify({
            'success': False,
            'message': f'Failed to clear tally: {str(e)}'
        }), 500


@admin_bp.route('/results/<int:election_id>/tally')
@admin_required
def get_tally_results(election_id):
    if not TALLY_VOTE_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'Tally system not available.'
        }), 500
    
    election = Election.query.get_or_404(election_id)
    
    tally_records = TallyVote.query.filter_by(election_id=election_id).all()
    if not tally_records:
        # 🚫 REMOVED: GET_TALLY_NO_RESULTS audit log (GET request - not a data modification)
        
        return jsonify({
            'success': False,
            'message': 'No official tally results found.',
            'suggestion': 'Use the Tally Votes button to create official results.'
        }), 404
    
    results = []
    total_votes = 0
    
    for tally in tally_records:
        candidate = Candidate.query.get(tally.candidate_id)
        if candidate:
            results.append({
                'candidate_id': candidate.id,
                'name': f"{candidate.first_name} {candidate.last_name}",
                'position': candidate.position.name if candidate.position else "N/A",
                'vote_count': tally.vote_count,
                'tally_timestamp': tally.tally_timestamp.isoformat()
            })
            total_votes += tally.vote_count
    
    last_tally = max(tally_records, key=lambda x: x.tally_timestamp)
    
    # 🚫 REMOVED: GET_TALLY_RESULTS audit log (GET request - not a data modification)
    
    return jsonify({
        'success': True,
        'data': {
            'election_id': election.id,
            'election_title': election.title,
            'total_candidates_tallied': len(results),
            'total_votes_tallied': total_votes,
            'last_tally_timestamp': last_tally.tally_timestamp.isoformat(),
            'results': sorted(results, key=lambda x: x['vote_count'], reverse=True)
        }
    })



@admin_bp.route('/results/<int:election_id>/pdf')
@admin_required
def election_results_pdf(election_id):
    """Generate PDF using ReportLab - Footer in page footer area on last page only"""
    from datetime import datetime as dt
    
    election = Election.query.get_or_404(election_id)
    
    # Get chairman name
    chairman_name = request.args.get('chairman', '').strip()
    if not chairman_name:
        chairman_name = "COMELEC CHAIRMAN"
    
    tz = pytz.timezone('Asia/Manila')
    now = datetime.now(tz)
    
    # Get election status
    election_start = election.start_date
    election_end = election.end_date
    if election_start.tzinfo is None:
        election_start = tz.localize(election_start)
    if election_end.tzinfo is None:
        election_end = tz.localize(election_end)
    
    if election_end < now:
        status = "Completed"
    elif election_start <= now <= election_end:
        status = "Active"
    else:
        status = "Upcoming"
    
    # Check if tallied
    is_tallied = False
    try:
        from admin.models import TallyVote
        tally_record = TallyVote.query.filter_by(election_id=election_id).first()
        is_tallied = tally_record is not None
    except:
        is_tallied = False
    
    if not is_tallied:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'PDF results are only available after official tally.'
            }), 403
        flash('PDF results are only available after official tally.', 'warning')
        return redirect(url_for('admin.election_results', election_id=election_id))
    
    # Get all data
    all_votes = Vote.query.filter_by(election_id=election_id).all()
    unique_voters = len(set(vote.student_id for vote in all_votes))
    candidates = Candidate.query.filter_by(election_id=election_id).all()
    
    # Position limits
    position_limits = {}
    election_positions = ElectionPosition.query.filter_by(election_id=election_id).all()
    for ep in election_positions:
        position_limits[ep.position_id] = ep.max_votes
    
    # Get tally data
    from admin.models import TallyVote
    tally_records = TallyVote.query.filter_by(election_id=election_id).all()
    tally_dict = {t.candidate_id: t.vote_count for t in tally_records}
    
    # Helper function to get candidate display name
    def get_candidate_display_name(candidate):
        if candidate.candidate_type == 'studio':
            return candidate.studio_name or "Unknown Studio"
        else:
            return f"{candidate.first_name or ''} {candidate.last_name or ''}".strip() or "Unknown Candidate"
    
    # ===== HELPER FUNCTION: Get vote timeline for a candidate (for tie-breaking) =====
    def get_candidate_vote_timeline(candidate_id, election_id):
        """Get all timestamps when votes were cast for a candidate"""
        all_votes = Vote.query.filter_by(election_id=election_id).all()
        timestamps = []
        
        for vote in all_votes:
            candidate_ids = vote.voted_candidate_ids
            if candidate_id in candidate_ids:
                # Use cast_timestamp (actual voting time) as primary
                if vote.cast_timestamp:
                    timestamps.append(vote.cast_timestamp)
                elif vote.created_at:
                    timestamps.append(vote.created_at)
        
        timestamps.sort()
        return timestamps
    
    # ===== CALCULATE POSITION-SPECIFIC VOTER COUNTS (for percentages) =====
    position_voter_counts = {}
    candidate_position_map = {c.id: c.position_id for c in candidates}
    
    for ep in election_positions:
        position_id = ep.position_id
        voters_for_position = set()
        
        for vote in all_votes:
            if not vote.finder_hash:
                continue
                
            try:
                import json
                finder_data = json.loads(vote.finder_hash)
                candidate_ids = []
                
                if isinstance(finder_data, dict):
                    if 'hashes' in finder_data and isinstance(finder_data['hashes'], list):
                        for item in finder_data['hashes']:
                            if isinstance(item, dict) and 'candidate_id' in item:
                                candidate_ids.append(item['candidate_id'])
                    elif 'candidate_ids' in finder_data:
                        candidate_ids = finder_data['candidate_ids']
                elif isinstance(finder_data, list):
                    for item in finder_data:
                        if isinstance(item, dict) and 'candidate_id' in item:
                            candidate_ids.append(item['candidate_id'])
                
                for cid in candidate_ids:
                    cand_position_id = candidate_position_map.get(cid)
                    if cand_position_id == position_id:
                        if vote.student_id:
                            voters_for_position.add(vote.student_id)
                        else:
                            voters_for_position.add(f"anon_{vote.id}")
                        break
                        
            except (json.JSONDecodeError, Exception):
                continue
        
        position_voter_counts[position_id] = len(voters_for_position)
        print(f"PDF - Position {ep.position_id}: Voters = {len(voters_for_position)}")
    
    # Build candidate results
    candidate_results = []
    candidates_by_position = {}
    
    for candidate in candidates:
        position_name = candidate.position.name if candidate.position else "Unknown"
        vote_count = tally_dict.get(candidate.id, 0)
        display_name = get_candidate_display_name(candidate)
        
        # Get position-specific voter count for percentage calculation
        pos_voter_count = position_voter_counts.get(candidate.position_id, 1)
        percentage = round((vote_count / pos_voter_count * 100), 2) if pos_voter_count > 0 else 0
        
        candidate_data = {
            'id': candidate.id,
            'first_name': candidate.first_name,
            'last_name': candidate.last_name,
            'studio_name': candidate.studio_name,
            'candidate_type': candidate.candidate_type or "student",
            'display_name': display_name,
            'position': position_name,
            'position_id': candidate.position_id,
            'vote_count': vote_count,
            'percentage': percentage,
            'is_winner': False,
            'tie_broken': False,
            'reached_at': None
        }
        candidate_results.append(candidate_data)
        
        if position_name not in candidates_by_position:
            candidates_by_position[position_name] = {
                'id': candidate.position_id,
                'max_votes': position_limits.get(candidate.position_id, 1),
                'voter_count': pos_voter_count,
                'candidates': []
            }
        candidates_by_position[position_name]['candidates'].append(candidate_data)
    
    # ===== DETERMINE WINNERS WITH TIE-BREAKING (SAME AS RESULTS PAGE) =====
    winners_by_position = {}
    for position_name, pos_data in candidates_by_position.items():
        # Sort candidates by vote count (descending)
        pos_data['candidates'].sort(key=lambda x: x['vote_count'], reverse=True)
        max_winners = pos_data['max_votes']
        
        # Check for ties at the winner threshold
        tie_vote_count = None
        if len(pos_data['candidates']) > max_winners:
            last_winner_index = max_winners - 1
            last_winner_votes = pos_data['candidates'][last_winner_index]['vote_count']
            next_candidate_votes = pos_data['candidates'][max_winners]['vote_count']
            
            if last_winner_votes == next_candidate_votes and last_winner_votes > 0:
                tie_vote_count = last_winner_votes
                print(f"PDF - TIE DETECTED for {position_name} at {tie_vote_count} votes!")
        
        if tie_vote_count:
            # Get all candidates tied for the last winner position
            tied_candidates = [c for c in pos_data['candidates'] 
                              if c['vote_count'] == tie_vote_count]
            
            # Get non-tied winners (those with more votes)
            non_tied_winners = [c for c in pos_data['candidates'] 
                               if c['vote_count'] > tie_vote_count]
            
            # For each tied candidate, find when they reached the tie_vote_count
            for candidate in tied_candidates:
                timestamps = get_candidate_vote_timeline(candidate['id'], election_id)
                if len(timestamps) >= tie_vote_count:
                    candidate['reached_at'] = timestamps[tie_vote_count - 1]
                elif timestamps:
                    candidate['reached_at'] = timestamps[-1]
                else:
                    candidate['reached_at'] = None
                
                candidate_display = candidate['display_name']
                print(f"PDF - {candidate_display}: Reached {tie_vote_count} votes at {candidate['reached_at']}")
            
            # Sort tied candidates by who reached first (earliest timestamp wins)
            tied_candidates.sort(key=lambda x: x.get('reached_at') or dt.max)
            
            # Determine how many winners needed from tied group
            winners_needed = max_winners - len(non_tied_winners)
            
            # Mark all candidates as winner or loser
            for candidate in pos_data['candidates']:
                if candidate['vote_count'] > tie_vote_count:
                    candidate['is_winner'] = True
                    candidate['tie_broken'] = False
                elif candidate['vote_count'] == tie_vote_count:
                    # Check if this candidate is among the earliest to reach
                    if candidate in tied_candidates[:winners_needed]:
                        candidate['is_winner'] = True
                        candidate['tie_broken'] = True
                    else:
                        candidate['is_winner'] = False
                        candidate['tie_broken'] = True
                else:
                    candidate['is_winner'] = False
                    candidate['tie_broken'] = False
            
            # Collect winners for display
            winners = [c for c in pos_data['candidates'] if c.get('is_winner')]
            if winners:
                winners_by_position[position_name] = winners
        else:
            # No tie, normal winner selection
            winners = []
            for i, candidate in enumerate(pos_data['candidates']):
                if i < max_winners and candidate['vote_count'] > 0:
                    candidate['is_winner'] = True
                    winners.append(candidate)
                else:
                    candidate['is_winner'] = False
            if winners:
                winners_by_position[position_name] = winners
    
    # ===== SORT WINNERS TABLE BY POSITION ID =====
    # Get all position IDs for sorting
    position_id_map = {}
    for position_name in winners_by_position.keys():
        # Find position ID from candidates_by_position
        for pos_name, pos_data in candidates_by_position.items():
            if pos_name == position_name:
                position_id_map[position_name] = pos_data['id']
                break
    
    # Sort winners_by_position by position ID
    sorted_winners = sorted(winners_by_position.items(), key=lambda x: position_id_map.get(x[0], 9999))
    winners_by_position = dict(sorted_winners)
    
    # Calculate overall statistics
    if election.department_id:
        total_eligible_voters = Student.query.filter_by(department_id=election.department_id).count()
    else:
        total_eligible_voters = Student.query.count()
    
    voter_turnout = round((unique_voters / total_eligible_voters * 100), 2) if total_eligible_voters > 0 else 0
    students_not_voted = total_eligible_voters - unique_voters
    
    # ===== REPORTLAB PDF GENERATION WITH FOOTER IN PAGE MARGIN =====
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import Image as RLImage
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from flask import make_response
    import os
    from flask import current_app
    
    AVAILABLE_WIDTH = 7.5 * inch
    
    buffer = BytesIO()
    
    # Create custom canvas with footer on last page only
    class FooterCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            self.chairman_name = kwargs.pop('chairman_name', 'COMELEC CHAIRMAN')
            canvas.Canvas.__init__(self, *args, **kwargs)
            self.page_count = 0
            self.footer_drawn = False
        
        def showPage(self):
            self.page_count += 1
            canvas.Canvas.showPage(self)
        
        def draw_footer(self):
            """Draw footer on the last page only - in the page footer area"""
            if self.footer_drawn:
                return
            
            page_width = letter[0]
            # Position at bottom of page
            footer_y = 0.6 * inch
            
            # Draw chairman name
            self.setFont('Helvetica-Bold', 11)
            self.drawCentredString(page_width / 2, footer_y + 0.25 * inch, self.chairman_name)
            
            # Draw signature line (short length, not full width)
            self.setFont('Helvetica', 8)
            line_length = 2.5 * inch
            line_x = (page_width - line_length) / 2
            self.line(line_x, footer_y + 0.17 * inch, line_x + line_length, footer_y + 0.17 * inch)
            
            # Draw title - IMPROVED SPACING (moved down from line)
            self.setFont('Helvetica-Oblique', 9)
            self.drawCentredString(page_width / 2, footer_y + 0.06 * inch, "COMELEC Chairperson")
            
            # Draw footer logo - SAME WIDTH AS HEADER, 0.5 INCH HEIGHT
            footer_logo_path = os.path.join(current_app.root_path, 'admin', 'static', 'images', 'CTU FOOTER.png')
            if os.path.exists(footer_logo_path):
                try:
                    from reportlab.lib.utils import ImageReader
                    img = ImageReader(footer_logo_path)
                    # Width matches header (7.5 inches), height is 0.5 inches
                    img_width = AVAILABLE_WIDTH  # 7.5 inches
                    img_height = 0.5 * inch      # 0.5 inches tall
                    x = (page_width - img_width) / 2
                    y = footer_y - img_height - 0.05 * inch
                    # Set preserveAspectRatio=False to force exact dimensions
                    self.drawImage(img, x, y, width=img_width, height=img_height, preserveAspectRatio=False)
                except:
                    pass
            
            self.footer_drawn = True
        
        def save(self):
            # Draw footer on the last page before saving
            if self.page_count > 0:
                self.draw_footer()
            canvas.Canvas.save(self)
    
    # Create document with standard margins
    doc = SimpleDocTemplate(buffer,
                           pagesize=letter,
                           topMargin=0.6*inch,
                           bottomMargin=0.8*inch,
                           leftMargin=0.5*inch,
                           rightMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    story = []
    
    # ===== CUSTOM STYLES =====
    # Adjusted font sizes - smaller
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                  fontSize=14, alignment=TA_CENTER, spaceAfter=5,
                                  fontName='Helvetica-Bold')
    
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                     fontSize=12, alignment=TA_CENTER, spaceAfter=5,
                                     fontName='Helvetica-Bold')
    
    date_style = ParagraphStyle('Date', parent=styles['Normal'],
                                 fontSize=10, alignment=TA_CENTER, spaceAfter=15)
    
    # ===== HEADER LOGO =====
    logo_path = os.path.join(current_app.root_path, 'admin', 'static', 'images', 'CTU HEADER.png')
    if os.path.exists(logo_path):
        try:
            # Add negative space before logo to pull it up into the margin
            story.append(Spacer(1, -0.2*inch))
            img = RLImage(logo_path, width=AVAILABLE_WIDTH, height=1.0*inch)
            story.append(img)
            story.append(Spacer(1, 0.05*inch))
        except:
            pass
    
    story.append(Paragraph("Cebu Technological University - Moalboal Campus", title_style))
    story.append(Paragraph(f"{election.title} ELECTION", subtitle_style))
    story.append(Paragraph(now.strftime('%B %d, %Y'), date_style))
    story.append(Spacer(1, 0.1*inch))
    
    # ===== SUMMARY CARDS =====
    summary_data = [
        ['TOTAL VOTERS', 'TURNOUT RATE', 'CANDIDATES', 'STATUS'],
        [str(unique_voters), f"{voter_turnout}%", str(len(candidate_results)), status]
    ]
    
    col_width = AVAILABLE_WIDTH / 4
    summary_table = Table(summary_data, colWidths=[col_width] * 4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e0e0')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,1), 14),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,1), (0,1), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0,1), (0,1), colors.HexColor('#01579b')),
        ('BACKGROUND', (1,1), (1,1), colors.HexColor('#e8f5e9')),
        ('TEXTCOLOR', (1,1), (1,1), colors.HexColor('#1b5e20')),
        ('BACKGROUND', (2,1), (2,1), colors.HexColor('#f3e5f5')),
        ('TEXTCOLOR', (2,1), (2,1), colors.HexColor('#4a148c')),
        ('BACKGROUND', (3,1), (3,1), colors.HexColor('#fff3e0')),
        ('TEXTCOLOR', (3,1), (3,1), colors.HexColor('#bf360c')),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.1*inch))
    
    # ===== TURNOUT STATS =====
    turnout_data = [
        ['REGISTERED', 'VOTED', 'ABSENT'],
        [str(total_eligible_voters), str(unique_voters), str(students_not_voted)]
    ]
    
    col_width = AVAILABLE_WIDTH / 3
    turnout_table = Table(turnout_data, colWidths=[col_width] * 3)
    turnout_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e0e0')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,1), 14),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,1), (0,1), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0,1), (0,1), colors.HexColor('#01579b')),
        ('BACKGROUND', (1,1), (1,1), colors.HexColor('#e8f5e9')),
        ('TEXTCOLOR', (1,1), (1,1), colors.HexColor('#1b5e20')),
        ('BACKGROUND', (2,1), (2,1), colors.HexColor('#fff3e0')),
        ('TEXTCOLOR', (2,1), (2,1), colors.HexColor('#bf360c')),
    ]))
    story.append(turnout_table)
    story.append(Spacer(1, 0.1*inch))
    
    # ===== ENCRYPTION NOTE =====
    note_style = ParagraphStyle('Note', parent=styles['Normal'],
                                 fontSize=8, alignment=TA_CENTER,
                                 textColor=colors.HexColor('#666666'),
                                 fontName='Helvetica-Oblique')
    story.append(Paragraph("Privacy Protected Results: These results are calculated using Paillier Homomorphic Encryption. Individual votes remain private while ensuring accurate totals.", note_style))
    story.append(Spacer(1, 0.15*inch))
    
    # ===== WINNERS SECTION =====
    winners_title_style = ParagraphStyle('WinnersTitle', parent=styles['Heading2'],
                                         fontSize=14, alignment=TA_CENTER,
                                         textColor=colors.HexColor('#1b5e20'),
                                         backColor=colors.HexColor('#d0ebd0'),
                                         spaceAfter=8, spaceBefore=8,
                                         borderPadding=8, fontName='Helvetica-Bold')
    story.append(Paragraph("OFFICIAL ELECTION WINNERS", winners_title_style))
    story.append(Spacer(1, 0.05*inch))
    
    # Winners table - SORTED BY POSITION ID (already sorted above)
    winners_data = [['Position', 'Winner(s)', 'Votes']]
    for position_name, winners in winners_by_position.items():
        if len(winners) == 1:
            winner = winners[0]
            winner_display = winner['display_name']
            winners_data.append([position_name, winner_display, str(winner['vote_count'])])
        else:
            winners_data.append([position_name, f"{len(winners)} WINNERS", ''])
            for winner in winners:
                winner_display = winner['display_name']
                winners_data.append(['', f"  • {winner_display}", str(winner['vote_count'])])
    
    winners_table = Table(winners_data, colWidths=[AVAILABLE_WIDTH * 0.35, AVAILABLE_WIDTH * 0.45, AVAILABLE_WIDTH * 0.2])
    winners_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1b5e20')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN', (2,0), (2,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(winners_table)
    story.append(PageBreak())
    
    # ===== DETAILED RESULTS SECTION =====
    detail_title_style = ParagraphStyle('DetailTitle', parent=styles['Heading2'],
                                         fontSize=14, alignment=TA_CENTER,
                                         textColor=colors.HexColor('#01579b'),
                                         backColor=colors.HexColor('#b8d9f5'),
                                         spaceAfter=8, spaceBefore=8,
                                         borderPadding=8, fontName='Helvetica-Bold')
    story.append(Paragraph("DETAILED RESULTS BY POSITION", detail_title_style))
    story.append(Spacer(1, 0.05*inch))
    
    # Group by position
    position_groups = {}
    for candidate in candidate_results:
        pos_name = candidate['position']
        pos_id = candidate['position_id']
        if pos_name not in position_groups:
            position_groups[pos_name] = {'id': pos_id, 'candidates': []}
        position_groups[pos_name]['candidates'].append(candidate)
    
    # Sort positions by ID
    sorted_positions = sorted(position_groups.items(), key=lambda x: x[1]['id'])
    
    for position_name, group in sorted_positions:
        candidates_list = group['candidates']
        # Sort by vote count (descending) - for display, winners already determined
        candidates_list.sort(key=lambda x: x['vote_count'], reverse=True)
        max_winners = position_limits.get(group['id'], 1)
        
        pos_title_style = ParagraphStyle('PosTitle', parent=styles['Heading3'],
                                         fontSize=11, spaceAfter=6, 
                                         textColor=colors.HexColor('#000000'),
                                         fontName='Helvetica-Bold')
        story.append(Paragraph(f"{position_name}{' (' + str(max_winners) + ' WINNERS)' if max_winners > 1 else ''}", pos_title_style))
        
        # UPDATED: Show display_name (studio_name or full name) and percentage
        detail_data = [['Rank', 'Candidate', 'Votes', 'Percentage', 'Result']]
        for idx, candidate in enumerate(candidates_list[:max_winners+5], 1):
            # Use the pre-determined winner status from tie-breaking logic
            is_winner = candidate.get('is_winner', False)
            result = "✓ WINNER" if is_winner and candidate['vote_count'] > 0 else "—"
            detail_data.append([str(idx),
                               candidate['display_name'],
                               str(candidate['vote_count']),
                               f"{candidate['percentage']}%",
                               result])
        
        detail_table = Table(detail_data, colWidths=[AVAILABLE_WIDTH * 0.10, AVAILABLE_WIDTH * 0.43, AVAILABLE_WIDTH * 0.12, AVAILABLE_WIDTH * 0.12, AVAILABLE_WIDTH * 0.23])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e0e0')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (2,0), (4,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('TEXTCOLOR', (4,1), (4,max_winners), colors.HexColor('#1b5e20')),
            ('FONTNAME', (4,1), (4,max_winners), 'Helvetica-Bold'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(detail_table)
        story.append(Spacer(1, 0.12*inch))
    
    # Build PDF with custom canvas for footer on last page only
    doc.build(story, canvasmaker=lambda filename, pagesize=letter, **kwargs: FooterCanvas(filename, pagesize, chairman_name=chairman_name, **kwargs))
    buffer.seek(0)
    
    # Return PDF
    filename = f"{election.title}_Official_Results_{now.strftime('%Y%m%d_%H%M')}.pdf"
    filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@admin_bp.route('/results/<int:election_id>/test-tally')
@admin_required
def test_tally_calculation(election_id):
    election = Election.query.get_or_404(election_id)
    candidates = Candidate.query.filter_by(election_id=election_id).all()
    
    if not candidates:
        return jsonify({
            'success': False,
            'message': 'No candidates found'
        }), 400
    
    results = []
    total_votes = 0
    
    for candidate in candidates:
        vote_count = count_votes_for_candidate(candidate.id, election_id)
        total_votes += vote_count
        
        results.append({
            'candidate_id': candidate.id,
            'name': f"{candidate.first_name} {candidate.last_name}",
            'position': candidate.position.name if candidate.position else "N/A",
            'vote_count': vote_count
        })
    
    results.sort(key=lambda x: x['vote_count'], reverse=True)
    
    winners_by_position = {}
    for result in results:
        position = result['position']
        if position not in winners_by_position:
            winners_by_position[position] = result
        elif result['vote_count'] > winners_by_position[position]['vote_count']:
            winners_by_position[position] = result
    
    return jsonify({
        'success': True,
        'data': {
            'election': {
                'id': election.id,
                'title': election.title,
                'status': election.status
            },
            'total_votes_calculated': total_votes,
            'candidates_count': len(results),
            'winners': list(winners_by_position.values()),
            'results': results,
            'note': 'This is a test calculation. Results are NOT saved.'
        }
    })


@admin_bp.route('/profile')
@login_required
def admin_profile():
    # Check if user is admin using role
    if current_user.role != 'Admin':
        return "Unauthorized", 403

    # Get the admin data
    admin = Admin.query.get(current_user.id)

    # Simple stats (replace with actual queries later)
    total_elections = 12
    total_votes = 340

    # 🚫 REMOVED: Profile page view audit log (not a data modification)

    return render_template('admin_profile.html', 
                           admin=admin,
                           total_elections=total_elections,
                           total_votes=total_votes)



@admin_bp.route('/debug-candidate-order/<int:election_id>')
@admin_required
def debug_candidate_order(election_id):
    """Debug endpoint to see candidate order in votes"""
    import json
    
    output = []
    output.append("=" * 60)
    output.append(f"DEBUG FOR ELECTION ID: {election_id}")
    output.append("=" * 60)
    
    # Get current candidates sorted by ID
    current_candidates = Candidate.query.filter_by(election_id=election_id)\
                                        .order_by(Candidate.id).all()
    
    output.append("\n📋 CURRENT CANDIDATES (Sorted by ID - used for tallying):")
    output.append("-" * 50)
    for idx, c in enumerate(current_candidates):
        output.append(f"  Index {idx}: ID {c.id} - {c.first_name} {c.last_name} ({c.position.name if c.position else 'N/A'})")
    
    # Get all votes for this election
    votes = Vote.query.filter_by(election_id=election_id).all()
    output.append(f"\n📊 TOTAL VOTES: {len(votes)}")
    
    if not votes:
        output.append("❌ No votes found!")
        return "<br>".join(output)
    
    # Check first 3 votes to see stored candidate_order
    output.append("\n🔍 CHECKING FIRST 3 VOTES:")
    output.append("-" * 50)
    
    for i, vote in enumerate(votes[:3]):
        output.append(f"\n--- VOTE #{i+1} (ID: {vote.id}) ---")
        
        if not vote.finder_hash:
            output.append("  ❌ No finder_hash found!")
            continue
            
        try:
            finder_data = json.loads(vote.finder_hash)
            output.append(f"  📦 finder_data keys: {list(finder_data.keys())}")
            
            # Check if candidate_order exists
            if 'candidate_order' in finder_data:
                stored_order = finder_data['candidate_order']
                output.append(f"  ✅ candidate_order FOUND! Length: {len(stored_order)}")
                output.append(f"  📍 Stored order: {stored_order}")
                
                # Compare with current order
                current_ids = [c.id for c in current_candidates]
                if stored_order == current_ids:
                    output.append("  ✅ PERFECT MATCH! Orders are identical.")
                else:
                    output.append("  ⚠️ MISMATCH DETECTED!")
                    for idx, (stored, current) in enumerate(zip(stored_order, current_ids)):
                        if stored != current:
                            output.append(f"     Index {idx}: Stored ID {stored} vs Current ID {current}")
            else:
                output.append("  ❌ NO candidate_order in this vote!")
                output.append(f"  📦 Available keys: {list(finder_data.keys())}")
                
                # Show what IS stored
                if 'hashes' in finder_data:
                    output.append(f"  📍 Hashes count: {len(finder_data['hashes'])}")
                    for h in finder_data['hashes'][:3]:
                        output.append(f"     - Candidate ID: {h.get('candidate_id')}")
                        
        except json.JSONDecodeError as e:
            output.append(f"  ❌ JSON decode error: {e}")
        except Exception as e:
            output.append(f"  ❌ Error: {e}")
    
    # Summary
    output.append("\n" + "=" * 60)
    output.append("SUMMARY")
    output.append("=" * 60)
    
    # Count how many votes have candidate_order
    votes_with_order = 0
    for vote in votes:
        if vote.finder_hash:
            try:
                finder_data = json.loads(vote.finder_hash)
                if 'candidate_order' in finder_data:
                    votes_with_order += 1
            except:
                pass
    
    output.append(f"✅ Votes with candidate_order: {votes_with_order} / {len(votes)}")
    
    if votes_with_order == 0:
        output.append("\n❌ PROBLEM: No votes have candidate_order stored!")
        output.append("   This is why tallying is miscounting votes.")
        output.append("\n🔧 SOLUTION: You need to:")
        output.append("   1. Update submit_vote to include 'candidate_order'")
        output.append("   2. Or manually fix existing votes (use fix script below)")
    
    return "<br>".join(output)


from admin.models import VoteDistribution
from datetime import datetime
import pytz

@admin_bp.route('/vote-distribution')
@admin_required
def vote_distribution():
    """Main vote distribution page with year filtering"""
    
    # ✅ GET YEAR FILTER FROM SESSION (set by dashboard)
    year = session.get('admin_current_year')
    
    # Parse year to date range (January 1 to December 31)
    start_date = None
    end_date = None
    if year:
        try:
            year_int = int(year)
            start_date = datetime(year_int, 1, 1)
            end_date = datetime(year_int, 12, 31, 23, 59, 59)
        except (ValueError, TypeError):
            start_date = None
            end_date = None
    
    # Get elections filtered by year
    election_query = Election.query
    
    if start_date and end_date:
        election_query = election_query.filter(
            Election.start_date >= start_date,
            Election.start_date <= end_date
        )
    
    elections = election_query.order_by(Election.created_at.desc()).all()
    
    if not year and elections:

        latest_election = Election.query.order_by(Election.start_date.desc()).first()
        if latest_election and latest_election.start_date:
            year = latest_election.start_date.year
            session['admin_current_year'] = year
    
    all_elections = Election.query.order_by(Election.start_date.asc()).all()
    available_years = set()
    for election in all_elections:
        if election.start_date:
            available_years.add(election.start_date.year)
    available_years = sorted(list(available_years), reverse=True)
    
    return render_template('vote_distribution.html', 
                          elections=elections, 
                          available_years=available_years,  
                          current_year=year)  


@admin_bp.route('/api/vote-distribution/<int:election_id>')
@admin_required
def get_vote_distribution(election_id):
    """API endpoint to get vote distribution data from vote_distributions table"""
    
    election = Election.query.get_or_404(election_id)
    
    # Check if distribution data exists
    distribution_exists = VoteDistribution.query.filter_by(election_id=election_id).first()
    if not distribution_exists:
        return jsonify({
            'success': False,
            'message': 'No distribution data found. Please populate first.',
            'suggestion': 'Use the populate endpoint to generate data.'
        }), 404
    
    # Get all candidates for this election with position ordering
    candidates = Candidate.query.filter_by(election_id=election_id)\
        .join(Position)\
        .order_by(Position.id, Candidate.last_name)\
        .all()
    
    # Determine grouping type from first record (assuming consistent)
    first_record = VoteDistribution.query.filter_by(election_id=election_id).first()
    grouping_type = first_record.grouping_type if first_record else 'department'
    
    # Build distribution data FROM vote_distributions TABLE
    distribution_data = []
    positions_with_ids = {}  # Store position names with their IDs for sorting
    
    for candidate in candidates:
        # Store position ID for sorting
        if candidate.position:
            positions_with_ids[candidate.position.name] = candidate.position.id
        
        # Get distribution records for this candidate
        dist_records = VoteDistribution.query.filter_by(
            election_id=election_id,
            candidate_id=candidate.id
        ).all()
        
        total_votes = sum(d.vote_count for d in dist_records)
        
        # Get photo URL
        photo_url = None
        if candidate.photo:
            photo_url = url_for('admin.static', filename='images/' + candidate.photo)
        
        # ===== UPDATED: Determine display name and initials based on candidate type =====
        if candidate.candidate_type == 'studio':
            display_name = candidate.studio_name or "Unknown Studio"
            initials = candidate.studio_name[0].upper() if candidate.studio_name else "S"
        else:
            display_name = f"{candidate.first_name or ''} {candidate.last_name or ''}".strip() or "Unknown Candidate"
            initials = f"{candidate.first_name[0] if candidate.first_name else ''}{candidate.last_name[0] if candidate.last_name else ''}".upper() or "??"
        
        candidate_data = {
            'id': candidate.id,
            'name': display_name,
            'initials': initials,
            'position': candidate.position.name if candidate.position else "Unknown",
            'position_id': candidate.position_id,
            'total_votes': total_votes,
            'photo': photo_url,
            'breakdown': []
        }
        
        # Add breakdown based on what's available with formatted percentages
        for dist in dist_records:
            name = None
            item_id = None
            
            if dist.department_id:
                dept = Department.query.get(dist.department_id)
                name = dept.name if dept else None
                item_id = dist.department_id
            elif dist.course_id:
                course = Course.query.get(dist.course_id)
                name = course.course_name if course else None
                item_id = dist.course_id
            elif dist.program_type_id:
                from student.models import ProgramType
                prog = ProgramType.query.get(dist.program_type_id)
                name = prog.name if prog else None
                item_id = dist.program_type_id
            else:
                name = dist.grouping_name
                item_id = None
            
            if name:
                # Format percentage to 2 decimal places
                percentage_formatted = f"{dist.percentage:.2f}" if dist.percentage else "0.00"
                
                candidate_data['breakdown'].append({
                    'id': item_id,
                    'name': name,
                    'votes': dist.vote_count,
                    'percentage': float(percentage_formatted),  # Send as float for consistency
                    'percentage_display': f"{percentage_formatted}%"  # Display format
                })
        
        # Sort by vote count (highest first)
        candidate_data['breakdown'].sort(key=lambda x: x['votes'], reverse=True)
        
        distribution_data.append(candidate_data)
    
    # Calculate stats - count unique groups based on grouping type
    if grouping_type == 'department':
        unique_groups = db.session.query(VoteDistribution.department_id)\
            .filter_by(election_id=election_id)\
            .filter(VoteDistribution.department_id.isnot(None))\
            .distinct().count()
        group_label = 'Departments with Votes'
    elif grouping_type == 'course':
        unique_groups = db.session.query(VoteDistribution.course_id)\
            .filter_by(election_id=election_id)\
            .filter(VoteDistribution.course_id.isnot(None))\
            .distinct().count()
        group_label = 'Courses with Votes'
    elif grouping_type == 'program_type':
        unique_groups = db.session.query(VoteDistribution.program_type_id)\
            .filter_by(election_id=election_id)\
            .filter(VoteDistribution.program_type_id.isnot(None))\
            .distinct().count()
        group_label = 'Program Types (Day/Night)'
    else:
        unique_groups = 0
        group_label = 'Groups with Votes'
    
    total_votes = sum(d.vote_count for d in VoteDistribution.query.filter_by(election_id=election_id).all())
    total_candidates = len(candidates)
    
    # Calculate voter turnout
    total_students = Student.query.count()
    voted_students = db.session.query(Vote.student_id)\
        .filter_by(election_id=election_id)\
        .distinct().count()
    turnout = round((voted_students / total_students * 100), 1) if total_students > 0 else 0
    
    # Group by position - use position ID for sorting
    positions = {}
    for candidate in distribution_data:
        pos_name = candidate['position']
        if pos_name not in positions:
            positions[pos_name] = []
        positions[pos_name].append(candidate)
    
    # Sort candidates by votes within each position
    for pos in positions:
        positions[pos].sort(key=lambda x: x['total_votes'], reverse=True)
    
    # Create sorted positions list based on position IDs
    sorted_position_names = sorted(positions.keys(), key=lambda x: positions_with_ids.get(x, 999))
    
    return jsonify({
        'success': True,
        'election': {
            'id': election.id,
            'title': election.title,
            'scope': election.scope
        },
        'stats': {
            'total_candidates': total_candidates,
            'total_departments': unique_groups,
            'total_votes': total_votes,
            'voter_turnout': turnout,
            'group_label': group_label
        },
        'positions': positions,
        'positions_order': sorted_position_names,  # Add positions order
        'all_candidates': distribution_data,
        'grouping_type': grouping_type
    })



@admin_bp.route('/api/vote-distribution/populate/<int:election_id>', methods=['POST'])
@admin_required
def populate_vote_distribution(election_id):
    """Manually populate vote_distributions table from votes"""
    import json
    from collections import defaultdict
    
    try:
        election = Election.query.get_or_404(election_id)
        
        # Clear existing distribution data for this election
        VoteDistribution.query.filter_by(election_id=election_id).delete()
        
        # Get all votes for this election
        votes = Vote.query.filter_by(election_id=election_id).all()
        
        if not votes:
            return jsonify({'success': False, 'message': 'No votes found'}), 404
        
        # Determine grouping strategy
        grouping_type = determine_grouping_strategy(election)
        print(f"📊 Using grouping strategy: {grouping_type}")
        
        # Count votes by candidate and grouping entity
        vote_counts = defaultdict(int)
        counted_combinations = set()
        
        for vote in votes:
            if not vote.finder_hash:
                print(f"⚠️ Vote {vote.id} has no finder_hash")
                continue
                
            try:
                finder_data = json.loads(vote.finder_hash)
                
                # Get student's details
                student = Student.query.get(vote.student_id)
                if not student:
                    continue
                
                # Get grouping ID based on strategy
                grouping_id = get_grouping_id(student, grouping_type)
                if not grouping_id:
                    continue
                
                # Get grouping name for display
                grouping_name = get_grouping_name(grouping_id, grouping_type)
                
                # Extract candidate IDs from finder_hash
                candidate_ids = []
                
                if isinstance(finder_data, dict):
                    if 'hashes' in finder_data and isinstance(finder_data['hashes'], list):
                        for item in finder_data['hashes']:
                            if isinstance(item, dict) and 'candidate_id' in item:
                                candidate_ids.append(item['candidate_id'])
                    elif 'candidate_ids' in finder_data and isinstance(finder_data['candidate_ids'], list):
                        candidate_ids = finder_data['candidate_ids']
                elif isinstance(finder_data, list):
                    for item in finder_data:
                        if isinstance(item, dict) and 'candidate_id' in item:
                            candidate_ids.append(item['candidate_id'])
                
                # Count the votes
                for candidate_id in candidate_ids:
                    vote_candidate_key = (vote.id, candidate_id)
                    
                    if vote_candidate_key not in counted_combinations:
                        counted_combinations.add(vote_candidate_key)
                        vote_counts[(candidate_id, grouping_id, grouping_name)] += 1
                            
            except json.JSONDecodeError as e:
                print(f"❌ JSON decode error for vote {vote.id}: {e}")
                continue
            except Exception as e:
                print(f"❌ Error processing vote {vote.id}: {e}")
                continue
        
        if not vote_counts:
            return try_alternative_extraction(election_id, election)
        
        # Insert into VoteDistribution table
        records_created = 0
        for (candidate_id, grouping_id, grouping_name), count in vote_counts.items():
            candidate = Candidate.query.get(candidate_id)
            if candidate:
                dist_kwargs = {
                    'election_id': election_id,
                    'candidate_id': candidate_id,
                    'vote_count': count,
                    'position_id': candidate.position_id,
                    'position_name': candidate.position.name if candidate.position else None,
                    'grouping_type': grouping_type,
                    'grouping_name': grouping_name
                }
                
                if grouping_type == 'department':
                    dist_kwargs['department_id'] = grouping_id
                elif grouping_type == 'course':
                    dist_kwargs['course_id'] = grouping_id
                elif grouping_type == 'program_type':
                    dist_kwargs['program_type_id'] = grouping_id
                
                dist = VoteDistribution(**dist_kwargs)
                db.session.add(dist)
                records_created += 1
        
        db.session.commit()
        
        # Calculate percentages
        if records_created > 0:
            VoteDistribution.calculate_percentages(election_id)
        
        # Audit log
        username = getattr(current_user, 'username', 'Unknown')
        ip = request.remote_addr
        
        log_audit(
            action='POPULATE_VOTE_DISTRIBUTION',
            description=f"Admin user '{username}' populated vote distribution for election '{election.title}' (ID: {election_id}) from IP: {ip} | Strategy: {grouping_type} | Records created: {records_created}"
        )
        
        return jsonify({
            'success': True,
            'message': f'Populated {records_created} distribution records using {grouping_type} grouping',
            'records_created': records_created,
            'strategy': grouping_type
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error in populate_vote_distribution: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({'success': False, 'error': str(e)}), 500


def try_alternative_extraction(election_id, election):
    """Alternative method to extract votes if the first method fails"""
    print("🔄 Trying alternative extraction method...")
    
    from sqlalchemy import text
    
    connection = db.session.connection()
    result = connection.execute(
        text("SELECT id, finder_hash FROM votes WHERE election_id = :eid LIMIT 5"),
        {"eid": election_id}
    )
    
    samples = []
    for row in result:
        samples.append({"id": row[0], "hash": row[1]})
    
    print(f"📊 Sample votes: {samples}")
    
    return jsonify({
        'success': False,
        'message': 'No distribution records could be created. Please check the console for debug info.',
        'samples': samples
    }), 400


def determine_grouping_strategy(election):
    """Determine how to group vote distribution"""
    if election.scope == 'department':
        if election.department_id:
            courses_count = Course.query.filter_by(department_id=election.department_id).count()
            if courses_count > 1:
                return 'course'
            else:
                return 'program_type'
    return 'department'


def get_grouping_id(student, grouping_type):
    """Get the appropriate grouping ID based on strategy"""
    if grouping_type == 'department':
        return student.department_id
    elif grouping_type == 'course':
        return student.course_id
    elif grouping_type == 'program_type':
        return student.program_type_id
    return None


def get_grouping_name(grouping_id, grouping_type):
    """Get the display name for a grouping ID"""
    if not grouping_id:
        return None
    
    if grouping_type == 'department':
        dept = Department.query.get(grouping_id)
        return dept.name if dept else None
    elif grouping_type == 'course':
        course = Course.query.get(grouping_id)
        return course.course_name if course else None
    elif grouping_type == 'program_type':
        from student.models import ProgramType
        prog_type = ProgramType.query.get(grouping_id)
        return prog_type.name if prog_type else None
    
    return None




@admin_bp.route("/audit-logs", methods=["GET"])
def audit_logs():
    import pytz  # Add this import
    
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    export = request.args.get("export")

    query = AuditLog.query

    # 🔎 SEARCH FILTER
    if search:
        query = query.filter(
            or_(
                AuditLog.action.ilike(f"%{search}%"),
                AuditLog.role.ilike(f"%{search}%"),
                AuditLog.description.ilike(f"%{search}%")
            )
        )

    # 📅 DATE FILTER (FIXED PROPERLY)
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(AuditLog.timestamp >= start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            # Add 1 day to include full end date
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.timestamp <= end_date_obj)
        except ValueError:
            pass

    # ORDER BY LATEST FIRST
    query = query.order_by(desc(AuditLog.timestamp))

    # 📥 EXPORT CSV
    if export == "csv":
        logs = query.all()

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["ID", "User ID", "Role", "Action", "Description", "IP Address", "Timestamp (Manila Time)"])

        tz = pytz.timezone('Asia/Manila')
        
        for log in logs:
            # Convert UTC to Manila time for CSV export
            if log.timestamp:
                manila_time = log.timestamp.replace(tzinfo=pytz.UTC).astimezone(tz)
                timestamp_str = manila_time.strftime("%Y-%m-%d %H:%M:%S")
            else:
                timestamp_str = ""
                
            writer.writerow([
                log.id,
                log.user_id,
                log.role,
                log.action,
                log.description,
                log.ip_address,
                timestamp_str
            ])

        output.seek(0)

        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="audit_logs.csv"
        )

    # 📄 PAGINATION
    logs = query.paginate(page=page, per_page=20)
    
    # Pass pytz to template
    return render_template(
        "audit_logs.html",
        logs=logs,
        search=search,
        start_date=start_date,
        end_date=end_date,
        pytz=pytz  # Add this
    )


@admin_bp.route("/audit-logs-ajax", methods=["GET"])
def audit_logs_ajax():
    """AJAX endpoint for audit logs pagination"""
    import pytz
    
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = AuditLog.query

    # 🔎 SEARCH FILTER
    if search:
        query = query.filter(
            or_(
                AuditLog.action.ilike(f"%{search}%"),
                AuditLog.role.ilike(f"%{search}%"),
                AuditLog.description.ilike(f"%{search}%")
            )
        )

    # 📅 DATE FILTER
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(AuditLog.timestamp >= start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.timestamp <= end_date_obj)
        except ValueError:
            pass

    # ORDER BY LATEST FIRST
    query = query.order_by(desc(AuditLog.timestamp))

    # 📄 PAGINATION
    logs = query.paginate(page=page, per_page=20)
    
    # Return just the table partial
    return render_template(
        "partials/audit_logs_table.html",
        logs=logs,
        search=search,
        start_date=start_date,
        end_date=end_date,
        pytz=pytz
    )



    
# ------------------- Modified Logout Route ------------------- #
@admin_bp.route('/logout')
@login_required
def logout():
    """Logout admin and clear all sessions"""
    username = current_user.username
    logout_user()
    
    # Clear ALL admin-related sessions
    session.pop('access_granted', None)
    session.pop('access_time', None)
    session.pop('pre_2fa_admin_id', None)
    
    # Get current secret path
    secret_path = AccessCode.get_secret_path()
    
    log_audit(
        action='LOGOUT',
        description=f"Admin user '{username}' logged out"
    )
    
    
    # Redirect to the dynamic access path
    return redirect(url_for('admin.dynamic_access', secret_path=secret_path))